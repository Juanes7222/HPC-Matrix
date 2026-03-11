import os
import sys
from typing import cast

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    Series,
    make_border, set_col_width, style_data_cell, style_header_cell,
    write_title_row, save_figure, plot_lines,
)

RESULTS_DIR = "results"
CHARTS_DIR  = os.path.join(RESULTS_DIR, "charts")
OUTPUT_FILE = os.path.join(RESULTS_DIR, "reporte_hpc.xlsx")
REPETITIONS = 10

MODE_ORDER = [
    ("sequential",  "Sequential"),
    ("threads_2t",  "2 Threads"),
    ("threads_4t",  "4 Threads"),
    ("threads_6t",  "6 Threads"),
    ("threads_8t",  "8 Threads"),
    ("threads_12t", "12 Threads"),
    ("concurrent",  "Processes"),
]

COLORS = [
    "#1F4E79", "#2E75B6", "#70AD47", "#ED7D31",
    "#FFC000", "#C00000", "#7030A0",
]

def load_csv(mode_key: str) -> pd.DataFrame | None:
    path = os.path.join(RESULTS_DIR, f"data_{mode_key}.csv")
    if not os.path.exists(path):
        return None

    df = pd.read_csv(path)
    df.columns = [c.strip().lower() for c in df.columns]
    df = df.rename(columns={
        "size": "matrix_size", "n": "matrix_size",
        "dim": "matrix_size",  "dimension": "matrix_size",
        "rep": "repetition",   "repeticion": "repetition",
        "time_ms": "wall_time_ms", "elapsed_ms": "wall_time_ms",
        "wall_time": "wall_time_ms", "time": "wall_time_ms",
    })

    required = {"matrix_size", "repetition", "wall_time_ms"}
    missing  = required - set(df.columns)
    if missing:
        print(f"  [WARN] {path}: missing columns {missing}. Skipping.")
        return None

    if "exit_code" in df.columns:
        bad = (df["exit_code"] != 0).sum()
        if bad:
            print(f"  [WARN] {path}: dropping {bad} failed row(s) (exit_code != 0)")
        df = df[df["exit_code"] == 0].copy()

    df["matrix_size"]  = df["matrix_size"].astype(int)
    df["repetition"]   = df["repetition"].astype(int)
    df["wall_time_ms"] = df["wall_time_ms"].astype(float) / 1000.0
    return df


def collect_sizes(available_modes: list) -> list[int]:
    sizes: set[int] = set()
    for mode_key, _ in available_modes:
        df = load_csv(mode_key)
        if df is not None:
            sizes.update(int(s) for s in df["matrix_size"].unique())
    return sorted(sizes)


def build_averages(available_modes: list,
                   matrix_sizes: list[int]) -> dict[str, dict[int, float]]:
    avgs: dict[str, dict[int, float]] = {}
    for mode_key, _ in available_modes:
        df         = load_csv(mode_key)
        avgs[mode_key] = {}
        if df is None:
            continue
        for size in matrix_sizes:
            subset = df[df["matrix_size"] == size]["wall_time_ms"]
            if not subset.empty:
                avgs[mode_key][size] = float(subset.mean())
    return avgs

def chart_execution_time(available_modes: list, avgs: dict,
                         matrix_sizes: list[int]) -> str:
    series = [
        Series(label=label, data=avgs.get(key, {}),
               color=COLORS[i % len(COLORS)])
        for i, (key, label) in enumerate(available_modes)
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 6))
        plot_lines(ax, [s for s in series if s.data])
        ax.set_title("Execution Time vs Matrix Dimension",
                     fontsize=14, fontweight="bold", pad=12)
        ax.set_xlabel("Matrix dimension N (NxN)", fontsize=11)
        ax.set_ylabel("Average wall time (s)", fontsize=11)
        ax.legend(fontsize=9, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "chart_execution_time.png")


def chart_speedup(available_modes: list, avgs: dict,
                  matrix_sizes: list[int]) -> str | None:
    seq_avgs = avgs.get("sequential", {})
    if not seq_avgs:
        print("  [WARN] No sequential data found; skipping speedup chart.")
        return None

    parallel = [
        (key, label, i)
        for i, (key, label) in enumerate(available_modes)
        if key != "sequential"
    ]
    series = [
        Series(
            label=label,
            data={s: seq_avgs[s] / avgs[key][s]
                  for s in matrix_sizes
                  if s in avgs.get(key, {}) and s in seq_avgs and avgs[key][s] > 0},
            color=COLORS[i % len(COLORS)],
        )
        for key, label, i in parallel
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.axhline(y=1, color="#AAAAAA", lw=1, ls="--", label="Baseline (seq = 1)")
        plot_lines(ax, [s for s in series if s.data])
        ax.set_title("Speedup vs Matrix Dimension",
                     fontsize=14, fontweight="bold", pad=12)
        ax.set_xlabel("Matrix dimension N (NxN)", fontsize=11)
        ax.set_ylabel("Speedup  (T_seq / T_parallel)", fontsize=11)
        ax.legend(fontsize=9, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "chart_speedup.png")

def chart_per_mode(mode_key: str, label: str,
                   df: pd.DataFrame | None,
                   matrix_sizes: list[int]) -> str | None:
    if df is None or df.empty:
        return None

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        x_pos    = list(range(len(matrix_sizes)))
        avg_vals = []

        for xi, size in enumerate(matrix_sizes):
            subset = df[df["matrix_size"] == size]["wall_time_ms"]
            if subset.empty:
                avg_vals.append(None)
                continue
            ax.scatter([xi] * len(subset), subset.to_numpy(),
                       color="#2E75B6", alpha=0.45, s=35, zorder=3)
            avg_vals.append(float(subset.mean()))

        valid_x = [x_pos[i] for i, v in enumerate(avg_vals) if v is not None]
        valid_y = [v for v in avg_vals if v is not None]
        ax.plot(valid_x, valid_y, color="#1F4E79", lw=2,
                marker="D", markersize=6, label="Average", zorder=4)

        ax.set_xticks(x_pos)
        ax.set_xticklabels([str(s) for s in matrix_sizes], fontsize=9)
        ax.set_title(f"{label}  —  Wall time per repetition",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Matrix dimension N", fontsize=10)
        ax.set_ylabel("Wall time (s)", fontsize=10)
        ax.legend(fontsize=9)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, f"chart_{mode_key}.png")

def write_mode_sheet(wb: Workbook, mode_key: str, label: str,
                     seq_avgs: dict, matrix_sizes: list[int],
                     chart_path: str | None = None) -> None:
    df     = load_csv(mode_key)
    ws     = wb.create_sheet(title=label)
    ws.freeze_panes = "B3"
    is_seq = (mode_key == "sequential")
    n_cols = REPETITIONS + 3

    write_title_row(ws, f"Matrix Multiplication  |  {label}", n_cols, row=1)
    ws.row_dimensions[1].height = 28

    ws.row_dimensions[2].height = 22
    headers = (["Dimension (N)"]
               + [f"Rep {i}" for i in range(1, REPETITIONS + 1)]
               + ["Average (s)", "Speedup"])
    for col, h in enumerate(headers, 1):
        style_header_cell(ws.cell(2, col), h,
                          bg="mid" if col > 1 else "dark")

    set_col_width(ws, 1, 16)
    for col in range(2, REPETITIONS + 2):
        set_col_width(ws, col, 13)
    set_col_width(ws, REPETITIONS + 2, 15)
    set_col_width(ws, REPETITIONS + 3, 12)

    first_rep = get_column_letter(2)
    last_rep  = get_column_letter(REPETITIONS + 1)
    avg_col   = get_column_letter(REPETITIONS + 2)

    for row_idx, size in enumerate(matrix_sizes, 3):
        alt = (row_idx % 2 == 0)

        dim = ws.cell(row_idx, 1, value=f"{size} x {size}")
        dim.font      = Font(name=FONT_NAME, bold=True, size=10)
        dim.fill      = PatternFill("solid", fgColor=C["light"])
        dim.alignment = Alignment(horizontal="center")
        dim.border    = make_border()

        for rep in range(1, REPETITIONS + 1):
            c = ws.cell(row_idx, rep + 1)
            if df is not None:
                mask = (df["matrix_size"] == size) & (df["repetition"] == rep)
                vals = cast(pd.Series, df.loc[mask, "wall_time_ms"])
                if not vals.empty:
                    c.value = round(float(vals.iloc[0]), 3)
            c.number_format = "0.000"
            c.font          = Font(name=FONT_NAME, size=10)
            c.fill          = PatternFill("solid",
                                          fgColor=C["alt"] if alt else C["white"])
            c.alignment     = Alignment(horizontal="right")
            c.border        = make_border()

        avg_c = ws.cell(row_idx, REPETITIONS + 2)
        avg_c.value = (
            f"=IFERROR(AVERAGE({first_rep}{row_idx}:{last_rep}{row_idx}),\"N/A\")"
        )
        avg_c.number_format = "0.000"
        avg_c.font      = Font(name=FONT_NAME, bold=True, size=10)
        avg_c.fill      = PatternFill("solid", fgColor=C["light"])
        avg_c.alignment = Alignment(horizontal="right")
        avg_c.border    = make_border()

        sp_c = ws.cell(row_idx, REPETITIONS + 3)
        if is_seq:
            sp_c.value = 1.0
        else:
            seq_avg = seq_avgs.get(size)
            if df is not None and seq_avg is not None:
                subset = df[df["matrix_size"] == size]["wall_time_ms"]
                sp_c.value = (round(seq_avg / float(subset.mean()), 4)
                              if not subset.empty and subset.mean() > 0
                              else "N/A")
            else:
                sp_c.value = "N/A"

        sp_c.number_format = "0.0000"
        sp_c.font      = Font(name=FONT_NAME, bold=True, size=10, color=C["green_fg"])
        sp_c.fill      = PatternFill("solid", fgColor=C["green_bg"])
        sp_c.alignment = Alignment(horizontal="right")
        sp_c.border    = make_border()

    if chart_path and os.path.exists(chart_path):
        img        = XLImage(chart_path)
        img.width  = 620
        img.height = 340
        ws.add_image(img, f"{get_column_letter(n_cols + 2)}2")

def write_summary_sheet(wb: Workbook, available_modes: list,
                        avgs: dict, seq_avgs: dict,
                        matrix_sizes: list[int],
                        time_chart: str | None,
                        speedup_chart: str | None) -> None:
    ws = wb.create_sheet(title="Summary")
    ws.freeze_panes = "B3"

    n_modes  = len(available_modes)
    last_col = get_column_letter(1 + n_modes * 2)
    write_title_row(ws, "Comparative Summary  |  HPC Matrix Multiplication",
                    1 + n_modes * 2, row=1)
    ws.row_dimensions[1].height = 28

    style_header_cell(ws.cell(2, 1), "Dimension (N)")
    set_col_width(ws, 1, 16)
    ws.row_dimensions[2].height = 32

    for m_idx, (_, label) in enumerate(available_modes):
        avg_col = 2 + m_idx * 2
        sp_col  = avg_col + 1
        style_header_cell(ws.cell(2, avg_col), f"{label}\nAverage (s)",
                          bg="mid", size=9)
        style_header_cell(ws.cell(2, sp_col), f"{label}\nSpeedup",
                          bg="mid", size=9)
        set_col_width(ws, avg_col, 14)
        set_col_width(ws, sp_col,  12)

    for row_idx, size in enumerate(matrix_sizes, 3):
        alt = (row_idx % 2 == 0)

        dim = ws.cell(row_idx, 1, value=f"{size} x {size}")
        dim.font      = Font(name=FONT_NAME, bold=True, size=10)
        dim.fill      = PatternFill("solid", fgColor=C["light"])
        dim.alignment = Alignment(horizontal="center")
        dim.border    = make_border()

        for m_idx, (mode_key, _) in enumerate(available_modes):
            avg_col  = 2 + m_idx * 2
            sp_col   = avg_col + 1
            mode_avg = avgs.get(mode_key, {}).get(size)
            seq_avg  = seq_avgs.get(size)

            avg_c = ws.cell(row_idx, avg_col)
            avg_c.value = round(mode_avg, 3) if mode_avg is not None else "N/D"
            avg_c.number_format = "0.000"
            avg_c.font      = Font(name=FONT_NAME, size=10)
            avg_c.fill      = PatternFill("solid",
                                          fgColor=C["alt"] if alt else C["white"])
            avg_c.alignment = Alignment(horizontal="right")
            avg_c.border    = make_border()

            sp_c = ws.cell(row_idx, sp_col)
            if mode_key == "sequential":
                sp_c.value = 1.0
            elif mode_avg and mode_avg > 0 and seq_avg:
                sp_c.value = round(seq_avg / mode_avg, 4)
            else:
                sp_c.value = "N/A"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            sp_c.fill      = PatternFill("solid",
                                         fgColor=C["green_bg"] if not alt
                                         else C["green_alt"])
            sp_c.alignment = Alignment(horizontal="right")
            sp_c.border    = make_border()

    chart_row = len(matrix_sizes) + 5
    for anchor, path in [("A", time_chart), ("M", speedup_chart)]:
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = 720
            img.height = 430
            ws.add_image(img, f"{anchor}{chart_row}")


def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    available_modes = []
    for mode_key, label in MODE_ORDER:
        path = os.path.join(RESULTS_DIR, f"data_{mode_key}.csv")
        if os.path.exists(path):
            available_modes.append((mode_key, label))
            print(f"  [OK]  data_{mode_key}.csv")
        else:
            print(f"  [--]  data_{mode_key}.csv  (not found, skipping)")

    if not available_modes:
        print("\nNo CSVs found in results/. Run the benchmark first.")
        sys.exit(1)

    matrix_sizes = collect_sizes(available_modes)
    print(f"\n  Sizes detected: {matrix_sizes}")

    avgs     = build_averages(available_modes, matrix_sizes)
    seq_avgs = avgs.get("sequential", {})
    if not seq_avgs:
        print("  [WARN] Sequential data not found. Speedup will show N/A.")

    print("\n  Generating charts...")
    per_mode_charts = {}
    for mode_key, label in available_modes:
        df   = load_csv(mode_key)
        path = chart_per_mode(mode_key, label, df, matrix_sizes)
        per_mode_charts[mode_key] = path
        if path:
            print(f"    {os.path.basename(path)}")

    time_chart    = chart_execution_time(available_modes, avgs, matrix_sizes)
    speedup_chart = chart_speedup(available_modes, avgs, matrix_sizes)
    for p in [time_chart, speedup_chart]:
        if p:
            print(f"    {os.path.basename(p)}")

    print("\n  Building Excel workbook...")
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    write_summary_sheet(wb, available_modes, avgs, seq_avgs, matrix_sizes,
                        time_chart, speedup_chart)

    for mode_key, label in available_modes:
        print(f"    Sheet: {label}")
        write_mode_sheet(wb, mode_key, label, seq_avgs, matrix_sizes,
                         chart_path=per_mode_charts.get(mode_key))

    wb.save(OUTPUT_FILE)
    print(f"\n  Excel report : {OUTPUT_FILE}")
    print(f"  Charts folder: {CHARTS_DIR}/")


if __name__ == "__main__":
    print("HPC Report Generator")
    print("=" * 40)
    main()