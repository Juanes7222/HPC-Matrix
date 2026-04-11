"""
report_utils.py  --  Shared utilities for HPC benchmark report generation.

Provides styling helpers, chart primitives, and generic sheet writers
used by report.py, report_cache.py, and report_opt.py.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

C: dict[str, str] = {
    "dark":     "1F4E79",
    "mid":      "2E75B6",
    "light":    "D6E4F0",
    "alt":      "F2F9FF",
    "white":    "FFFFFF",
    "green_bg": "E2EFDA",
    "green_fg": "375623",
    "green_alt":"EAF4E2",
    "red_bg":   "FCE4D6",
    "red_fg":   "843C0C",
    "gold_bg":  "FFF2CC",
    "gold_fg":  "7F6000",
    "grey_bg":  "D8D8D8",
    "grey_fg":  "404040",
    "border":   "BDD7EE",
}

FONT_NAME = "Arial"

CHART_STYLE: dict[str, Any] = {
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "axes.grid":         True,
    "grid.color":        "#E8E8E8",
    "grid.linewidth":    0.7,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "font.family":       "DejaVu Sans",
}

def make_border(color: str | None = None) -> Border:
    s = Side(style="thin", color=color or C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def style_header_cell(cell, value: str, bg: str = "dark", size: int = 10) -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=True, color=C["white"], size=size)
    cell.fill      = PatternFill("solid", fgColor=C.get(bg, bg))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = make_border()


def style_data_cell(cell, value: Any, fmt: str | None = None,
                    bg: str = "white", fg: str = "000000",
                    bold: bool = False, align: str = "right") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, color=C.get(fg, fg), size=10)
    cell.fill      = PatternFill("solid", fgColor=C.get(bg, bg))
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = make_border()
    if fmt:
        cell.number_format = fmt


def write_title_row(ws, text: str, n_cols: int, row: int = 1) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    cell            = ws.cell(row=row, column=1, value=text)
    cell.font       = Font(name=FONT_NAME, bold=True, size=13, color=C["white"])
    cell.fill       = PatternFill("solid", fgColor=C["dark"])
    cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

def save_figure(fig, charts_dir: str, name: str) -> str:
    path = os.path.join(charts_dir, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path

@dataclass
class Series:
    label: str
    data:  dict[int | str, float]
    color: str


def plot_lines(ax, series: list[Series], log_scale: bool = False) -> None:
    """Draw one line per Series onto ax."""
    for s in series:
        xs = sorted(s.data)
        ys = [s.data[x] for x in xs]
        ax.plot(xs, ys, marker="o", lw=2.5, color=s.color, label=s.label)

    ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"{int(x)}"))
    if log_scale:
        all_vals = [v for s in series for v in s.data.values() if v > 0]
        ax.set_yscale("log")
        if all_vals:
            ax.set_ylim(bottom=min(all_vals) * 0.4)
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda y, _: f"{y:,.3g}"))


def plot_grouped_bars(ax, groups: list[int | str],
                      series: list[Series],
                      bar_width: float = 0.35,
                      label_values: bool = False) -> None:
    """Draw grouped bars: one group per element in groups, one bar per Series."""
    n = len(series)
    x = list(range(len(groups)))
    for i, s in enumerate(series):
        vals    = [s.data.get(g, 0) for g in groups]
        offsets = [xi + (i - n / 2 + 0.5) * bar_width for xi in x]
        bars    = ax.bar(offsets, vals, width=bar_width * 0.9,
                         color=s.color, label=s.label,
                         edgecolor="white", alpha=0.85)
        if label_values:
            for bar, val in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height() * 1.02,
                        f"{val:,.0f}", ha="center", va="bottom",
                        fontsize=8, fontweight="bold")
    ax.set_xticks(x)


def plot_simple_bars(ax, labels: list[str], values: list[float],
                     color: str | list[str], label_fmt: str = "{:.2f}x") -> None:
    """Draw a single series of bars with value labels on top."""
    bars = ax.bar(range(len(labels)), values,
                  color=color, edgecolor="white")
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(values) * 0.01,
                label_fmt.format(val),
                ha="center", va="bottom", fontsize=8, fontweight="bold")
    ax.set_xticks(range(len(labels)))


def cv_per_group(df: pd.DataFrame, group_col: str,
                 size_col: str, time_col: str,
                 group_key: str, sizes: list[int]) -> list[float]:
    """Return CV (%) for group_key across sizes."""
    result = []
    for s in sizes:
        vals = df[(df[group_col] == group_key) & (df[size_col] == s)][time_col]
        result.append(vals.std() / vals.mean() * 100 if len(vals) > 1 else 0.0)
    return result

@dataclass
class ColConfig:
    header: str
    width:  float
    attr:   str
    fmt:    str | None = None
    align:  str = "right"


def write_raw_data_sheet(wb: Workbook, df: pd.DataFrame,
                         sheet_title: str, cols: list[ColConfig]) -> None:
    ws = wb.create_sheet("Raw Data")
    ws.freeze_panes = "A3"
    write_title_row(ws, sheet_title, len(cols))

    for ci, col in enumerate(cols, 1):
        style_header_cell(ws.cell(2, ci), col.header, bg="mid")
        set_col_width(ws, ci, col.width)
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(df.itertuples(index=False), 3):
        bg = "alt" if ri % 2 == 0 else "white"
        for ci, col in enumerate(cols, 1):
            style_data_cell(ws.cell(ri, ci), getattr(row, col.attr),
                            fmt=col.fmt, bg=bg, align=col.align)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df) + 2}"

def write_charts_sheet(wb: Workbook, title: str, chart_paths: list[str],
                       anchors: list[str] | None = None,
                       img_w: int = 750, img_h: int = 370,
                       n_title_cols: int = 12) -> None:
    ws = wb.create_sheet("Charts")
    write_title_row(ws, title, n_title_cols)

    default_anchors = ["A2", "M2", "A28", "M28", "A54", "M54"]
    used_anchors    = anchors or default_anchors

    for anchor, path in zip(used_anchors, chart_paths):
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = img_w
            img.height = img_h
            ws.add_image(img, anchor)

def compute_averages(df: pd.DataFrame,
                     group_col: str,
                     size_col: str,
                     time_col: str,
                     groups: list[str],
                     sizes: list[int]) -> dict[str, dict[int, float]]:
    """Return {group: {size: avg_time}} for all group/size combos."""
    avgs: dict[str, dict[int, float]] = {}
    for group in groups:
        sub        = df[df[group_col] == group]
        avgs[group] = {}
        for size in sizes:
            vals = sub[sub[size_col] == size][time_col]
            if not vals.empty:
                avgs[group][size] = float(vals.mean())
    return avgs