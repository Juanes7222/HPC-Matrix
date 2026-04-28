"""
report_geekbench.py  --  Geekbench 6 comparison report (web scraper edition)
==============================================================================

Lee la URL del resultado desde results_geekbench/<label>/result_url.txt,
hace scraping de browser.geekbench.com y genera:
  - Excel workbook  (reporte_geekbench.xlsx)
  - PNG charts      (charts_geekbench/)
  - LaTeX table     (geekbench_table.tex)

Usage:
    python3 tests/benchmarks/report_geekbench.py
    python3 tests/benchmarks/report_geekbench.py machine1 machine2

Requisitos:
    pip install requests beautifulsoup4 matplotlib openpyxl
"""

from __future__ import annotations

import os
import re
import sys
import time

import requests
from bs4 import BeautifulSoup

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths & constants
# ---------------------------------------------------------------------------
BASE_DIR   = "tests/benchmarks/results_geekbench"
LABELS     = sys.argv[1:3] if len(sys.argv) >= 3 else None
CHARTS_DIR = os.path.join(BASE_DIR, "charts_geekbench")
XLSX_PATH  = os.path.join(BASE_DIR, "reporte_geekbench.xlsx")
TEX_PATH   = os.path.join(BASE_DIR, "geekbench_table.tex")

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/135.0.0.0 Safari/537.36"
    )
}

MACHINE_COLORS = ["#2E75B6", "#C00000"]

# Win/loss colors used in workload charts
WIN_COLOR  = "#43A047"
LOSE_COLOR = "#E53935"
TIE_COLOR  = "#9E9E9E"

# Colors for the SC/MC parallel % difference chart
SC_COLOR = "#1E88E5"
MC_COLOR = "#FB8C00"

FONT_NAME = "Arial"
C = {
    "dark":     "1F4E79",
    "mid":      "2E75B6",
    "light":    "D6E4F0",
    "alt":      "F2F9FF",
    "white":    "FFFFFF",
    "green_bg": "E2EFDA",
    "green_fg": "375623",
    "red_bg":   "FCE4D6",
    "red_fg":   "C00000",
    "border":   "BDD7EE",
}

CHART_STYLE = {
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
    "axes.grid":         True,
    "grid.color":        "#E8E8E8",
    "axes.spines.top":   False,
    "axes.spines.right": False,
}


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
def make_border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def style_header_cell(cell, value, bg="dark", size=10):
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=True, color=C["white"], size=size)
    cell.fill      = PatternFill("solid", fgColor=C.get(bg, bg))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = make_border()


def style_data_cell(cell, value, fmt=None, bg="white", fg="000000",
                    bold=False, align="right"):
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, color=C.get(fg, fg), size=10)
    cell.fill      = PatternFill("solid", fgColor=C.get(bg, bg))
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = make_border()
    if fmt:
        cell.number_format = fmt


def write_title_row(ws, text: str, n_cols: int, row: int = 1):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    cell           = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(name=FONT_NAME, bold=True, size=13, color=C["white"])
    cell.fill      = PatternFill("solid", fgColor=C["dark"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def save_figure(fig, name: str) -> str:
    path = os.path.join(CHARTS_DIR, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path


# ---------------------------------------------------------------------------
# Discovery helpers
# ---------------------------------------------------------------------------
def discover_machines() -> list[str]:
    """Returns labels whose folder contains result_url.txt or geekbench_summary.txt."""
    if not os.path.isdir(BASE_DIR):
        return []
    return sorted(
        d for d in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, d))
        and (
            os.path.isfile(os.path.join(BASE_DIR, d, "result_url.txt"))
            or os.path.isfile(os.path.join(BASE_DIR, d, "geekbench_summary.txt"))
        )
    )


def get_result_url(label: str) -> str:
    """
    Returns the Geekbench result URL for the given label.
    Checks in order: result_url.txt, then geekbench_summary.txt via regex.
    """
    url_file = os.path.join(BASE_DIR, label, "result_url.txt")
    if os.path.isfile(url_file):
        url = open(url_file).read().strip()
        if url.startswith("http"):
            return url

    summary_file = os.path.join(BASE_DIR, label, "geekbench_summary.txt")
    if os.path.isfile(summary_file):
        text = open(summary_file).read()
        m = re.search(
            r"https://browser\.geekbench\.com/v6/cpu/(\d+)(?!/claim)",
            text
        )
        if m:
            return m.group(0)

    raise FileNotFoundError(
        f"[{label}] No se encontro result_url.txt ni URL valida en geekbench_summary.txt.\n"
        f"  Carpeta esperada: {os.path.join(BASE_DIR, label)}"
    )


def load_machine_info(label: str) -> str:
    path = os.path.join(BASE_DIR, label, "machine_info.txt")
    return open(path).read() if os.path.isfile(path) else "No machine info available."


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------
def _int(s) -> int:
    try:
        return int(str(s).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0


def scrape_geekbench(url: str, label: str) -> dict:
    """
    Extracts scores and metadata from browser.geekbench.com/v6/cpu/<id>.
    Uses curl_cffi to impersonate a real browser TLS fingerprint and
    bypass Cloudflare's bot detection.
    """
    from curl_cffi import requests as cffi_requests

    print(f"  [{label}] Descargando: {url}")
    resp = cffi_requests.get(url, impersonate="chrome136", timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # ... resto del parsing sin cambios

    single_core = 0
    multi_core  = 0
    for div in soup.find_all("div", class_="score-container"):
        note = div.find("div", class_="note")
        val  = div.find("div", class_="score")
        if not note or not val:
            continue
        note_text = note.get_text(strip=True)
        v         = _int(val.get_text(strip=True))
        if "Single" in note_text:
            single_core = v
        elif "Multi" in note_text:
            multi_core = v

    plat_div = soup.find("div", class_="platform-info")
    platform = plat_div.get_text(strip=True) if plat_div else ""

    system_info: dict[str, str] = {}
    cpu_info:    dict[str, str] = {}
    memory_info: dict[str, str] = {}

    for table in soup.find_all("table", class_="system-table"):
        header  = table.find("th")
        section = header.get_text(strip=True) if header else ""
        parsed: dict[str, str] = {}
        for row in table.find_all("tr"):
            name_td = row.find("td", class_="system-name") or \
                      row.find("td", class_=lambda c: c and "name" in c.split())
            val_td  = row.find("td", class_="system-value") or \
                      row.find("td", class_=lambda c: c and "value" in c.split())
            if name_td and val_td:
                parsed[name_td.get_text(strip=True)] = val_td.get_text(strip=True)
        if "CPU" in section:
            cpu_info = parsed
        elif "Memory" in section:
            memory_info = parsed
        elif "System" in section:
            system_info = parsed

    sc_map: dict[str, dict] = {}
    mc_map: dict[str, dict] = {}

    for table in soup.find_all("table", class_="benchmark-table"):
        th_name = table.find("th", class_="name")
        if not th_name:
            continue
        is_single = "Single" in th_name.get_text(strip=True)
        target    = sc_map if is_single else mc_map

        for row in table.find_all("tr"):
            name_td  = row.find("td", class_="name")
            score_td = row.find("td", class_="score")
            if not name_td or not score_td:
                continue
            name = name_td.get_text(" ", strip=True).split("\n")[0].strip()
            raw  = score_td.get_text("\n", strip=True).split("\n")
            target[name] = {
                "score": _int(raw[0]),
                "desc":  raw[1].strip() if len(raw) > 1 else "",
            }

    all_names = list(dict.fromkeys(list(sc_map) + list(mc_map)))
    workloads = []
    for name in all_names:
        sc = sc_map.get(name, {})
        mc = mc_map.get(name, {})
        workloads.append({
            "name":        name,
            "single_core": sc.get("score", 0),
            "single_desc": sc.get("desc",  ""),
            "multi_core":  mc.get("score", 0),
            "multi_desc":  mc.get("desc",  ""),
        })

    return {
        "label":       label,
        "url":         url,
        "single_core": single_core,
        "multi_core":  multi_core,
        "system":      system_info,
        "cpu":         cpu_info,
        "memory":      memory_info,
        "platform":    platform,
        "workloads":   workloads,
    }


# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------
def _pct_diff(v1: float, v2: float) -> float:
    """Percentage difference of v2 relative to v1. Returns 0.0 if v1 is zero."""
    return (v2 - v1) / v1 * 100 if v1 > 0 else 0.0


def _fmt_pct(value: float) -> str:
    return f"+{value:.1f}%" if value >= 0 else f"{value:.1f}%"


def _bar_value_labels(ax, bars, values: list, max_val: float, fontsize: int = 8):
    """Adds value labels above each bar."""
    for bar, val in zip(bars, values):
        if val == 0:
            continue
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + max_val * 0.010,
            f"{val:,}",
            ha="center", va="bottom",
            fontsize=fontsize, fontweight="bold", color="#333333",
        )


def _pct_diff_annotations(ax, bars_a, bars_b,
                           vals_a: list, vals_b: list,
                           max_val: float, fontsize: int = 8):
    """
    Draws a percentage-difference badge centered above each bar pair.
    Positive = b is better than a (green), negative = a is better (red).
    """
    for b1, b2, v1, v2 in zip(bars_a, bars_b, vals_a, vals_b):
        if v1 == 0 or v2 == 0:
            continue
        pct   = _pct_diff(v1, v2)
        x_mid = (b1.get_x() + b1.get_width() / 2 + b2.get_x() + b2.get_width() / 2) / 2
        y_top = max(b1.get_height(), b2.get_height()) + max_val * 0.075
        color = "#2E7D32" if pct >= 0 else "#B71C1C"
        ax.text(
            x_mid, y_top, _fmt_pct(pct),
            ha="center", va="bottom",
            fontsize=fontsize, fontweight="bold", color=color,
            bbox=dict(
                boxstyle="round,pad=0.28",
                facecolor="white",
                edgecolor=color,
                alpha=0.88,
                linewidth=0.8,
            ),
        )


def _workload_colors(machines: list[str], all_names: list[str], score_fn) -> list[list]:
    """
    Returns a list of color lists (one per machine) where each bar is colored
    WIN_COLOR if that machine wins the workload, LOSE_COLOR if it loses.
    """
    if len(machines) < 2:
        return [[MACHINE_COLORS[0]] * len(all_names)]

    colors_m1, colors_m2 = [], []
    for name in all_names:
        v1 = score_fn(machines[0], name)
        v2 = score_fn(machines[1], name)
        if v1 > v2:
            colors_m1.append(WIN_COLOR)
            colors_m2.append(LOSE_COLOR)
        elif v2 > v1:
            colors_m1.append(LOSE_COLOR)
            colors_m2.append(WIN_COLOR)
        else:
            colors_m1.append(TIE_COLOR)
            colors_m2.append(TIE_COLOR)
    return [colors_m1, colors_m2]


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------
def chart_overall(machines: list[str], data: dict) -> str:
    labels    = ["Single-Core Score", "Multi-Core Score"]
    bar_width = 0.35
    x_pos     = [0, 1]
    max_val   = max(data[m]["multi_core"] for m in machines) or 1

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5.5))
        bar_groups: list = []
        val_groups: list = []

        for mi, m in enumerate(machines):
            vals    = [data[m]["single_core"], data[m]["multi_core"]]
            offsets = [xi + (mi - len(machines) / 2 + 0.5) * bar_width for xi in x_pos]
            bars    = ax.bar(offsets, vals, width=bar_width * 0.9,
                             color=MACHINE_COLORS[mi], label=m,
                             edgecolor="white", alpha=0.88)
            _bar_value_labels(ax, bars, vals, max_val, fontsize=10)
            bar_groups.append(bars)
            val_groups.append(vals)

        if len(machines) > 1:
            _pct_diff_annotations(
                ax, bar_groups[0], bar_groups[1],
                val_groups[0], val_groups[1],
                max_val, fontsize=10,
            )

        ax.set_xticks(x_pos)
        ax.set_xticklabels(labels, fontsize=11)
        ax.set_ylabel("Geekbench 6 Score", fontsize=10)
        ax.set_ylim(0, max_val * 1.28)
        ax.set_title("Geekbench 6 — Overall Score Comparison",
                     fontsize=13, fontweight="bold", pad=14)
        ax.legend(fontsize=10)
        fig.tight_layout(pad=2.0)
        return save_figure(fig, "overall_scores.png")


def chart_workloads(machines: list[str], data: dict,
                    metric: str, title: str, filename: str) -> str:
    all_names = list(dict.fromkeys(
        w["name"] for m in machines for w in data[m]["workloads"]
    ))

    def get_score(m: str, name: str) -> int:
        for w in data[m]["workloads"]:
            if w["name"] == name:
                return w[metric]
        return 0

    bar_width   = 0.35
    x           = list(range(len(all_names)))
    score_lists = [[get_score(m, n) for n in all_names] for m in machines]
    max_val     = max((v for sl in score_lists for v in sl), default=1) or 1
    colors      = _workload_colors(machines, all_names, get_score)

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(max(13, len(all_names) * 1.35), 7.0))
        bar_groups = []

        for mi, m in enumerate(machines):
            offsets = [xi + (mi - len(machines) / 2 + 0.5) * bar_width for xi in x]
            bars    = ax.bar(offsets, score_lists[mi], width=bar_width * 0.9,
                             color=colors[mi], edgecolor="white", alpha=0.88)
            _bar_value_labels(ax, bars, score_lists[mi], max_val, fontsize=7)
            bar_groups.append(bars)

        if len(machines) > 1:
            _pct_diff_annotations(
                ax, bar_groups[0], bar_groups[1],
                score_lists[0], score_lists[1],
                max_val, fontsize=7,
            )
            legend_handles = [
                Patch(facecolor=MACHINE_COLORS[mi],
                      label=f"{machines[mi]}  ({'left' if mi == 0 else 'right'} bar)")
                for mi in range(len(machines))
            ] + [
                Patch(facecolor=WIN_COLOR,  label="Winner"),
                Patch(facecolor=LOSE_COLOR, label="Loser"),
            ]
        else:
            legend_handles = [Patch(facecolor=MACHINE_COLORS[0], label=machines[0])]

        ax.set_xticks(x)
        ax.set_xticklabels(all_names, rotation=38, ha="right", fontsize=8)
        ax.set_ylabel("Score", fontsize=10)
        ax.set_ylim(0, max_val * 1.32)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=12)
        ax.legend(handles=legend_handles, fontsize=9, loc="upper right")
        fig.tight_layout(pad=2.2)
        return save_figure(fig, filename)


def chart_speedup(machines: list[str], data: dict) -> str:
    """Bar chart showing % difference (m2 vs m1) per workload for multi-core."""
    if len(machines) < 2:
        return ""
    m1, m2 = machines[0], machines[1]

    sc1_map = {w["name"]: w["multi_core"] for w in data[m1]["workloads"]}
    sc2_map = {w["name"]: w["multi_core"] for w in data[m2]["workloads"]}

    names: list[str] = []
    pcts:  list[float] = []
    for name, v1 in sc1_map.items():
        v2 = sc2_map.get(name, 0)
        if v1 > 0 and v2 > 0:
            names.append(name)
            pcts.append(_pct_diff(v1, v2))

    bar_colors = [WIN_COLOR if p >= 0 else LOSE_COLOR for p in pcts]
    abs_max    = max(abs(p) for p in pcts) if pcts else 1.0

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(max(12, len(names) * 1.2), 5.5))
        bars = ax.bar(range(len(names)), pcts, color=bar_colors,
                      edgecolor="white", alpha=0.88)

        for bar, val in zip(bars, pcts):
            label  = _fmt_pct(val)
            color  = "#1B5E20" if val >= 0 else "#B71C1C"
            y_text = val + abs_max * 0.04 if val >= 0 else val - abs_max * 0.04
            va     = "bottom" if val >= 0 else "top"
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                y_text, label,
                ha="center", va=va,
                fontsize=8, fontweight="bold", color=color,
            )

        ax.axhline(0, color="#888888", lw=1.5, ls="--", zorder=0)
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=38, ha="right", fontsize=8)
        ax.set_ylabel(f"% difference  ({m2} vs {m1})", fontsize=10)
        ax.set_title(f"Multi-Core % Difference:  {m2}  vs  {m1}",
                     fontsize=12, fontweight="bold", pad=12)
        y_pad = abs_max * 0.18
        ax.set_ylim(min(pcts) - y_pad, max(pcts) + y_pad)
        fig.tight_layout(pad=2.0)
        return save_figure(fig, "speedup_ratio.png")


def chart_pct_diff_sc_mc(machines: list[str], data: dict) -> str:
    """
    Grouped bar chart with SC and MC % differences (m2 vs m1) side by side
    for each workload. SC bars in blue, MC bars in orange.
    """
    if len(machines) < 2:
        return ""
    m1, m2 = machines[0], machines[1]

    all_names = list(dict.fromkeys(w["name"] for w in data[m1]["workloads"]))

    def score(mm: str, name: str, key: str) -> int:
        for w in data[mm]["workloads"]:
            if w["name"] == name:
                return w[key]
        return 0

    sc_pcts = [_pct_diff(score(m1, n, "single_core"), score(m2, n, "single_core"))
               for n in all_names]
    mc_pcts = [_pct_diff(score(m1, n, "multi_core"),  score(m2, n, "multi_core"))
               for n in all_names]

    bar_width = 0.35
    x         = list(range(len(all_names)))
    all_vals  = sc_pcts + mc_pcts
    abs_max   = max(abs(v) for v in all_vals) if all_vals else 1.0

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(max(13, len(all_names) * 1.35), 6.0))

        sc_offsets = [xi - bar_width * 0.5 for xi in x]
        mc_offsets = [xi + bar_width * 0.5 for xi in x]

        sc_bars = ax.bar(sc_offsets, sc_pcts, width=bar_width * 0.9,
                         color=SC_COLOR, edgecolor="white", alpha=0.85,
                         label="Single-Core")
        mc_bars = ax.bar(mc_offsets, mc_pcts, width=bar_width * 0.9,
                         color=MC_COLOR, edgecolor="white", alpha=0.85,
                         label="Multi-Core")

        for bars, pcts_list in [(sc_bars, sc_pcts), (mc_bars, mc_pcts)]:
            for bar, val in zip(bars, pcts_list):
                if val == 0:
                    continue
                y_text = val + abs_max * 0.03 if val >= 0 else val - abs_max * 0.03
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    y_text, _fmt_pct(val),
                    ha="center", va="bottom" if val >= 0 else "top",
                    fontsize=7, fontweight="bold", color="#333333",
                )

        ax.axhline(0, color="#888888", lw=1.5, ls="--", zorder=0)
        ax.set_xticks(x)
        ax.set_xticklabels(all_names, rotation=38, ha="right", fontsize=8)
        ax.set_ylabel(f"% difference  ({m2} vs {m1})", fontsize=10)
        ax.set_title(
            f"Single-Core & Multi-Core % Difference:  {m2}  vs  {m1}",
            fontsize=12, fontweight="bold", pad=12,
        )
        y_pad = abs_max * 0.20
        ax.set_ylim(min(all_vals) - y_pad, max(all_vals) + y_pad)
        ax.legend(fontsize=10)
        fig.tight_layout(pad=2.2)
        return save_figure(fig, "pct_diff_sc_mc.png")


# ---------------------------------------------------------------------------
# Excel sheets
# ---------------------------------------------------------------------------
def write_overview_sheet(wb: Workbook, machines: list[str], data: dict) -> None:
    ws     = wb.create_sheet("Overview")
    n_cols = 3 if len(machines) == 1 else 4
    write_title_row(ws, "Geekbench 6 — Machine Comparison", n_cols)

    headers = ["Metric"] + machines + (["Ratio (M2/M1)"] if len(machines) > 1 else [])
    for ci, h in enumerate(headers, 1):
        style_header_cell(ws.cell(2, ci), h, bg="mid")
        set_col_width(ws, ci, 22)
    ws.row_dimensions[2].height = 22

    for ri, (label, key) in enumerate([("Single-Core Score", "single_core"),
                                        ("Multi-Core Score",  "multi_core")], 3):
        bg = "alt" if ri % 2 == 0 else "white"
        v1 = data[machines[0]][key]
        style_data_cell(ws.cell(ri, 1), label, bg="light", bold=True, align="left")
        style_data_cell(ws.cell(ri, 2), v1, fmt="#,##0", bg=bg)
        if len(machines) > 1:
            v2 = data[machines[1]][key]
            style_data_cell(ws.cell(ri, 3), v2, fmt="#,##0", bg=bg)
            if v1 > 0:
                ratio  = round(v2 / v1, 4)
                fg_key = "green_fg" if ratio >= 1 else "red_fg"
                bg_key = "green_bg" if ratio >= 1 else "red_bg"
                style_data_cell(ws.cell(ri, 4), ratio, fmt="0.0000",
                                bg=bg_key, fg=fg_key, bold=True)
            else:
                style_data_cell(ws.cell(ri, 4), "-", bg=bg, align="center")

    ri = 6
    ws.merge_cells(f"A{ri}:{get_column_letter(n_cols)}{ri}")
    style_header_cell(ws.cell(ri, 1), "Result URLs", bg="mid")
    ri += 1
    for m in machines:
        style_data_cell(ws.cell(ri, 1), m, bg="light", bold=True, align="left")
        url_cell           = ws.cell(ri, 2, value=data[m]["url"])
        url_cell.font      = Font(name=FONT_NAME, color="0563C1", size=10, underline="single")
        url_cell.alignment = Alignment(horizontal="left")
        ws.merge_cells(f"B{ri}:{get_column_letter(n_cols)}{ri}")
        ri += 1


def write_workloads_sheet(wb: Workbook, machines: list[str], data: dict) -> None:
    ws = wb.create_sheet("Workloads")
    n_ratio = 1 if len(machines) > 1 else 0
    n_cols  = 1 + len(machines) * 2 + n_ratio + len(machines)
    write_title_row(ws, "Geekbench 6 — Workload Breakdown", n_cols)

    headers = ["Workload"]
    for m in machines:
        headers += [f"{m}\nSingle-Core", f"{m}\nMulti-Core"]
    if len(machines) > 1:
        headers.append("Ratio MC\n(M2/M1)")
    for m in machines:
        headers.append(f"{m}\nDescription (SC)")

    for ci, h in enumerate(headers, 1):
        style_header_cell(ws.cell(2, ci), h, bg="mid")
        set_col_width(ws, ci, 22 if ci == 1 else 15)
    ws.row_dimensions[2].height = 30

    all_names = list(dict.fromkeys(
        w["name"] for m in machines for w in data[m]["workloads"]
    ))

    def get_wb(m, name):
        for w in data[m]["workloads"]:
            if w["name"] == name:
                return w
        return {"single_core": 0, "multi_core": 0, "single_desc": "", "multi_desc": ""}

    for ri, name in enumerate(all_names, 3):
        bg = "alt" if ri % 2 == 0 else "white"
        style_data_cell(ws.cell(ri, 1), name, bg="light", bold=True, align="left")
        col = 2
        first_mc = 0
        for mi, m in enumerate(machines):
            w = get_wb(m, name)
            style_data_cell(ws.cell(ri, col),     w["single_core"], fmt="#,##0", bg=bg)
            style_data_cell(ws.cell(ri, col + 1), w["multi_core"],  fmt="#,##0", bg=bg)
            if mi == 0:
                first_mc = w["multi_core"]
            col += 2
        if len(machines) > 1:
            second_mc = get_wb(machines[1], name)["multi_core"]
            if first_mc > 0 and second_mc > 0:
                ratio  = round(second_mc / first_mc, 4)
                fg_key = "green_fg" if ratio >= 1 else "red_fg"
                bg_key = "green_bg" if ratio >= 1 else "red_bg"
                style_data_cell(ws.cell(ri, col), ratio, fmt="0.0000",
                                bg=bg_key, fg=fg_key, bold=True)
            else:
                style_data_cell(ws.cell(ri, col), "-", bg=bg, align="center")
            col += 1
        for m in machines:
            style_data_cell(ws.cell(ri, col), get_wb(m, name)["single_desc"],
                            bg=bg, align="left")
            col += 1


def write_sysinfo_sheet(wb: Workbook, machines: list[str], data: dict) -> None:
    ws = wb.create_sheet("System Info")
    write_title_row(ws, "Hardware & System Specifications", 3)
    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 36)
    set_col_width(ws, 3, 52)

    ri = 2
    for m in machines:
        d = data[m]
        ws.merge_cells(f"A{ri}:C{ri}")
        c            = ws.cell(ri, 1, value=f"  {m}  —  {d['url']}")
        c.font       = Font(name=FONT_NAME, bold=True, size=11, color=C["white"])
        c.fill       = PatternFill("solid", fgColor=C["mid"])
        c.alignment  = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[ri].height = 20
        ri += 1

        for sec_title, sec_data in [("System", d["system"]),
                                     ("CPU",    d["cpu"]),
                                     ("Memory", d["memory"])]:
            if not sec_data:
                continue
            ws.merge_cells(f"B{ri}:C{ri}")
            c2           = ws.cell(ri, 2, value=sec_title)
            c2.font      = Font(name=FONT_NAME, bold=True, size=10, color=C["white"])
            c2.fill      = PatternFill("solid", fgColor=C["dark"])
            c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[ri].height = 18
            ri += 1
            for key, val in sec_data.items():
                bg = "alt" if ri % 2 == 0 else "white"
                style_data_cell(ws.cell(ri, 2), key, bg="light", bold=True, align="left")
                style_data_cell(ws.cell(ri, 3), val, bg=bg, align="left")
                ri += 1

        style_data_cell(ws.cell(ri, 2), "Platform / Version",
                        bg="light", bold=True, align="left")
        style_data_cell(ws.cell(ri, 3), d.get("platform", ""), bg="white", align="left")
        ri += 2

        info = load_machine_info(m)
        if info != "No machine info available.":
            ws.merge_cells(f"A{ri}:C{ri}")
            c3           = ws.cell(ri, 1, value=f"  {m}  — machine_info.txt")
            c3.font      = Font(name=FONT_NAME, bold=True, size=10, color=C["white"])
            c3.fill      = PatternFill("solid", fgColor=C["dark"])
            c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[ri].height = 18
            ri += 1
            for line in info.splitlines():
                bg = "alt" if ri % 2 == 0 else "white"
                ws.merge_cells(f"A{ri}:C{ri}")
                style_data_cell(ws.cell(ri, 1), line, bg=bg, align="left")
                ri += 1
            ri += 1


def write_charts_sheet(wb: Workbook, paths: list[str]) -> None:
    from openpyxl.drawing.image import Image as XLImage
    ws = wb.create_sheet("Charts")
    write_title_row(ws, "Visual Comparison — Geekbench 6", 12)
    anchors = ["A2", "M2", "A28", "M28", "A54", "M54"]
    for anchor, path in zip(anchors, paths):
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = 700
            img.height = 390
            ws.add_image(img, anchor)


# ---------------------------------------------------------------------------
# LaTeX table
# ---------------------------------------------------------------------------
def write_latex_table(machines: list[str], data: dict) -> None:
    m1 = machines[0]
    m2 = machines[1] if len(machines) > 1 else None

    def s(mm, name, key):
        for w in data[mm]["workloads"]:
            if w["name"] == name:
                return w[key]
        return 0

    out = []
    out.append(r"\begin{table}[htbp]")
    out.append(r"\centering")
    out.append(r"\small")
    out.append(r"\caption{Comparativa de rendimiento Geekbench 6}")
    out.append(r"\label{tab:geekbench}")

    if m2:
        out.append(r"\begin{tabular}{lrrrrrr}")
        out.append(r"\toprule")
        out.append(f" & \\multicolumn{{2}}{{c}}{{{m1}}} & \\multicolumn{{2}}{{c}}{{{m2}}} & \\\\")
        out.append(r"\cmidrule(lr){2-3}\cmidrule(lr){4-5}")
        out.append(r"Workload & SC & MC & SC & MC & Ratio MC \\")
        out.append(r"\midrule")
        sc1 = data[m1]["single_core"]; mc1 = data[m1]["multi_core"]
        sc2 = data[m2]["single_core"]; mc2 = data[m2]["multi_core"]
        r0  = f"{mc2/mc1:.3f}" if mc1 > 0 else "--"
        out.append(f"\\textbf{{Overall}} & {sc1:,} & {mc1:,} & {sc2:,} & {mc2:,} & {r0} \\\\")
        out.append(r"\midrule")
        all_names = list(dict.fromkeys(
            w["name"] for mm in machines for w in data[mm]["workloads"]
        ))
        for name in all_names:
            v_sc1 = s(m1, name, "single_core"); v_mc1 = s(m1, name, "multi_core")
            v_sc2 = s(m2, name, "single_core"); v_mc2 = s(m2, name, "multi_core")
            r_str = f"{v_mc2/v_mc1:.3f}" if v_mc1 > 0 and v_mc2 > 0 else "--"
            safe  = name.replace("&", r"\&").replace("#", r"\#")
            out.append(f"\\quad {safe} & {v_sc1:,} & {v_mc1:,} & {v_sc2:,} & {v_mc2:,} & {r_str} \\\\")
        out.append(r"\bottomrule")
        out.append(r"\end{tabular}")
    else:
        out.append(r"\begin{tabular}{lrr}")
        out.append(r"\toprule")
        out.append(r"Workload & Single-Core & Multi-Core \\")
        out.append(r"\midrule")
        out.append(f"\\textbf{{Overall}} & {data[m1]['single_core']:,} & {data[m1]['multi_core']:,} \\\\")
        out.append(r"\midrule")
        for w in data[m1]["workloads"]:
            safe = w["name"].replace("&", r"\&").replace("#", r"\#")
            out.append(f"\\quad {safe} & {w['single_core']:,} & {w['multi_core']:,} \\\\")
        out.append(r"\bottomrule")
        out.append(r"\end{tabular}")

    out.append(r"\end{table}")
    with open(TEX_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")
    print(f"  LaTeX -> {TEX_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    machines = LABELS or discover_machines()
    if not machines:
        print(f"[ERROR] No se encontraron carpetas con result_url.txt en {BASE_DIR}/")
        print("  Ejecuta bench_geekbench.sh en cada maquina primero.")
        sys.exit(1)
    if len(machines) > 2:
        machines = machines[:2]
        print(f"  Usando las primeras dos maquinas: {machines}")

    print(f"Maquinas: {machines}")

    data: dict[str, dict] = {}
    for m in machines:
        url     = get_result_url(m)
        data[m] = scrape_geekbench(url, m)
        time.sleep(1)

    for m in machines:
        cpu = data[m]["cpu"].get("Name", "CPU desconocida")
        print(f"  {m}: SC={data[m]['single_core']:,}  MC={data[m]['multi_core']:,}  ({cpu})")

    print("\nGenerando graficas...")
    chart_paths = [
        chart_overall(machines, data),
        chart_workloads(machines, data, "single_core",
                        "Single-Core Workload Scores", "workloads_single.png"),
        chart_workloads(machines, data, "multi_core",
                        "Multi-Core Workload Scores",  "workloads_multi.png"),
        chart_speedup(machines, data),
        chart_pct_diff_sc_mc(machines, data),
    ]
    chart_paths = [p for p in chart_paths if p]
    for p in chart_paths:
        print(f"  {os.path.basename(p)}")

    print("\nConstruyendo workbook Excel...")
    excel_wb = Workbook()
    if excel_wb.active:
        excel_wb.remove(excel_wb.active)
    write_overview_sheet(excel_wb, machines, data)
    write_workloads_sheet(excel_wb, machines, data)
    write_sysinfo_sheet(excel_wb, machines, data)
    write_charts_sheet(excel_wb, chart_paths)
    excel_wb.save(XLSX_PATH)
    print(f"  Excel -> {XLSX_PATH}")

    print("\nGenerando tabla LaTeX...")
    write_latex_table(machines, data)

    print("\nListo.")


if __name__ == "__main__":
    main()