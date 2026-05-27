from __future__ import annotations

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

FONT = "Calibri"
BEST_MARKER = "-march=native"
SIZES = [400, 800, 1600, 3200, 6400]
PROCESSES = [2, 3, 4]
C_DARK = "1F4E79"
C_MID = "2E75B6"
C_LIGHT = "BDD7EE"
C_ALT = "EBF3FB"
C_SEP = "D9E1F2"
C_SP_BG = "E2EFDA"
C_SP_FG = "375623"
C_REF_BG = "FFF2CC"
CHART_STYLE = {
    "axes.facecolor": "#F5F5F5",
    "figure.facecolor": "white",
    "axes.grid": True,
    "grid.color": "#DDDDDD",
    "grid.linestyle": "--",
    "grid.linewidth": 0.6,
    "axes.spines.top": False,
    "axes.spines.right": False,
}
SEQ_COLOR = "#FF6600"
NOOPT_COLORS = {1: "#AAAAAA", 2: "#4472C4", 3: "#70AD47", 4: "#ED7D31"}
OPT_COLORS = {2: "#1F4E79", 3: "#375623", 4: "#843C0C"}


def thin_border() -> Border:
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(cell, value, bg=C_MID, fg="FFFFFF", bold=True, size=10, align="center"):
    cell.value = value
    cell.font = Font(name=FONT, bold=bold, size=size, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()


def dat(cell, value, fmt=None, bg="FFFFFF", fg="000000", bold=False, align="right"):
    cell.value = value
    cell.font = Font(name=FONT, bold=bold, size=10, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt


def summ(cell, value, fmt=None, bold=False):
    dat(cell, value, fmt=fmt, bg=C_ALT, fg=C_DARK, bold=bold)


def sp_cell(cell, value, fmt="0.00", is_ref=False):
    bg = C_REF_BG if is_ref else C_SP_BG
    fg = "000000" if is_ref else C_SP_FG
    dat(cell, value, fmt=fmt, bg=bg, fg=fg, bold=True)


def title_row(ws, text, n_cols, row=1):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols + 1)
    c = ws.cell(row, 2, value=text)
    c.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def section_header(ws, text, row, n_cols):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols + 1)
    c = ws.cell(row, 2, value=text)
    c.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_MID)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def impl_header(ws, text, row, n_cols):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols + 1)
    c = ws.cell(row, 2, value=text)
    c.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def tag_from_flags(flags: str) -> str:
    return "opt" if BEST_MARKER in str(flags) else "noopt"


def load_data(compiler_csv: str, mpi_csv: str):
    df_compiler = pd.read_csv(compiler_csv)
    df_mpi = pd.read_csv(mpi_csv)
    df_compiler.columns = [c.strip().lower() for c in df_compiler.columns]
    df_mpi.columns = [c.strip().lower() for c in df_mpi.columns]
    df_compiler["flags"] = df_compiler["flags"].astype(str).str.strip('"')
    df_mpi["flags"] = df_mpi["flags"].astype(str).str.strip('"')

    seq = df_compiler[df_compiler["impl"] == "seq_std"].copy()
    seq["tag"] = seq["flags"].apply(tag_from_flags)
    seq["processes"] = 0
    seq["total_ms"] = seq["wall_time_ms"]
    seq_opt = seq[seq["tag"] == "opt"].copy()

    df_mpi["tag"] = df_mpi["flags"].apply(tag_from_flags)
    return seq_opt, df_mpi


def build_seq_reps(seq_opt: pd.DataFrame) -> pd.DataFrame:
    return (
        seq_opt[["matrix_size", "repetition", "total_ms"]]
        .pivot(index="repetition", columns="matrix_size", values="total_ms")
        .reindex(columns=SIZES)
        .sort_index()
    )


def build_mpi_reps(df_mpi: pd.DataFrame, tag: str, procs: int) -> pd.DataFrame:
    sub = df_mpi[(df_mpi["tag"] == tag) & (df_mpi["processes"] == procs)]
    return (
        sub[["matrix_size", "repetition", "total_ms"]]
        .pivot(index="repetition", columns="matrix_size", values="total_ms")
        .reindex(columns=SIZES)
        .sort_index()
    )


def avg_by_size(df: pd.DataFrame) -> pd.Series:
    return df.groupby("matrix_size")["total_ms"].mean().reindex(SIZES)


def mpi_avg(df_mpi: pd.DataFrame, tag: str) -> pd.DataFrame:
    sub = df_mpi[df_mpi["tag"] == tag]
    return (
        sub.groupby(["processes", "matrix_size"])["total_ms"]
        .mean()
        .unstack("matrix_size")
        .reindex(index=PROCESSES, columns=SIZES)
    )


def save_fig(fig, out_dir: str, name: str) -> str:
    path = os.path.join(out_dir, name)
    fig.savefig(path, dpi=160, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def make_charts(seq_avg, mpi_noopt_avg, mpi_opt_avg, sp_noopt, sp_opt, charts_dir: str):
    x_labels = [str(s) for s in SIZES]

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.plot(x_labels, seq_avg.values, marker="o", lw=2, color=SEQ_COLOR, label="Sequential (with -O3)")
        for p in PROCESSES:
            ax.plot(x_labels, mpi_noopt_avg.loc[p].values, marker="s", lw=1.8, color=NOOPT_COLORS[p], label=f"MPI {p} process(es) — no opt")
        ax.set_yscale("log")
        ax.set_title("Execution Time: Sequential vs MPI (no compiler opt)", fontsize=12, fontweight="bold")
        ax.set_xlabel("Matrix size N", fontsize=10)
        ax.set_ylabel("Average time (ms) — log scale", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
        chart_time_noopt = save_fig(fig, charts_dir, "time_seq_vs_mpi_noopt.png")

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.plot(x_labels, seq_avg.values, marker="o", lw=2, color=SEQ_COLOR, label="Sequential (with -O3)")
        for p in [2, 3, 4]:
            ax.plot(x_labels, mpi_opt_avg.loc[p].values, marker="s", lw=1.8, color=OPT_COLORS[p], label=f"MPI {p} process(es) — with opt")
        ax.set_yscale("log")
        ax.set_title("Execution Time: Sequential vs MPI (with -O3 opt)", fontsize=12, fontweight="bold")
        ax.set_xlabel("Matrix size N", fontsize=10)
        ax.set_ylabel("Average time (ms) — log scale", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
        chart_time_opt = save_fig(fig, charts_dir, "time_seq_vs_mpi_opt.png")

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--", label="Baseline (speedup = 1)")
        for p in PROCESSES:
            ax.plot(x_labels, sp_noopt.loc[p].values, marker="o", lw=1.8, color=NOOPT_COLORS[p], label=f"MPI {p} process(es)")
        ax.set_title("Speedup: Sequential (opt) vs MPI (no opt)", fontsize=12, fontweight="bold")
        ax.set_xlabel("Matrix size N", fontsize=10)
        ax.set_ylabel("Speedup (seq_opt / mpi_noopt)", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
        chart_sp_noopt = save_fig(fig, charts_dir, "speedup_seq_vs_mpi_noopt.png")

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--", label="Baseline (speedup = 1)")
        for p in [2, 3, 4]:
            ax.plot(x_labels, sp_opt.loc[p].values, marker="o", lw=1.8, color=OPT_COLORS[p], label=f"MPI {p} process(es)")
        ax.set_title("Speedup: Sequential (opt) vs MPI (with opt)", fontsize=12, fontweight="bold")
        ax.set_xlabel("Matrix size N", fontsize=10)
        ax.set_ylabel("Speedup (seq_opt / mpi_opt)", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
        chart_sp_opt = save_fig(fig, charts_dir, "speedup_seq_vs_mpi_opt.png")

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        for sz, color in {1600: "#4472C4", 3200: "#70AD47", 6400: "#ED7D31"}.items():
            ys = [sp_opt.loc[p, sz] if (p in sp_opt.index and not np.isnan(sp_opt.loc[p, sz])) else None for p in [2, 3, 4]]
            valid = [(x, y) for x, y in zip([2, 3, 4], ys) if y is not None]
            if valid:
                ax.plot([v[0] for v in valid], [v[1] for v in valid], marker="o", lw=1.8, color=color, label=f"N={sz:,} (opt)")
        ax.plot([2, 3, 4], [1, 1.5, 2], color="#AAAAAA", lw=1.2, ls="--", label="Ideal linear scaling")
        ax.set_title("Speedup vs Process Count (MPI with opt, large N)", fontsize=12, fontweight="bold")
        ax.set_xlabel("Number of MPI processes", fontsize=10)
        ax.set_ylabel("Speedup vs sequential (opt)", fontsize=10)
        ax.set_xticks([2, 3, 4])
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
        chart_sp_procs = save_fig(fig, charts_dir, "speedup_vs_processes.png")

    return {
        "chart_time_noopt": chart_time_noopt,
        "chart_time_opt": chart_time_opt,
        "chart_sp_noopt": chart_sp_noopt,
        "chart_sp_opt": chart_sp_opt,
        "chart_sp_procs": chart_sp_procs,
    }


def write_raw_sheet(wb, sheet_name, sheet_title, variants):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 3
    n_cols = 1 + len(SIZES)
    title_row(ws, sheet_title, n_cols, row=2)
    section_header(ws, "Table 1 — Individual Measurements (ms)", 4, n_cols)
    ws.column_dimensions["B"].width = 8
    for ci, _ in enumerate(SIZES, 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    cur = 5
    for vi, variant in enumerate(variants):
        impl_header(ws, variant["label"], cur, n_cols)
        cur += 1
        hdr(ws.cell(cur, 2), "Rep", bg=C_LIGHT, fg="000000", size=9)
        for ci, sz in enumerate(SIZES, 3):
            hdr(ws.cell(cur, ci), f"N = {sz:,}", bg=C_LIGHT, fg="000000", size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        reps_df = variant["reps_df"]
        for rep in range(1, 11):
            alt_bg = C_ALT if rep % 2 == 0 else "FFFFFF"
            dat(ws.cell(cur, 2), rep, fmt="0", bg=C_LIGHT, bold=True, align="center")
            for ci, sz in enumerate(SIZES, 3):
                val = reps_df.loc[rep, sz] if (rep in reps_df.index and sz in reps_df.columns) else None
                if val is not None and not pd.isna(val):
                    dat(ws.cell(cur, ci), round(float(val), 3), fmt="#,##0.000", bg=alt_bg)
                else:
                    dat(ws.cell(cur, ci), "N/A", bg=alt_bg, align="center")
            ws.row_dimensions[cur].height = 16
            cur += 1

        for label, idx in [("Average", 0), ("Std Dev", 1), ("CV (%)", 2)]:
            hdr(ws.cell(cur, 2), label, bg=C_ALT, fg=C_DARK, size=9, align="left")
            for ci, sz in enumerate(SIZES, 3):
                if sz not in reps_df.columns:
                    summ(ws.cell(cur, ci), "N/A")
                    continue
                vals = reps_df[sz].dropna().values
                if len(vals) == 0:
                    summ(ws.cell(cur, ci), "N/A")
                    continue
                mean = float(np.mean(vals))
                std = float(np.std(vals))
                cv = std / mean * 100 if mean > 0 else 0.0
                show = mean if idx == 0 else (std if idx == 1 else cv)
                fmt = "#,##0.000" if idx < 2 else "0.00"
                summ(ws.cell(cur, ci), round(show, 3 if idx < 2 else 2), fmt=fmt, bold=(idx == 0))
            ws.row_dimensions[cur].height = 16
            cur += 1

        if vi < len(variants) - 1:
            for ci in range(2, n_cols + 2):
                ws.cell(cur, ci).fill = PatternFill("solid", fgColor=C_SEP)
            ws.row_dimensions[cur].height = 6
            cur += 1


def write_speedup_sheet(wb, sheet_name, sheet_title, ref_label, ref_avg, variants_sp, chart_time_path, chart_sp_path, extra_chart=None):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 3
    last_col = 2 + len(SIZES) * 2
    n_cols = 1 + len(SIZES) * 2 + 1
    title_row(ws, sheet_title, n_cols, row=2)
    section_header(ws, "Table 2 — Average Execution Time and Speedup", 4, n_cols)
    ws.column_dimensions["B"].width = 38
    for si in range(len(SIZES)):
        ac = 3 + si * 2
        sc = ac + 1
        ws.column_dimensions[get_column_letter(ac)].width = 14
        ws.column_dimensions[get_column_letter(sc)].width = 10
    ws.column_dimensions[get_column_letter(last_col)].width = 12

    cur = 5
    hdr(ws.cell(cur, 2), "Implementation", bg=C_DARK, size=10)
    for si, sz in enumerate(SIZES):
        ac = 3 + si * 2
        sc = ac + 1
        hdr(ws.cell(cur, ac), f"N={sz:,}\nAvg (ms)", bg=C_MID, size=9)
        hdr(ws.cell(cur, sc), f"N={sz:,}\nSpeedup", bg=C_MID, size=9)
    hdr(ws.cell(cur, last_col), "Avg\nSpeedup", bg=C_DARK, size=9)
    ws.row_dimensions[cur].height = 28
    cur += 1

    dat(ws.cell(cur, 2), ref_label, bg=C_LIGHT, bold=True, align="left")
    for si, sz in enumerate(SIZES):
        ac = 3 + si * 2
        sc = ac + 1
        val = ref_avg.get(sz)
        dat(ws.cell(cur, ac), round(float(val), 3) if val is not None else "N/A", fmt="#,##0.000", bg=C_REF_BG)
        sp_cell(ws.cell(cur, sc), 1.0, fmt="0.00", is_ref=True)
    sp_cell(ws.cell(cur, last_col), 1.0, fmt="0.00", is_ref=True)
    cur += 1

    for ri, variant in enumerate(variants_sp):
        bg = C_ALT if ri % 2 == 0 else "FFFFFF"
        dat(ws.cell(cur, 2), variant["label"], bg=C_LIGHT, bold=True, align="left")
        sp_vals = []
        for si, sz in enumerate(SIZES):
            ac = 3 + si * 2
            sc = ac + 1
            avg_val = variant["avg"].get(sz)
            sp_val = variant["sp_series"].get(sz)
            dat(ws.cell(cur, ac), round(float(avg_val), 3) if avg_val is not None and not pd.isna(avg_val) else "N/A", fmt="#,##0.000", bg=bg)
            if sp_val is not None and not pd.isna(sp_val):
                sp_cell(ws.cell(cur, sc), round(float(sp_val), 4), fmt="0.0000")
                sp_vals.append(float(sp_val))
            else:
                dat(ws.cell(cur, sc), "N/A", bg=C_SP_BG, align="center")
        avg_sp = round(sum(sp_vals) / len(sp_vals), 2) if sp_vals else "N/A"
        if isinstance(avg_sp, float):
            sp_cell(ws.cell(cur, last_col), avg_sp, fmt="0.00")
        else:
            dat(ws.cell(cur, last_col), avg_sp, bg=C_SP_BG, align="center")
        cur += 1

    cur += 2
    for anchor, path in [(f"B{cur}", chart_time_path), (f"L{cur}", chart_sp_path)]:
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width = 600
            img.height = 350
            ws.add_image(img, anchor)
    if extra_chart and os.path.exists(extra_chart):
        img = XLImage(extra_chart)
        img.width = 600
        img.height = 350
        ws.add_image(img, f"B{cur + 24}")


def write_overview(wb, sp_noopt, sp_opt):
    from openpyxl.worksheet.hyperlink import Hyperlink

    ws = wb.create_sheet("Overview")
    wb.move_sheet("Overview", offset=-len(wb.sheetnames) + 1)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    for ci in range(3, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    title_row(ws, "MPI Speedup Report — Matrix Multiplication Benchmark", 5, row=2)
    ws.cell(4, 2, value="Reference baseline: Sequential implementation compiled with -O3 -march=native -funroll-loops -flto -ffast-math -fomit-frame-pointer")
    ws.cell(4, 2).font = Font(name=FONT, italic=True, size=10, color="595959")

    section_header(ws, "Sheet Index", 6, 5)
    sheets_info = [
        ("Raw Measurements — MPI No Opt", "Raw_MPI_NoOpt", "Raw execution time data for MPI without compiler optimization"),
        ("Raw Measurements — MPI With Opt", "Raw_MPI_Opt", "Raw execution time data for MPI with -O3 full optimization"),
        ("Speedup — Sequential vs MPI No Opt", "SP_Seq_vs_NoOpt", "Speedup analysis: seq (opt) vs MPI (no opt)"),
        ("Speedup — Sequential vs MPI Opt", "SP_Seq_vs_Opt", "Speedup analysis: seq (opt) vs MPI (with opt)"),
    ]
    for i, (label, target, desc) in enumerate(sheets_info, 8):
        c = ws.cell(i, 2, value=label)
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"'{target}'!B2")
        c.font = Font(name=FONT, color="1F4E79", underline="single", size=10)
        ws.cell(i, 3, value=desc).font = Font(name=FONT, size=10, color="595959")

    section_header(ws, "Summary — Average Speedup (Sequential opt vs MPI)", 14, 5)
    hdr(ws.cell(15, 2), "Configuration", bg=C_DARK, size=10)
    hdr(ws.cell(15, 3), "No-opt Avg Speedup", bg=C_MID, size=9)
    hdr(ws.cell(15, 4), "With-opt Avg Speedup", bg=C_MID, size=9)
    hdr(ws.cell(15, 5), "Best config (opt)", bg=C_MID, size=9)

    row = 16
    for p in PROCESSES:
        bg = C_ALT if row % 2 == 0 else "FFFFFF"
        dat(ws.cell(row, 2), f"MPI {p} process(es)", bg=C_LIGHT, bold=True, align="left")
        no_vals = sp_noopt.loc[p].dropna()
        avg_no = round(float(no_vals.mean()), 2) if len(no_vals) > 0 else "N/A"
        if p in sp_opt.index:
            opt_vals = sp_opt.loc[p].dropna()
            avg_opt = round(float(opt_vals.mean()), 2) if len(opt_vals) > 0 else "N/A"
        else:
            avg_opt = "N/A"
        if p in sp_opt.index and len(sp_opt.loc[p].dropna()) > 0:
            best_sz = int(sp_opt.loc[p].idxmax())
            best_v = round(float(sp_opt.loc[p].max()), 2)
            best_str = f"N={best_sz:,} → {best_v}x"
        else:
            best_str = "N/A"
        if isinstance(avg_no, float):
            sp_cell(ws.cell(row, 3), avg_no, fmt="0.00")
        else:
            dat(ws.cell(row, 3), avg_no, bg=bg, align="center")
        if isinstance(avg_opt, float):
            sp_cell(ws.cell(row, 4), avg_opt, fmt="0.00")
        else:
            dat(ws.cell(row, 4), avg_opt, bg=bg, align="center")
        dat(ws.cell(row, 5), best_str, bg=bg, align="center", fg=C_SP_FG if best_str != "N/A" else "000000", bold=True)
        row += 1


def build_report(compiler_csv: str, mpi_csv: str, output_xlsx: str, charts_dir: str | None = None):
    charts_dir = charts_dir or os.path.join(os.path.dirname(output_xlsx) or ".", "charts_mpi")
    os.makedirs(charts_dir, exist_ok=True)

    seq_opt, df_mpi = load_data(compiler_csv, mpi_csv)
    seq_reps = build_seq_reps(seq_opt)
    seq_avg = avg_by_size(seq_opt)
    mpi_noopt_avg = mpi_avg(df_mpi, "noopt")
    mpi_opt_avg = mpi_avg(df_mpi, "opt")
    sp_noopt = seq_avg / mpi_noopt_avg
    sp_opt = seq_avg / mpi_opt_avg
    charts = make_charts(seq_avg, mpi_noopt_avg, mpi_opt_avg, sp_noopt, sp_opt, charts_dir)

    wb = Workbook()
    wb.remove(wb.active)

    noopt_variants = [{"label": "Sequential — with -O3 (reference baseline)", "reps_df": seq_reps}]
    for p in PROCESSES:
        noopt_variants.append({"label": f"MPI {p} process(es) — no compiler optimization", "reps_df": build_mpi_reps(df_mpi, "noopt", p)})
    write_raw_sheet(wb, "Raw_MPI_NoOpt", "Raw Measurements: Sequential (opt) vs MPI — No Compiler Optimization", noopt_variants)

    opt_variants = [{"label": "Sequential — with -O3 (reference baseline)", "reps_df": seq_reps}]
    for p in [2, 3, 4]:
        opt_variants.append({"label": f"MPI {p} process(es) — with -O3 -march=native full optimization", "reps_df": build_mpi_reps(df_mpi, "opt", p)})
    write_raw_sheet(wb, "Raw_MPI_Opt", "Raw Measurements: Sequential (opt) vs MPI — With -O3 Compiler Optimization", opt_variants)

    sp_noopt_variants = []
    for p in PROCESSES:
        sp_noopt_variants.append({"label": f"MPI {p} process(es) — no compiler optimization", "avg": mpi_noopt_avg.loc[p], "sp_series": sp_noopt.loc[p]})
    write_speedup_sheet(wb, "SP_Seq_vs_NoOpt", "Speedup: Sequential (opt) vs MPI — No Compiler Optimization", "Sequential — with -O3 (reference)", seq_avg, sp_noopt_variants, charts["chart_time_noopt"], charts["chart_sp_noopt"])

    sp_opt_variants = []
    for p in [2, 3, 4]:
        sp_opt_variants.append({"label": f"MPI {p} process(es) — with -O3 -march=native full optimization", "avg": mpi_opt_avg.loc[p], "sp_series": sp_opt.loc[p]})
    write_speedup_sheet(wb, "SP_Seq_vs_Opt", "Speedup: Sequential (opt) vs MPI — With -O3 Full Optimization", "Sequential — with -O3 (reference)", seq_avg, sp_opt_variants, charts["chart_time_opt"], charts["chart_sp_opt"], extra_chart=charts["chart_sp_procs"])

    write_overview(wb, sp_noopt, sp_opt)
    wb.save(output_xlsx)


if __name__ == "__main__":
    build_report("./mpi/data_compiler.csv", "./mpi/data_mpi_nfs.csv", "mpi_speedup_report.xlsx")