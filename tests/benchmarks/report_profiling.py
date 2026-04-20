"""
report_profiling.py
...
Reads:
  - data_profiling.csv
  - raw/N*/gprof_report.txt
  - raw/N*/perf_stat.txt
  - raw/N*/perf_record_report.txt
  - raw/N*/perf_mem_report.txt        # <-- agregar
  - raw/N*/cachegrind_report.txt
  - raw/N*/massif_report.txt
  - raw/N*/timing_runs.txt
...
"""

from __future__ import annotations

import html
import math
import os
import re
import sys
import statistics
from typing import Any

import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

CSV_PATH = (
    sys.argv[1]
    if len(sys.argv) > 1
    else "tests/benchmarks/machine1/results_profiling/data_profiling.csv"
)

OUT_DIR = os.path.dirname(os.path.abspath(CSV_PATH))
RAW_DIR = os.path.join(OUT_DIR, "raw")
CHARTS_DIR = os.path.join(OUT_DIR, "charts_profiling")
XLSX_PATH = os.path.join(OUT_DIR, "reporte_profiling.xlsx")
HTML_PATH = os.path.join(OUT_DIR, "profiling_report.html")
LATEX_PATH = os.path.join(OUT_DIR, "profiling_table.tex")

COLORS = {
    "navy": "1F3557",
    "blue": "2E75B6",
    "cyan": "1F9EB7",
    "green": "70AD47",
    "orange": "ED7D31",
    "red": "C00000",
    "purple": "7030A0",
    "gray": "808080",
    "light": "DCE6F1",
    "lighter": "F7FBFF",
    "alt": "F4F8FC",
    "white": "FFFFFF",
    "border": "C9D6E2",
    "good_bg": "E2F0D9",
    "good_fg": "2F6B1E",
    "warn_bg": "FFF2CC",
    "warn_fg": "7F6000",
    "bad_bg": "FCE4D6",
    "bad_fg": "A61C00",
}

CHART_STYLE = {
    "figure.facecolor": "white",
    "axes.facecolor": "white",
    "axes.grid": True,
    "grid.color": "#E7EDF3",
    "grid.linestyle": "--",
    "grid.alpha": 0.75,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.titleweight": "bold",
    "axes.labelsize": 10,
    "axes.titlesize": 12,
    "legend.frameon": False,
    "font.size": 10,
}

plt.rcParams.update(CHART_STYLE)


def border() -> Border:
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def mk_fill(key: str) -> PatternFill:
    return PatternFill("solid", fgColor=COLORS[key])


def set_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def title_row(ws, text: str, cols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, size=14, color=COLORS["white"])
    c.fill = mk_fill("navy")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    ws.row_dimensions[row].height = 24


def header_cell(cell, value: Any, bg: str = "blue") -> None:
    cell.value = value
    cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["white"])
    cell.fill = mk_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()


def data_cell(
    cell,
    value: Any,
    fmt: str | None = None,
    bg: str = "white",
    bold: bool = False,
    align: str = "right",
    fg: str = "000000",
) -> None:
    cell.value = value
    cell.font = Font(name="Calibri", bold=bold, size=10, color=COLORS.get(fg, fg))
    cell.fill = mk_fill(bg if bg in COLORS else "white")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = border()
    if fmt:
        cell.number_format = fmt


def canonical_col(name: str) -> str:
    s = name.strip().lower()
    s = s.replace("%", "pct")
    s = s.replace("-", "_")
    s = s.replace(" ", "_")
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def read_text(path: str) -> str:
    if not os.path.exists(path):
        return ""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def theoretical_heap_mb(n: int) -> float:
    bytes_ = 3 * (n * n * 4 + n * 8)
    return bytes_ / 1024 / 1024


def safe_float(x: Any, default: float = math.nan) -> float:
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def safe_int(x: Any, default: int = 0) -> int:
    try:
        if pd.isna(x):
            return default
        return int(float(x))
    except Exception:
        return default


def fmt_num(x: Any, decimals: int = 3) -> str:
    if x is None:
        return "—"
    try:
        xf = float(x)
        if math.isnan(xf):
            return "—"
        return f"{xf:,.{decimals}f}"
    except Exception:
        return str(x)


def fmt_pct(x: Any, decimals: int = 2) -> str:
    if x is None:
        return "—"
    try:
        xf = float(x)
        if math.isnan(xf):
            return "—"
        return f"{xf:.{decimals}f}%"
    except Exception:
        return str(x)


def fmt_int(x: Any) -> str:
    try:
        xf = float(x)
        if math.isnan(xf):
            return "—"
        return f"{int(round(xf)):,}"
    except Exception:
        return "—"


def classify_ipc(ipc: float) -> str:
    if math.isnan(ipc):
        return "Sin dato"
    if ipc >= 2.5:
        return "Alto"
    if ipc >= 1.5:
        return "Medio"
    return "Bajo"


def classify_miss(pct: float) -> str:
    if math.isnan(pct):
        return "Sin dato"
    if pct < 1:
        return "Excelente"
    if pct < 5:
        return "Bueno"
    if pct < 15:
        return "Moderado"
    return "Alto"


def normalize_dataframe(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [canonical_col(c) for c in df.columns]

    rename_map = {
        "time_ms": "time_mean_ms",
        "matrixsize": "matrix_size",
        "llc_loads": "llc_loads",
        "llc_misses": "llc_misses",
        "llc_miss_pct": "llc_miss_pct",
        "dtlb_loads": "dtlb_loads",
        "dtlb_misses": "dtlb_misses",
        "dtlb_miss_pct": "dtlb_miss_pct",
    }
    df = df.rename(columns={c: rename_map.get(c, c) for c in df.columns})

    if "matrix_size" not in df.columns:
        raise ValueError("CSV inválido: falta la columna matrix_size")

    for col in df.columns:
        if col == "matrix_size":
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        else:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.sort_values("matrix_size").reset_index(drop=True)

    if "time_mean_ms" not in df.columns:
        df["time_mean_ms"] = math.nan
    if "time_std_ms" not in df.columns:
        df["time_std_ms"] = math.nan
    if "gflops" not in df.columns:
        df["gflops"] = math.nan
    if "ipc" not in df.columns:
        df["ipc"] = math.nan
    if "peak_heap_mb" not in df.columns:
        df["peak_heap_mb"] = math.nan

    df["theoretical_heap_mb"] = df["matrix_size"].apply(
        lambda n: theoretical_heap_mb(int(n)) if pd.notna(n) else math.nan
    )
    df["heap_ratio_pct"] = (
        100.0 * df["peak_heap_mb"] / df["theoretical_heap_mb"]
    )
    df["timing_cv_pct"] = 100.0 * df["time_std_ms"] / df["time_mean_ms"]

    if "instructions" in df.columns and "gflops" in df.columns and "time_mean_ms" in df.columns:
        ops = 2.0 * (df["matrix_size"].astype(float) ** 3)
        df["operations"] = ops
        df["ops_per_instruction"] = ops / df["instructions"]
    else:
        df["operations"] = math.nan
        df["ops_per_instruction"] = math.nan

    return df


def parse_timing_runs(path: str) -> dict[str, Any]:
    text = read_text(path)
    values = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            values.append(float(line))
        except ValueError:
            pass

    if not values:
        return {
            "values": [],
            "count": 0,
            "mean_ms": math.nan,
            "std_ms": math.nan,
            "min_ms": math.nan,
            "max_ms": math.nan,
            "median_ms": math.nan,
            "cv_pct": math.nan,
        }

    if len(values) == 1:
        std = 0.0
    else:
        std = statistics.stdev(values)

    mean = statistics.mean(values)
    return {
        "values": values,
        "count": len(values),
        "mean_ms": mean,
        "std_ms": std,
        "min_ms": min(values),
        "max_ms": max(values),
        "median_ms": statistics.median(values),
        "cv_pct": 100.0 * std / mean if mean else math.nan,
    }


def parse_gprof_flat(path: str, top_k: int = 8) -> dict[str, Any]:
    text = read_text(path)
    if not text:
        return {"sample_seconds": math.nan, "top": []}

    m = re.search(r"Each sample counts as\s+([\d.]+)\s+seconds", text)
    sample_seconds = float(m.group(1)) if m else math.nan

    in_flat = False
    rows = []
    for line in text.splitlines():
        if line.strip().startswith("Flat profile"):
            in_flat = True
            continue
        if in_flat and line.strip().startswith("%"):
            continue
        if in_flat and not line.strip():
            continue
        if in_flat and line.startswith("\f"):
            break
        if not in_flat:
            continue

        parts = line.split()
        if len(parts) < 7:
            continue
        try:
            pct_time = float(parts[0])
            cumulative = float(parts[1])
            self_seconds = float(parts[2])
        except ValueError:
            continue

        calls = None
        if parts[3].isdigit():
            calls = int(parts[3])
            name = parts[-1]
            self_ms_call = parts[4]
            total_ms_call = parts[5]
        else:
            name = parts[-1]
            self_ms_call = None
            total_ms_call = None

        rows.append(
            {
                "name": name,
                "pct_time": pct_time,
                "cumulative_s": cumulative,
                "self_s": self_seconds,
                "calls": calls,
                "self_ms_call": safe_float(self_ms_call) if self_ms_call else math.nan,
                "total_ms_call": safe_float(total_ms_call) if total_ms_call else math.nan,
            }
        )

    return {"sample_seconds": sample_seconds, "top": rows[:top_k]}


def parse_perf_record(path: str, top_k: int = 8) -> dict[str, Any]:
    text = read_text(path)
    if not text:
        return {"samples": 0, "event_count": 0, "top": []}

    samples = 0
    event_count = 0

    m = re.search(r"Samples:\s+(\d+)", text)
    if m:
        samples = int(m.group(1))
    m = re.search(r"Event count \(approx\.\):\s+([\d,]+)", text)
    if m:
        event_count = int(m.group(1).replace(",", ""))

    rows = []
    for line in text.splitlines():
        if line.lstrip().startswith("#") or not line.strip():
            continue

        m = re.match(
            r"^\s*([\d.]+)%\s+(\S+)\s+(\S+)\s+(\[[^\]]+\])\s+(.+?)\s*$",
            line,
        )
        if not m:
            continue

        rows.append(
            {
                "overhead_pct": float(m.group(1)),
                "command": m.group(2),
                "shared_object": m.group(3),
                "marker": m.group(4),
                "symbol": m.group(5),
            }
        )

    return {"samples": samples, "event_count": event_count, "top": rows[:top_k]}

def parse_perf_mem(path: str) -> dict[str, Any]:
    """
    Parses the output of `perf mem -t load report --stdio --sort=mem`.
    Format (AMD IBS / Intel PEBS):
        68.74%    439  L2 hit
        14.11%   1434  N/A
         9.72%    988  L1 hit

    Returns a dict with all sources and a filtered list excluding N/A entries.
    """
    text = read_text(path)
    if not text:
        return {"available": False, "sources": [], "sources_classified": []}

    sources = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        # Three-column format: overhead%  samples  label
        m = re.match(r"([\d.]+)%\s+(\d+)\s+(.+)", line)
        if not m:
            # Fallback: two-column format without sample count
            m2 = re.match(r"([\d.]+)%\s+(.+)", line)
            if not m2:
                continue
            pct   = float(m2.group(1))
            label = m2.group(2).strip()
            samples = 0
        else:
            pct     = float(m.group(1))
            samples = int(m.group(2))
            label   = m.group(3).strip()

        sources.append({"label": label, "pct": pct, "samples": samples})

    # Classified: exclude N/A entries and re-normalize percentages
    classified = [s for s in sources if s["label"].upper() != "N/A"]
    total_classified = sum(s["pct"] for s in classified)
    if total_classified > 0:
        for s in classified:
            s["pct_normalized"] = round(s["pct"] / total_classified * 100, 2)
    else:
        for s in classified:
            s["pct_normalized"] = s["pct"]

    na_entries = [s for s in sources if s["label"].upper() == "N/A"]
    na_pct = sum(s["pct"] for s in na_entries)

    return {
        "available": bool(sources),
        "sources": sources,                   
        "sources_classified": classified,     
        "na_pct": na_pct,                     
    }


def parse_cachegrind(path: str, top_k: int = 8, top_lines: int = 6) -> dict[str, Any]:
    text = read_text(path)
    if not text:
        return {"total_ir": 0, "top_functions": [], "hot_lines": []}

    total_ir = 0
    m = re.search(r"^\s*([\d,]+)\s+\(100\.0%\)\s+PROGRAM TOTALS", text, re.MULTILINE)
    if m:
        total_ir = int(m.group(1).replace(",", ""))

    top_functions = []
    in_func_summary = False
    for line in text.splitlines():
        if "-- Function:file summary" in line:
            in_func_summary = True
            continue
        if in_func_summary and line.startswith("-- Annotated source file"):
            break
        if not in_func_summary:
            continue

        m = re.match(r"^\>\s*([\d,]+)\s+\(([\d.]+)%.*?\)\s+(.+?)\s*$", line.strip())
        if not m:
            continue

        top_functions.append(
            {
                "ir": int(m.group(1).replace(",", "")),
                "pct": float(m.group(2)),
                "name": m.group(3),
            }
        )

    hot_lines = []
    for line in text.splitlines():
        m = re.match(r"^\s*([\d,]+)\s+\(([\d.]+)%\)\s+(.+?)\s*$", line)
        if not m:
            continue
        code = m.group(3).strip()
        if not code or code == "." or code.startswith("--") or code.startswith("Unannotated"):
            continue
        hot_lines.append(
            {
                "ir": int(m.group(1).replace(",", "")),
                "pct": float(m.group(2)),
                "code": code,
            }
        )

    hot_lines = sorted(hot_lines, key=lambda x: x["ir"], reverse=True)[:top_lines]
    return {
        "total_ir": total_ir,
        "top_functions": top_functions[:top_k],
        "hot_lines": hot_lines,
    }


def parse_massif(path: str) -> dict[str, Any]:
    text = read_text(path)
    if not text:
        return {
            "snapshots": 0,
            "peak_total_b": 0,
            "peak_useful_b": 0,
            "peak_extra_b": 0,
            "peak_stacks_b": 0,
            "peak_total_mb": math.nan,
            "peak_useful_pct": math.nan,
            "allocators": [],
        }

    snapshots = 0
    m = re.search(r"Number of snapshots:\s+(\d+)", text)
    if m:
        snapshots = int(m.group(1))

    rows = []
    lines = text.splitlines()
    peak_line_idx = None
    for idx, line in enumerate(lines):
        m = re.match(
            r"^\s*(\d+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s*$",
            line,
        )
        if not m:
            continue

        row = {
            "n": int(m.group(1)),
            "time_i": int(m.group(2).replace(",", "")),
            "total_b": int(m.group(3).replace(",", "")),
            "useful_b": int(m.group(4).replace(",", "")),
            "extra_b": int(m.group(5).replace(",", "")),
            "stacks_b": int(m.group(6).replace(",", "")),
            "line_idx": idx,
        }
        rows.append(row)

    if not rows:
        return {
            "snapshots": snapshots,
            "peak_total_b": 0,
            "peak_useful_b": 0,
            "peak_extra_b": 0,
            "peak_stacks_b": 0,
            "peak_total_mb": math.nan,
            "peak_useful_pct": math.nan,
            "allocators": [],
        }

    peak = max(rows, key=lambda x: x["total_b"])
    peak_line_idx = peak["line_idx"]

    allocators = []
    if peak_line_idx is not None:
        for line in lines[peak_line_idx + 1 : peak_line_idx + 20]:
            if line.startswith("--------------------------------------------------------------------------------"):
                break
            m = re.match(r"^\s*->\s*([\d.]+)%\s+\(([\d,]+)B\)\s+.+?:\s+(.+?)\s*$", line)
            if not m:
                continue
            allocators.append(
                {
                    "pct": float(m.group(1)),
                    "bytes": int(m.group(2).replace(",", "")),
                    "where": m.group(3),
                }
            )

    peak_total_mb = peak["total_b"] / 1024 / 1024
    peak_useful_pct = 100.0 * peak["useful_b"] / peak["total_b"] if peak["total_b"] else math.nan

    return {
        "snapshots": snapshots,
        "peak_total_b": peak["total_b"],
        "peak_useful_b": peak["useful_b"],
        "peak_extra_b": peak["extra_b"],
        "peak_stacks_b": peak["stacks_b"],
        "peak_total_mb": peak_total_mb,
        "peak_useful_pct": peak_useful_pct,
        "allocators": allocators,
    }


def collect_raw_analysis(df: pd.DataFrame) -> dict[int, dict[str, Any]]:
    results: dict[int, dict[str, Any]] = {}

    for n in df["matrix_size"].dropna().astype(int).tolist():
        raw_n = os.path.join(RAW_DIR, f"N{n}")
        results[n] = {
            "timing": parse_timing_runs(os.path.join(raw_n, "timing_runs.txt")),
            "gprof": parse_gprof_flat(os.path.join(raw_n, "gprof_report.txt")),
            "perf_record": parse_perf_record(os.path.join(raw_n, "perf_record_report.txt")),
            "perf_mem":     parse_perf_mem(os.path.join(raw_n, "perf_mem_report.txt")),
            "cachegrind": parse_cachegrind(os.path.join(raw_n, "cachegrind_report.txt")),
            "massif": parse_massif(os.path.join(raw_n, "massif_report.txt")),
            "perf_stat_text": read_text(os.path.join(raw_n, "perf_stat.txt")),
        }

    return results


def perf_llc_supported(perf_stat_text: str) -> bool:
    if not perf_stat_text:
        return True
    return "<not supported>" not in perf_stat_text.lower()


def save_figure(fig, name: str) -> str:
    path = os.path.join(CHARTS_DIR, name)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path


def annotate_points(ax, xs, ys, decimals=3, dy=8):
    for x, y in zip(xs, ys):
        if pd.isna(y):
            continue
        ax.annotate(
            f"{y:.{decimals}f}",
            (x, y),
            textcoords="offset points",
            xytext=(0, dy),
            ha="center",
            fontsize=8,
        )


def chart_performance(df: pd.DataFrame) -> str:
    sizes = df["matrix_size"].astype(int).tolist()

    fig, axs = plt.subplots(2, 2, figsize=(12, 8))

    axs[0, 0].errorbar(
        sizes,
        df["time_mean_ms"],
        yerr=df["time_std_ms"],
        marker="o",
        lw=2.4,
        capsize=4,
        color="#2E75B6",
    )
    axs[0, 0].set_title("Tiempo medio de ejecución")
    axs[0, 0].set_xlabel("N")
    axs[0, 0].set_ylabel("ms")

    axs[0, 1].plot(sizes, df["gflops"], marker="o", lw=2.4, color="#1F9EB7")
    axs[0, 1].set_title("Rendimiento")
    axs[0, 1].set_xlabel("N")
    axs[0, 1].set_ylabel("GFLOPS")
    annotate_points(axs[0, 1], sizes, df["gflops"], decimals=3)

    axs[1, 0].plot(sizes, df["ipc"], marker="D", lw=2.4, color="#7030A0")
    axs[1, 0].axhline(1.0, color="#BBBBBB", lw=1, ls="--")
    axs[1, 0].axhline(2.0, color="#BBBBBB", lw=1, ls="--")
    axs[1, 0].set_title("Eficiencia del pipeline")
    axs[1, 0].set_xlabel("N")
    axs[1, 0].set_ylabel("IPC")
    annotate_points(axs[1, 0], sizes, df["ipc"], decimals=3)

    axs[1, 1].plot(
        sizes,
        df["peak_heap_mb"],
        marker="o",
        lw=2.4,
        color="#70AD47",
        label="Massif (medido)",
    )
    axs[1, 1].plot(
        sizes,
        df["theoretical_heap_mb"],
        marker="s",
        lw=2.0,
        ls="--",
        color="#7F7F7F",
        label="Modelo teórico",
    )
    axs[1, 1].set_title("Memoria pico heap")
    axs[1, 1].set_xlabel("N")
    axs[1, 1].set_ylabel("MB")
    axs[1, 1].legend()

    for ax in axs.flat:
        ax.set_xticks(sizes)
        ax.set_xticklabels([f"N={n}" for n in sizes])

    fig.suptitle("Resumen de rendimiento y memoria", fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "01_performance_summary.png")


def chart_miss_rates(df: pd.DataFrame) -> str:
    sizes = df["matrix_size"].astype(int).tolist()
    fig, axs = plt.subplots(2, 2, figsize=(12, 8))

    charts = [
        ("cache_miss_pct", "Cache misses", "#C00000"),
        ("l1_miss_pct", "L1-D miss rate", "#ED7D31"),
        ("dtlb_miss_pct", "dTLB miss rate", "#2F5597"),
        ("branch_miss_pct", "Branch miss rate", "#9E480E"),
    ]

    for ax, (col, title, color) in zip(axs.flat, charts):
        if col in df.columns:
            ax.plot(sizes, df[col], marker="o", lw=2.2, color=color)
            annotate_points(ax, sizes, df[col], decimals=2)
        ax.set_title(title)
        ax.set_xlabel("N")
        ax.set_ylabel("%")
        ax.set_xticks(sizes)
        ax.set_xticklabels([f"N={n}" for n in sizes])

    fig.suptitle("Comportamiento de misses y predictor", fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "02_miss_rates.png")


def chart_counters(df: pd.DataFrame) -> str:
    sizes = df["matrix_size"].astype(int).tolist()
    fig, axs = plt.subplots(2, 2, figsize=(12, 8))

    if "instructions" in df.columns:
        axs[0, 0].plot(
            sizes, df["instructions"] / 1e9, marker="o", lw=2.3, color="#2E75B6"
        )
        axs[0, 0].set_title("Instructions")
        axs[0, 0].set_ylabel("Billions")

    if "cycles" in df.columns:
        axs[0, 1].plot(
            sizes, df["cycles"] / 1e9, marker="o", lw=2.3, color="#7030A0"
        )
        axs[0, 1].set_title("Cycles")
        axs[0, 1].set_ylabel("Billions")

    if "cache_refs" in df.columns:
        axs[1, 0].plot(
            sizes, df["cache_refs"] / 1e9, marker="o", lw=2.3, color="#C00000"
        )
        axs[1, 0].set_title("Cache references")
        axs[1, 0].set_ylabel("Billions")

    if "ops_per_instruction" in df.columns:
        axs[1, 1].plot(
            sizes,
            df["ops_per_instruction"],
            marker="o",
            lw=2.3,
            color="#70AD47",
        )
        axs[1, 1].set_title("Operaciones / instrucción")
        axs[1, 1].set_ylabel("ops/inst")

    for ax in axs.flat:
        ax.set_xlabel("N")
        ax.set_xticks(sizes)
        ax.set_xticklabels([f"N={n}" for n in sizes])

    fig.suptitle("Volumen de trabajo y eficiencia derivada", fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "03_counter_trends.png")


def chart_timing_variability(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["matrix_size"].max())
    timing = raw.get(largest_n, {}).get("timing", {})
    values = timing.get("values", [])

    fig, axs = plt.subplots(1, 2, figsize=(12, 4.8))

    if values:
        xs = list(range(1, len(values) + 1))
        axs[0].plot(xs, values, marker="o", lw=2.1, color="#2E75B6")
        axs[0].axhline(statistics.mean(values), color="#C00000", lw=1.2, ls="--")
        axs[0].set_title(f"Corridas de tiempo (N={largest_n})")
        axs[0].set_xlabel("Run")
        axs[0].set_ylabel("ms")

        axs[1].boxplot(values, vert=True, patch_artist=True, boxprops=dict(facecolor="#DCE6F1"))
        axs[1].scatter([1] * len(values), values, color="#2E75B6", alpha=0.75)
        axs[1].set_title("Dispersión temporal")
        axs[1].set_ylabel("ms")
        axs[1].set_xticks([1])
        axs[1].set_xticklabels([f"N={largest_n}"])
    else:
        for ax in axs:
            ax.text(0.5, 0.5, "Sin timing_runs.txt", ha="center", va="center", fontsize=12)
            ax.axis("off")

    fig.suptitle("Variabilidad de las corridas", fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "04_timing_variability.png")

def chart_mem_access(raw: dict[int, dict[str, Any]]) -> str | None:
    sizes = sorted(raw.keys())
    # Canonical order for memory hierarchy levels
    level_order = ["L1 hit", "L2 hit", "L3 hit", "Local RAM", "Remote RAM"]
    color_map = {
        "L1 hit":    "#2E75B6",
        "L2 hit":    "#70AD47",
        "L3 hit":    "#ED7D31",
        "Local RAM": "#C00000",
        "Remote RAM":"#7030A0",
    }

    # Collect all labels present across all sizes (excluding N/A)
    all_labels: list[str] = []
    valid_sizes = []
    size_na: dict[int, float] = {}

    for n in sizes:
        mem = raw[n].get("perf_mem", {})
        if not mem.get("available"):
            continue
        valid_sizes.append(n)
        size_na[n] = mem.get("na_pct", 0.0)
        for s in mem.get("sources_classified", []):
            if s["label"] not in all_labels:
                all_labels.append(s["label"])

    if not valid_sizes:
        return None

    # Sort labels by canonical order, then alphabetically for unknowns
    def label_rank(lbl):
        try:
            return level_order.index(lbl)
        except ValueError:
            return len(level_order)

    all_labels = sorted(all_labels, key=label_rank)

    # Build matrix: rows=sizes, cols=labels
    data: dict[str, list[float]] = {lbl: [] for lbl in all_labels}
    for n in valid_sizes:
        sources = {
            s["label"]: s["pct_normalized"]
            for s in raw[n].get("perf_mem", {}).get("sources_classified", [])
        }
        for lbl in all_labels:
            data[lbl].append(sources.get(lbl, 0.0))

    fig, ax = plt.subplots(figsize=(11, max(4, len(valid_sizes) * 1.5)))
    y_labels = [f"N={n}" for n in valid_sizes]
    bottoms = [0.0] * len(valid_sizes)
    default_colors = ["#2E75B6", "#70AD47", "#ED7D31", "#C00000", "#7030A0", "#1F9EB7"]

    for i, lbl in enumerate(all_labels):
        values = data[lbl]
        color = color_map.get(lbl, default_colors[i % len(default_colors)])
        bars = ax.barh(y_labels, values, left=bottoms, label=lbl,
                       color=color, height=0.55)
        for bar, val in zip(bars, values):
            if val >= 5.0:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%",
                    ha="center", va="center",
                    fontsize=8.5, color="white", fontweight="bold",
                )
        bottoms = [b + v for b, v in zip(bottoms, values)]

    # Annotate N/A percentage per size
    for i, n in enumerate(valid_sizes):
        na = size_na.get(n, 0.0)
        if na > 0:
            ax.text(
                102, i,
                f"N/A: {na:.1f}%",
                va="center", fontsize=8, color="#808080",
            )

    ax.set_xlabel("% de accesos clasificados (excluye N/A de IBS)")
    ax.set_xlim(0, 118)
    ax.set_title("Distribución de accesos a memoria por nivel de jerarquía (perf mem / IBS)")
    ax.legend(loc="lower right", fontsize=9, title="Nivel")
    fig.tight_layout()
    return save_figure(fig, "06_mem_access_distribution.png")


def _barh(ax, rows: list[dict[str, Any]], label_key: str, value_key: str, title: str, color: str):
    if not rows:
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center", fontsize=11)
        ax.axis("off")
        return

    rows = rows[:5]
    labels = [str(r[label_key])[:34] for r in rows][::-1]
    values = [float(r[value_key]) for r in rows][::-1]

    bars = ax.barh(labels, values, color=color)
    for bar, val in zip(bars, values):
        ax.text(bar.get_width() + max(values) * 0.02, bar.get_y() + bar.get_height() / 2, f"{val:.2f}", va="center", fontsize=8)
    ax.set_title(title)
    ax.set_xlabel("%")


def chart_hotspots(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> str:
    largest_n = int(df["matrix_size"].max())
    gprof_rows = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_rows = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    cache_rows = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])

    fig, axs = plt.subplots(1, 3, figsize=(16, 5.6))

    _barh(axs[0], gprof_rows, "name", "pct_time", f"gprof top functions (N={largest_n})", "#C00000")
    _barh(axs[1], perf_rows, "symbol", "overhead_pct", "perf record hotspots", "#2E75B6")
    _barh(axs[2], cache_rows, "name", "pct", "cachegrind instruction share", "#70AD47")

    fig.suptitle("Hotspots del caso más grande", fontsize=14, fontweight="bold")
    fig.tight_layout()
    return save_figure(fig, "05_hotspots_largest_n.png")


def build_chart_pack(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> list[str]:
    os.makedirs(CHARTS_DIR, exist_ok=True)
    paths = [
        chart_performance(df),
        chart_miss_rates(df),
        chart_counters(df),
        chart_timing_variability(df, raw),
        chart_hotspots(df, raw),
        chart_mem_access(raw),
    ]
    return [p for p in paths if p and os.path.exists(p)]


def build_overview_rows(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> list[tuple[str, str, str]]:
    largest_n = int(df["matrix_size"].max())
    row_max_gflops = df.loc[df["gflops"].idxmax()] if df["gflops"].notna().any() else None
    row_max_ipc = df.loc[df["ipc"].idxmax()] if df["ipc"].notna().any() else None
    row_max_heap = df.loc[df["peak_heap_mb"].idxmax()] if df["peak_heap_mb"].notna().any() else None
    row_max_l1 = df.loc[df["l1_miss_pct"].idxmax()] if "l1_miss_pct" in df.columns and df["l1_miss_pct"].notna().any() else None

    timing = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    cache_top = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])
    massif = raw.get(largest_n, {}).get("massif", {})

    rows = []

    if row_max_gflops is not None:
        rows.append(
            (
                "Mejor rendimiento",
                f"N={int(row_max_gflops['matrix_size'])}",
                f"{fmt_num(row_max_gflops['gflops'], 4)} GFLOPS",
            )
        )
    if row_max_ipc is not None:
        rows.append(
            (
                "Mejor IPC",
                f"N={int(row_max_ipc['matrix_size'])}",
                f"{fmt_num(row_max_ipc['ipc'], 4)} ({classify_ipc(float(row_max_ipc['ipc']))})",
            )
        )
    if row_max_heap is not None:
        rows.append(
            (
                "Mayor heap pico",
                f"N={int(row_max_heap['matrix_size'])}",
                f"{fmt_num(row_max_heap['peak_heap_mb'], 3)} MB",
            )
        )
    if row_max_l1 is not None:
        rows.append(
            (
                "Peor L1 miss rate",
                f"N={int(row_max_l1['matrix_size'])}",
                f"{fmt_pct(row_max_l1['l1_miss_pct'], 2)}",
            )
        )

    if timing.get("count", 0) > 0:
        rows.append(
            (
                f"Variabilidad temporal N={largest_n}",
                f"{timing['count']} runs",
                f"CV={fmt_pct(timing['cv_pct'], 2)} | rango={fmt_num(timing['min_ms'], 3)}–{fmt_num(timing['max_ms'], 3)} ms",
            )
        )

    if gprof_top:
        rows.append(
            (
                f"Hotspot gprof N={largest_n}",
                gprof_top[0]["name"],
                f"{fmt_pct(gprof_top[0]['pct_time'], 2)} del tiempo de CPU",
            )
        )

    if perf_top:
        rows.append(
            (
                f"Hotspot perf record N={largest_n}",
                perf_top[0]["symbol"],
                f"{fmt_pct(perf_top[0]['overhead_pct'], 2)} de overhead muestral",
            )
        )

    if cache_top:
        rows.append(
            (
                f"Hotspot cachegrind N={largest_n}",
                cache_top[0]["name"],
                f"{fmt_pct(cache_top[0]['pct'], 2)} de instrucciones",
            )
        )

    if massif:
        rows.append(
            (
                f"Massif peak N={largest_n}",
                f"{fmt_num(massif.get('peak_total_mb'), 3)} MB",
                f"heap útil={fmt_pct(massif.get('peak_useful_pct'), 2)}",
            )
        )

    return rows


def write_overview_sheet(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Overview")
    title_row(ws, "Executive Profiling Report | Sequential Matrix Multiplication", 4)

    for col, width in enumerate([28, 22, 26, 58], 1):
        set_width(ws, col, width)

    header_cell(ws.cell(3, 1), "Indicador", "blue")
    header_cell(ws.cell(3, 2), "Contexto", "blue")
    header_cell(ws.cell(3, 3), "Valor", "blue")
    header_cell(ws.cell(3, 4), "Interpretación", "blue")
    ws.freeze_panes = "A4"

    rows = build_overview_rows(df, raw)
    r = 4
    for label, context, value in rows:
        bg = "alt" if r % 2 == 0 else "white"
        note = ""
        if "IPC" in label:
            note = "Refleja cuánta instrucción útil se retira por ciclo."
        elif "L1 miss" in label:
            note = "Captura presión de localidad en accesos de datos."
        elif "heap" in label.lower():
            note = "Contrasta costo de reserva dinámica frente al modelo teórico."
        elif "gprof" in label.lower() or "perf" in label.lower() or "cachegrind" in label.lower():
            note = "Señala el punto dominante donde conviene optimizar primero."
        elif "Variabilidad" in label:
            note = "CV bajo indica estabilidad experimental entre ejecuciones."
        elif "rendimiento" in label.lower():
            note = "Relaciona 2·N³ operaciones con el tiempo medio medido."

        data_cell(ws.cell(r, 1), label, bg=bg, align="left", bold=True)
        data_cell(ws.cell(r, 2), context, bg=bg, align="center")
        data_cell(ws.cell(r, 3), value, bg=bg, align="center")
        data_cell(ws.cell(r, 4), note, bg=bg, align="left")
        r += 1

    r += 1
    title_row(ws, "Per-size quick view", 8, row=r)
    r += 1

    cols = [
        "N",
        "Time mean (ms)",
        "Std (ms)",
        "GFLOPS",
        "IPC",
        "L1 miss %",
        "Cache miss %",
        "Heap peak (MB)",
    ]
    for i, c in enumerate(cols, 1):
        header_cell(ws.cell(r, i), c, "cyan")
    r += 1

    for _, row in df.iterrows():
        bg = "alt" if r % 2 == 0 else "white"
        values = [
            int(row["matrix_size"]),
            safe_float(row.get("time_mean_ms")),
            safe_float(row.get("time_std_ms")),
            safe_float(row.get("gflops")),
            safe_float(row.get("ipc")),
            safe_float(row.get("l1_miss_pct")),
            safe_float(row.get("cache_miss_pct")),
            safe_float(row.get("peak_heap_mb")),
        ]
        fmts = ["0", "0.000", "0.000", "0.0000", "0.0000", "0.00", "0.00", "0.000"]
        for cidx, (v, f) in enumerate(zip(values, fmts), 1):
            align = "center" if cidx == 1 else "right"
            data_cell(ws.cell(r, cidx), v, fmt=f, bg=bg, align=align)
        r += 1


def write_metrics_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Metrics")
    title_row(ws, "All Metrics from data_profiling.csv", len(df.columns))
    ws.freeze_panes = "A3"

    for cidx, col in enumerate(df.columns, 1):
        header_cell(ws.cell(2, cidx), col, "blue")
        set_width(ws, cidx, max(12, min(22, len(col) + 2)))

    for ridx, row in enumerate(df.itertuples(index=False), 3):
        bg = "alt" if ridx % 2 == 0 else "white"
        for cidx, col in enumerate(df.columns, 1):
            val = getattr(row, col)
            if col == "matrix_size":
                data_cell(ws.cell(ridx, cidx), int(val), fmt="0", bg=bg, align="center")
            elif isinstance(val, (float, int)) and not pd.isna(val):
                fmt = "0.0000"
                if "pct" in col:
                    fmt = "0.00"
                elif "instructions" in col or "cycles" in col or "refs" in col or "misses" in col or "operations" in col:
                    fmt = "#,##0"
                elif "heap" in col or "time" in col:
                    fmt = "0.000"
                data_cell(ws.cell(ridx, cidx), float(val), fmt=fmt, bg=bg)
            else:
                data_cell(ws.cell(ridx, cidx), "—", bg=bg, align="center")


def write_raw_summary_sheet(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Raw Summary")
    cols = [
        "N",
        "Timing runs",
        "Timing mean (ms)",
        "Timing std (ms)",
        "Timing CV %",
        "gprof top",
        "gprof %",
        "perf top",
        "perf %",
        "cachegrind top",
        "cachegrind %",
        "cachegrind total Ir",
        "massif peak MB",
        "massif useful %",
        "Top allocator",
        "LLC counters",
    ]
    title_row(ws, "Summaries parsed from raw reports", len(cols))
    ws.freeze_panes = "A3"
    widths = [8, 12, 16, 14, 12, 26, 10, 28, 10, 34, 12, 18, 16, 14, 30, 14]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        header_cell(ws.cell(2, i), c, "blue")
        set_width(ws, i, w)

    r = 3
    for n in df["matrix_size"].astype(int).tolist():
        entry = raw.get(n, {})
        timing = entry.get("timing", {})
        gprof = entry.get("gprof", {}).get("top", [])
        perf = entry.get("perf_record", {}).get("top", [])
        cache = entry.get("cachegrind", {})
        massif = entry.get("massif", {})
        perf_txt = entry.get("perf_stat_text", "")

        bg = "alt" if r % 2 == 0 else "white"

        gprof_name = gprof[0]["name"] if gprof else "—"
        gprof_pct = gprof[0]["pct_time"] if gprof else math.nan
        perf_name = perf[0]["symbol"] if perf else "—"
        perf_pct = perf[0]["overhead_pct"] if perf else math.nan

        cache_top = cache.get("top_functions", [])
        cache_name = cache_top[0]["name"] if cache_top else "—"
        cache_pct = cache_top[0]["pct"] if cache_top else math.nan

        allocs = massif.get("allocators", [])
        top_alloc = allocs[0]["where"] if allocs else "—"
        llc_status = "supported" if perf_llc_supported(perf_txt) else "not supported"

        values = [
            n,
            timing.get("count", 0),
            timing.get("mean_ms", math.nan),
            timing.get("std_ms", math.nan),
            timing.get("cv_pct", math.nan),
            gprof_name,
            gprof_pct,
            perf_name,
            perf_pct,
            cache_name,
            cache_pct,
            cache.get("total_ir", 0),
            massif.get("peak_total_mb", math.nan),
            massif.get("peak_useful_pct", math.nan),
            top_alloc,
            llc_status,
        ]

        for cidx, val in enumerate(values, 1):
            cell = ws.cell(r, cidx)
            if cidx in (1, 2):
                data_cell(cell, val, fmt="0", bg=bg, align="center")
            elif cidx in (3, 4, 13):
                data_cell(cell, val, fmt="0.000", bg=bg)
            elif cidx in (5, 7, 9, 10, 14):
                data_cell(cell, val, fmt="0.00", bg=bg)
            elif cidx == 12:
                data_cell(cell, val, fmt="#,##0", bg=bg)
            else:
                data_cell(cell, val, bg=bg, align="left" if isinstance(val, str) else "right")
        r += 1


def write_hot_lines_sheet(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Hot Lines")
    title_row(ws, "Top annotated lines from cachegrind", 4)
    ws.freeze_panes = "A3"
    headers = ["N", "Instruction refs", "Share %", "Source line"]
    widths = [10, 18, 12, 95]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws.cell(2, i), h, "blue")
        set_width(ws, i, w)

    r = 3
    for n in df["matrix_size"].astype(int).tolist():
        hot_lines = raw.get(n, {}).get("cachegrind", {}).get("hot_lines", [])
        if not hot_lines:
            bg = "alt" if r % 2 == 0 else "white"
            data_cell(ws.cell(r, 1), n, fmt="0", bg=bg, align="center")
            data_cell(ws.cell(r, 2), "—", bg=bg, align="center")
            data_cell(ws.cell(r, 3), "—", bg=bg, align="center")
            data_cell(ws.cell(r, 4), "Sin líneas anotadas", bg=bg, align="left")
            r += 1
            continue

        for item in hot_lines:
            bg = "alt" if r % 2 == 0 else "white"
            data_cell(ws.cell(r, 1), n, fmt="0", bg=bg, align="center")
            data_cell(ws.cell(r, 2), item["ir"], fmt="#,##0", bg=bg)
            data_cell(ws.cell(r, 3), item["pct"], fmt="0.00", bg=bg)
            data_cell(ws.cell(r, 4), item["code"], bg=bg, align="left")
            r += 1


def write_charts_sheet(wb: Workbook, chart_paths: list[str]) -> None:
    ws = wb.create_sheet("Charts")
    title_row(ws, "Visual report", 20)
    anchors = ["A3", "L3", "A28", "L28", "A53", "L53"]
    for anchor, path in zip(anchors, chart_paths):
        if not os.path.exists(path):
            continue
        img = XLImage(path)
        img.width = 760
        img.height = 430
        ws.add_image(img, anchor)


def write_notes_sheet(wb: Workbook, df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Analysis Notes")
    title_row(ws, "Auto-generated interpretation notes", 3)
    for i, w in enumerate([28, 22, 82], 1):
        set_width(ws, i, w)

    header_cell(ws.cell(2, 1), "Tema", "blue")
    header_cell(ws.cell(2, 2), "Contexto", "blue")
    header_cell(ws.cell(2, 3), "Comentario", "blue")

    notes = []
    largest_n = int(df["matrix_size"].max())
    last = df[df["matrix_size"] == largest_n].iloc[0]

    notes.append((
        "Escalamiento",
        "Tiempo y GFLOPS",
        "Compara el crecimiento de tiempo con el de GFLOPS para detectar si el incremento de tamaño favorece el aprovechamiento del pipeline o expone aún más el cuello de botella de memoria."
    ))
    notes.append((
        "IPC",
        f"N={largest_n}",
        f"IPC={fmt_num(last.get('ipc'), 4)} → clasificación: {classify_ipc(safe_float(last.get('ipc')))}."
    ))
    if "cache_miss_pct" in df.columns:
        notes.append((
            "Cache misses",
            f"N={largest_n}",
            f"Cache miss global={fmt_pct(last.get('cache_miss_pct'), 2)}; interpretación={classify_miss(safe_float(last.get('cache_miss_pct')))}."
        ))
    if "l1_miss_pct" in df.columns:
        notes.append((
            "L1-D",
            f"N={largest_n}",
            f"L1 miss={fmt_pct(last.get('l1_miss_pct'), 2)}; este valor suele ser el mejor indicador de localidad efectiva del kernel i-j-k."
        ))
    if "dtlb_miss_pct" in df.columns:
        notes.append((
            "dTLB",
            f"N={largest_n}",
            f"dTLB miss={fmt_pct(last.get('dtlb_miss_pct'), 4)}; ayuda a separar problemas de traslación de problemas puramente de cache."
        ))

    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    if gprof_top:
        notes.append((
            "gprof",
            f"N={largest_n}",
            f"La función dominante es {gprof_top[0]['name']} con {fmt_pct(gprof_top[0]['pct_time'], 2)} del tiempo muestral."
        ))

    perf_top = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    if perf_top:
        notes.append((
            "perf record",
            f"N={largest_n}",
            f"El símbolo más pesado es {perf_top[0]['symbol']} con {fmt_pct(perf_top[0]['overhead_pct'], 2)} de overhead."
        ))

    cache_top = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])
    if cache_top:
        notes.append((
            "cachegrind",
            f"N={largest_n}",
            f"La distribución de instrucciones también concentra el costo en {cache_top[0]['name']} con {fmt_pct(cache_top[0]['pct'], 2)}."
        ))

    massif = raw.get(largest_n, {}).get("massif", {})
    if massif:
        notes.append((
            "Massif",
            f"N={largest_n}",
            f"Peak heap={fmt_num(massif.get('peak_total_mb'), 3)} MB; heap útil={fmt_pct(massif.get('peak_useful_pct'), 2)}."
        ))

    timing = raw.get(largest_n, {}).get("timing", {})
    if timing.get("count", 0):
        notes.append((
            "Estabilidad",
            f"N={largest_n}",
            f"Media={fmt_num(timing.get('mean_ms'), 3)} ms, std={fmt_num(timing.get('std_ms'), 3)} ms, CV={fmt_pct(timing.get('cv_pct'), 2)}."
        ))
        
    for n in df["matrix_size"].astype(int).tolist():
        mem_data = raw.get(n, {}).get("perf_mem", {})
        if not mem_data.get("available"):
            continue
        sources = mem_data["sources"]
        # Find RAM and L1 shares for this N
        ram_pct  = next((s["pct"] for s in sources if "RAM"  in s["label"]), None)
        l1_pct   = next((s["pct"] for s in sources if "L1"   in s["label"]), None)
        lfb_pct  = next((s["pct"] for s in sources if "LFB"  in s["label"]), None)
        summary  = ", ".join(f"{s['label']}={s['pct']:.1f}%" for s in sources)
        notes.append((
            "perf mem",
            f"N={n}",
            f"Distribución de accesos: {summary}."
            + (f" RAM={ram_pct:.1f}% indica presión alta sobre memoria principal." if ram_pct and ram_pct > 10 else ""),
        ))

    r = 3
    for topic, context, note in notes:
        bg = "alt" if r % 2 == 0 else "white"
        data_cell(ws.cell(r, 1), topic, bg=bg, align="left", bold=True)
        data_cell(ws.cell(r, 2), context, bg=bg, align="center")
        data_cell(ws.cell(r, 3), note, bg=bg, align="left")
        r += 1


def build_workbook(df: pd.DataFrame, raw: dict[int, dict[str, Any]], chart_paths: list[str]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_overview_sheet(wb, df, raw)
    write_metrics_sheet(wb, df)
    write_raw_summary_sheet(wb, df, raw)
    write_hot_lines_sheet(wb, df, raw)
    write_notes_sheet(wb, df, raw)
    write_charts_sheet(wb, chart_paths)

    wb.save(XLSX_PATH)


def html_card(label: str, value: str, note: str = "") -> str:
    return f"""
    <div class="card kpi">
      <div class="kpi-label">{html.escape(label)}</div>
      <div class="kpi-value">{html.escape(value)}</div>
      <div class="kpi-note">{html.escape(note)}</div>
    </div>
    """


def relative_chart_path(path: str) -> str:
    return os.path.relpath(path, OUT_DIR).replace(os.sep, "/")


def html_table_from_df(df: pd.DataFrame, cols: list[str]) -> str:
    headers = "".join(f"<th>{html.escape(c)}</th>" for c in cols)
    rows_html = []
    for _, r in df.iterrows():
        tds = []
        for c in cols:
            v = r.get(c, math.nan)
            if c == "matrix_size":
                tds.append(f"<td>N={int(v)}</td>")
            elif "pct" in c:
                tds.append(f"<td>{fmt_pct(v, 2)}</td>")
            elif c in ("instructions", "cycles", "cache_refs", "cache_misses", "l1_loads", "l1_misses", "llc_loads", "llc_misses", "dtlb_loads", "dtlb_misses", "branch_instructions", "branch_misses"):
                tds.append(f"<td>{fmt_int(v)}</td>")
            elif "heap" in c or "time" in c:
                tds.append(f"<td>{fmt_num(v, 3)}</td>")
            else:
                tds.append(f"<td>{fmt_num(v, 4)}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    return f"""
    <table>
      <thead><tr>{headers}</tr></thead>
      <tbody>{''.join(rows_html)}</tbody>
    </table>
    """


def html_hotspot_table(title: str, rows: list[dict[str, Any]], name_key: str, value_key: str) -> str:
    if not rows:
        return f"""
        <div class="card">
          <h3>{html.escape(title)}</h3>
          <p>Sin datos.</p>
        </div>
        """
    body = []
    for row in rows[:8]:
        name = html.escape(str(row.get(name_key, "—")))
        value = row.get(value_key, math.nan)
        body.append(f"<tr><td>{name}</td><td>{fmt_pct(value, 2)}</td></tr>")
    return f"""
    <div class="card">
      <h3>{html.escape(title)}</h3>
      <table>
        <thead><tr><th>Elemento</th><th>Share</th></tr></thead>
        <tbody>{''.join(body)}</tbody>
      </table>
    </div>
    """


def build_html_report(df: pd.DataFrame, raw: dict[int, dict[str, Any]], chart_paths: list[str]) -> None:
    largest_n = int(df["matrix_size"].max())
    largest = df[df["matrix_size"] == largest_n].iloc[0]

    best_gflops = df.loc[df["gflops"].idxmax()] if df["gflops"].notna().any() else None
    best_ipc = df.loc[df["ipc"].idxmax()] if df["ipc"].notna().any() else None
    max_heap = df.loc[df["peak_heap_mb"].idxmax()] if df["peak_heap_mb"].notna().any() else None

    timing = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    cache_top = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])
    hot_lines = raw.get(largest_n, {}).get("cachegrind", {}).get("hot_lines", [])
    massif = raw.get(largest_n, {}).get("massif", {})
    
    def _html_mem_table(raw_all: dict[int, dict]) -> str:
        rows_html = []
        for n in sorted(raw_all.keys()):
            sources = raw_all[n].get("perf_mem", {}).get("sources", [])
            if not sources:
                continue
            for s in sources:
                rows_html.append(
                    f"<tr><td>N={n}</td>"
                    f"<td>{html.escape(s['label'])}</td>"
                    f"<td>{s['pct']:.2f}%</td></tr>"
                )
        if not rows_html:
            return ""
        return f"""
        <div class="card">
        <h3>perf mem — distribución de accesos a memoria</h3>
        <table>
            <thead><tr><th>N</th><th>Nivel</th><th>Share</th></tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
        </table>
        </div>
        """

    mem_table_html = _html_mem_table(raw)

    cards = []
    if best_gflops is not None:
        cards.append(html_card("Best GFLOPS", f"N={int(best_gflops['matrix_size'])}", f"{fmt_num(best_gflops['gflops'], 4)} GFLOPS"))
    if best_ipc is not None:
        cards.append(html_card("Best IPC", f"N={int(best_ipc['matrix_size'])}", f"{fmt_num(best_ipc['ipc'], 4)}"))
    if max_heap is not None:
        cards.append(html_card("Peak heap", f"N={int(max_heap['matrix_size'])}", f"{fmt_num(max_heap['peak_heap_mb'], 3)} MB"))
    cards.append(html_card(f"N={largest_n}", "Largest case", f"time={fmt_num(largest['time_mean_ms'], 3)} ms | IPC={fmt_num(largest['ipc'], 4)}"))

    chart_imgs = []
    for p in chart_paths:
        rel = relative_chart_path(p)
        chart_imgs.append(
            f"""
            <figure class="card chart-card">
              <img src="{html.escape(rel)}" alt="{html.escape(os.path.basename(p))}" loading="lazy">
            </figure>
            """
        )

    hot_lines_html = ""
    if hot_lines:
        rows = "".join(
            f"<tr><td>{fmt_int(x['ir'])}</td><td>{fmt_pct(x['pct'], 2)}</td><td><code>{html.escape(x['code'])}</code></td></tr>"
            for x in hot_lines
        )
        hot_lines_html = f"""
        <div class="card">
          <h3>Cachegrind hottest source lines</h3>
          <table>
            <thead><tr><th>Ir</th><th>Share</th><th>Source</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>
        """

    allocators_html = ""
    if massif.get("allocators"):
        rows = "".join(
            f"<tr><td>{html.escape(a['where'])}</td><td>{fmt_pct(a['pct'], 2)}</td><td>{fmt_int(a['bytes'])} B</td></tr>"
            for a in massif["allocators"][:8]
        )
        allocators_html = f"""
        <div class="card">
          <h3>Massif allocation breakdown</h3>
          <table>
            <thead><tr><th>Allocator</th><th>Share</th><th>Bytes</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>
        """

    metrics_table = html_table_from_df(
        df,
        [
            "matrix_size",
            "time_mean_ms",
            "time_std_ms",
            "gflops",
            "ipc",
            "cache_miss_pct",
            "l1_miss_pct",
            "dtlb_miss_pct",
            "branch_miss_pct",
            "peak_heap_mb",
        ],
    )

    html_doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Profiling Report</title>
<style>
:root {{
  --bg: #f4f7fb;
  --surface: #ffffff;
  --surface-2: #f9fbfd;
  --text: #1f2a36;
  --muted: #5d6b79;
  --border: #d8e1ea;
  --primary: #1f3557;
  --accent: #2e75b6;
  --good: #2f6b1e;
  --shadow: 0 10px 30px rgba(31,53,87,.08);
}}
* {{ box-sizing: border-box; }}
body {{
  margin: 0;
  font-family: "Inter", "Segoe UI", Arial, sans-serif;
  color: var(--text);
  background: linear-gradient(180deg, #f8fbff 0%, var(--bg) 100%);
  line-height: 1.55;
}}
.container {{
  width: min(1380px, 94vw);
  margin: 0 auto;
  padding: 32px 0 48px;
}}
.hero {{
  background: linear-gradient(135deg, #1f3557 0%, #2e75b6 100%);
  color: white;
  border-radius: 22px;
  padding: 36px;
  box-shadow: var(--shadow);
  margin-bottom: 26px;
}}
.hero h1 {{
  margin: 0 0 8px;
  font-size: 2.1rem;
}}
.hero p {{
  margin: 0;
  color: rgba(255,255,255,.88);
  max-width: 82ch;
}}
.grid {{
  display: grid;
  gap: 18px;
}}
.kpis {{
  grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
  margin-bottom: 24px;
}}
.two {{
  grid-template-columns: repeat(auto-fit, minmax(420px, 1fr));
  margin-bottom: 24px;
}}
.charts {{
  grid-template-columns: repeat(auto-fit, minmax(480px, 1fr));
}}
.card {{
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: 18px;
  box-shadow: var(--shadow);
  padding: 20px;
}}
.kpi {{
  padding: 18px 20px;
}}
.kpi-label {{
  font-size: .8rem;
  text-transform: uppercase;
  letter-spacing: .08em;
  color: var(--muted);
  margin-bottom: 8px;
}}
.kpi-value {{
  font-size: 1.4rem;
  font-weight: 700;
  color: var(--primary);
  margin-bottom: 4px;
}}
.kpi-note {{
  color: var(--muted);
  font-size: .94rem;
}}
h2, h3 {{
  margin-top: 0;
  color: var(--primary);
}}
table {{
  width: 100%;
  border-collapse: collapse;
  font-size: .94rem;
}}
th, td {{
  padding: 10px 12px;
  border-bottom: 1px solid var(--border);
  text-align: left;
  vertical-align: top;
}}
th {{
  background: var(--surface-2);
  color: var(--primary);
}}
.chart-card img {{
  width: 100%;
  height: auto;
  border-radius: 12px;
  display: block;
}}
code {{
  font-family: "JetBrains Mono", "Consolas", monospace;
  white-space: pre-wrap;
}}
.section {{
  margin-top: 26px;
}}
.meta {{
  display: inline-flex;
  gap: 10px;
  flex-wrap: wrap;
  margin-top: 12px;
}}
.badge {{
  background: rgba(255,255,255,.14);
  border: 1px solid rgba(255,255,255,.18);
  padding: 6px 10px;
  border-radius: 999px;
  font-size: .9rem;
}}
@media (max-width: 720px) {{
  .hero {{
    padding: 24px;
  }}
  .hero h1 {{
    font-size: 1.6rem;
  }}
}}
</style>
</head>
<body>
  <div class="container">
    <section class="hero">
      <h1>Sequential Matrix Multiplication — Profiling Report</h1>
      <p>Professional report generated from CSV metrics plus raw reports from gprof, perf stat, perf record, cachegrind, massif, and timing runs.</p>
      <div class="meta">
        <span class="badge">Cases: {", ".join(f"N={int(n)}" for n in df["matrix_size"].tolist())}</span>
        <span class="badge">Largest N: {largest_n}</span>
        <span class="badge">Peak heap @ largest N: {fmt_num(massif.get('peak_total_mb'), 3)} MB</span>
        <span class="badge">Timing CV @ largest N: {fmt_pct(timing.get('cv_pct'), 2)}</span>
      </div>
    </section>

    <section class="grid kpis">
      {''.join(cards)}
    </section>

    <section class="section card">
      <h2>Trend table</h2>
      {metrics_table}
    </section>

    <section class="section grid two">
      {html_hotspot_table("gprof dominant functions", gprof_top, "name", "pct_time")}
      {html_hotspot_table("perf record dominant symbols", perf_top, "symbol", "overhead_pct")}
      {html_hotspot_table("cachegrind dominant functions", cache_top, "name", "pct")}
      {allocators_html}
      {hot_lines_html}
      {mem_table_html}
    </section>

    <section class="section grid charts">
      {''.join(chart_imgs)}
    </section>
  </div>
</body>
</html>
"""
    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html_doc)


def build_latex_table(df: pd.DataFrame) -> None:
    lines = [
        r"\begin{table}[h]",
        r"\centering",
        r"\caption{Caracterización de la implementación secuencial de multiplicación de matrices}",
        r"\label{tab:seq_profiling}",
        r"\begin{tabular}{rrrrrrrrrr}",
        r"\toprule",
        r"$N$ & Time mean (ms) & Std (ms) & GFLOPS & IPC & Cache Miss \% & L1 Miss \% & dTLB Miss \% & Branch Miss \% & Peak Heap (MB) \\",
        r"\midrule",
    ]
    for _, r in df.iterrows():
        lines.append(
            f"{int(r['matrix_size'])} & "
            f"{safe_float(r.get('time_mean_ms')):.3f} & "
            f"{safe_float(r.get('time_std_ms')):.3f} & "
            f"{safe_float(r.get('gflops')):.4f} & "
            f"{safe_float(r.get('ipc')):.4f} & "
            f"{safe_float(r.get('cache_miss_pct')):.2f} & "
            f"{safe_float(r.get('l1_miss_pct')):.2f} & "
            f"{safe_float(r.get('dtlb_miss_pct')):.4f} & "
            f"{safe_float(r.get('branch_miss_pct')):.4f} & "
            f"{safe_float(r.get('peak_heap_mb')):.3f} \\\\"
        )
    lines += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    with open(LATEX_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def print_console_summary(df: pd.DataFrame, raw: dict[int, dict[str, Any]]) -> None:
    largest_n = int(df["matrix_size"].max())
    last = df[df["matrix_size"] == largest_n].iloc[0]
    timing = raw.get(largest_n, {}).get("timing", {})
    gprof_top = raw.get(largest_n, {}).get("gprof", {}).get("top", [])
    perf_top = raw.get(largest_n, {}).get("perf_record", {}).get("top", [])
    cache_top = raw.get(largest_n, {}).get("cachegrind", {}).get("top_functions", [])

    print("============================================================")
    print(" Profiling report generated")
    print("============================================================")
    print(f"CSV   : {CSV_PATH}")
    print(f"XLSX  : {XLSX_PATH}")
    print(f"HTML  : {HTML_PATH}")
    print(f"LaTeX : {LATEX_PATH}")
    print(f"Charts: {CHARTS_DIR}")
    print("------------------------------------------------------------")
    print(f"Largest N        : {largest_n}")
    print(f"Time mean (ms)   : {fmt_num(last.get('time_mean_ms'), 3)}")
    print(f"GFLOPS           : {fmt_num(last.get('gflops'), 4)}")
    print(f"IPC              : {fmt_num(last.get('ipc'), 4)}")
    print(f"Cache miss %     : {fmt_pct(last.get('cache_miss_pct'), 2)}")
    print(f"L1 miss %        : {fmt_pct(last.get('l1_miss_pct'), 2)}")
    print(f"Peak heap (MB)   : {fmt_num(last.get('peak_heap_mb'), 3)}")
    if timing.get("count", 0):
        print(f"Timing CV %      : {fmt_pct(timing.get('cv_pct'), 2)}")
    if gprof_top:
        print(f"gprof hotspot    : {gprof_top[0]['name']} ({fmt_pct(gprof_top[0]['pct_time'], 2)})")
    if perf_top:
        print(f"perf hotspot     : {perf_top[0]['symbol']} ({fmt_pct(perf_top[0]['overhead_pct'], 2)})")
    if cache_top:
        print(f"cachegrind top   : {cache_top[0]['name']} ({fmt_pct(cache_top[0]['pct'], 2)})")
    print("============================================================")


def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    df = normalize_dataframe(CSV_PATH)
    raw = collect_raw_analysis(df)

    chart_paths = build_chart_pack(df, raw)
    build_workbook(df, raw, chart_paths)
    build_html_report(df, raw, chart_paths)
    build_latex_table(df)
    print_console_summary(df, raw)


if __name__ == "__main__":
    main()