"""
report.py  --  HPC benchmark report (bench_final.sh real data).

Table layout (pivoted for readability):
  One block per implementation:
    Header  : impl label  (merged, dark)
    Sub-hdr : Rep | N=400 | N=800 | N=1600 | N=3200 | N=6400
    Rows    : 1..10  (individual measurements)
    Summary : Avg / StdDev / CV%
    Blank separator before next block

Speedup table below (one row per impl):
  Impl | N=400 avg | sp | N=800 avg | sp | ... | Avg Speedup

Usage:
    python report.py [results_dir/] [csv_src_dir/]
    Defaults: results_dir = results_final/  csv_src = same as results_dir
"""

from __future__ import annotations

import math
import os
import sys
from dataclasses import dataclass

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    Series,
    make_border, set_col_width, style_data_cell, style_header_cell,
    write_title_row, save_figure, plot_lines,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

RESULTS_DIR = sys.argv[1].rstrip("/") if len(sys.argv) > 1 else "results_final"
SRC_DIR     = sys.argv[2].rstrip("/") if len(sys.argv) > 2 else RESULTS_DIR
OUTPUT_PATH = os.path.join(RESULTS_DIR, "reporte_final.xlsx")
CHARTS_DIR  = os.path.join(RESULTS_DIR, "charts_final")
BEST_MARKER = "-march=native"

COLORS: dict[str, str] = {
    "seq_std/best":      "#FF6600",
    "seq_cache/noopt":   "#2E75B6",
    "seq_cache/best":    "#1F4E79",
    "threads_2t/noopt":  "#AAAAAA",
    "threads_4t/noopt":  "#70AD47",
    "threads_6t/noopt":  "#4472C4",
    "threads_8t/noopt":  "#ED7D31",
    "threads_12t/noopt": "#7030A0",
    "threads_2t/best":   "#AAAAAA",
    "threads_4t/best":   "#70AD47",
    "threads_6t/best":   "#4472C4",
    "threads_8t/best":   "#ED7D31",
    "threads_12t/best":  "#7030A0",
    "conc/noopt":        "#843C0C",
    "conc/best":         "#C00000",
}

# Extra colors not in report_utils
C_EXTRA: dict[str, str] = {
    "summary_bg":  "EBF3FB",   # light blue for Avg/Std/CV rows
    "summary_fg":  "1F4E79",
    "impl_header": "1F4E79",   # per-impl block header bg
    "sep":         "D9E1F2",   # subtle separator row
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

_OPT_SUFFIX   = "con O3_full"
_NOOPT_SUFFIX = "sin optimizar"


@dataclass(frozen=True)
class Row:
    impl:    str
    threads: int
    tag:     str

    @property
    def label(self) -> str:
        opt = _OPT_SUFFIX if self.tag == "best" else _NOOPT_SUFFIX
        if self.impl == "threads":
            return f"Concurrencia {self.threads} hilos {opt}"
        if self.impl == "seq_std":
            return f"Secuencial naive {opt}"
        if self.impl == "seq_cache":
            return f"Cache line {opt}"
        if self.impl == "conc":
            return f"Procesos (fork) {opt}"
        return f"{self.impl} {opt}"

    @property
    def short_label(self) -> str:
        if self.impl == "threads":
            return f"threads_{self.threads}t/{self.tag}"
        return f"{self.impl}/{self.tag}"

    @property
    def color(self) -> str:
        return COLORS.get(self.short_label, "#888888")


def tag_of(flags: str) -> str:
    return "best" if BEST_MARKER in flags else "noopt"


AllReps = dict[Row, dict[int, list[tuple[int, float]]]]

# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_csv(name: str) -> pd.DataFrame | None:
    for base in [SRC_DIR, RESULTS_DIR]:
        path = os.path.join(base, f"data_{name}.csv")
        if os.path.exists(path):
            df = pd.read_csv(path)
            df.columns         = [c.strip().lower() for c in df.columns]
            df["flags"]        = df["flags"].str.strip('"')
            df["impl"]         = df["impl"].str.strip()
            df["threads"]      = df["threads"].astype(int)
            df["matrix_size"]  = df["matrix_size"].astype(int)
            df["repetition"]   = df["repetition"].astype(int)
            df["wall_time_ms"] = df["wall_time_ms"].astype(float)
            df["tag"]          = df["flags"].apply(tag_of)
            print(f"  [OK]  {path}  ({len(df)} rows)")
            return df
    print(f"  [--]  data_{name}.csv not found")
    return None


def build_all_reps(df: pd.DataFrame) -> AllReps:
    result: AllReps = {}
    for key, grp in df.groupby(["impl", "threads", "tag"]):
        impl, threads, tag = key  # type: ignore[misc]
        row = Row(impl=str(impl), threads=int(str(threads)), tag=str(tag))
        result[row] = {}
        for size, sub in grp.groupby("matrix_size"):
            pairs = sorted(
                [(int(str(r)), float(v))
                 for r, v in zip(sub["repetition"], sub["wall_time_ms"])],
                key=lambda x: x[0],
            )
            result[row][int(str(size))] = pairs
    return result


def avgs(reps: AllReps) -> dict[Row, dict[int, float]]:
    return {
        row: {s: sum(v for _, v in pairs) / len(pairs)
              for s, pairs in sizes.items()}
        for row, sizes in reps.items()
    }


def best_thread(avg_data: dict[Row, dict], sizes: list[int],
                tag: str, ref: Row) -> Row | None:
    ref_avgs = avg_data.get(ref, {})
    if not ref_avgs:
        return None
    best_row, best_sp = None, 0.0
    for row, times in avg_data.items():
        if row.impl != "threads" or row.tag != tag:
            continue
        sp_vals = [ref_avgs[s] / times[s]
                   for s in sizes
                   if s in times and s in ref_avgs and times[s] > 0]
        if sp_vals:
            sp = sum(sp_vals) / len(sp_vals)
            if sp > best_sp:
                best_sp, best_row = sp, row
    return best_row

# ---------------------------------------------------------------------------
# Low-level cell helpers
# ---------------------------------------------------------------------------

def _thick_border() -> Border:
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin",   color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thin)

def _thick_bottom() -> Border:
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin",   color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thick)

def _thin_border() -> Border:
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, value: str, bg: str, fg: str = "FFFFFF",
         size: int = 10, bold: bool = True,
         align: str = "center") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    cell.border    = _thin_border()


def _dat(cell, value, fmt: str | None = None,
         bg: str = "FFFFFF", fg: str = "000000",
         bold: bool = False, align: str = "right") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt


def _summary_dat(cell, value, fmt: str | None = None,
                 bold: bool = False) -> None:
    _dat(cell, value, fmt=fmt,
         bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"], bold=bold)

# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------

def chart_time(rows: list[Row], avg_data: dict, sizes: list,
               title: str, fname: str) -> str:
    series = [Series(label=r.short_label, data=avg_data[r], color=r.color)
              for r in rows if r in avg_data and avg_data[r]]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        plot_lines(ax, series, log_scale=True)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Dimensión N", fontsize=10)
        ax.set_ylabel("Tiempo promedio (ms)", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)


def chart_speedup(rows: list[Row], avg_data: dict, ref: Row,
                  sizes: list, title: str, fname: str) -> str:
    ref_avgs = avg_data.get(ref, {})
    series   = [
        Series(
            label=r.short_label,
            data={s: ref_avgs[s] / avg_data[r][s]
                  for s in sizes
                  if s in avg_data.get(r, {}) and s in ref_avgs
                  and avg_data[r][s] > 0},
            color=r.color,
        )
        for r in rows if r != ref and r in avg_data
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--",
                   label=f"ref: {ref.short_label}")
        plot_lines(ax, [s for s in series if s.data], log_scale=False)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Dimensión N", fontsize=10)
        ax.set_ylabel("Speedup", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)

# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

def write_sheet(wb: Workbook, name: str, title: str,
                rows: list[Row], all_reps: AllReps,
                ref: Row, sizes: list[int],
                ct: str, cs: str) -> None:
    ws  = wb.create_sheet(name)
    avg_data = avgs(all_reps)
    n_reps   = max(
        (len(p) for row in rows for p in all_reps.get(row, {}).values()),
        default=10,
    )

    # Raw table: 1 + n_sizes + 1 cols  (Rep | N=400..N=6400 | blank spacer)
    N_COLS = 1 + len(sizes)

    # -----------------------------------------------------------------------
    # Title
    # -----------------------------------------------------------------------
    write_title_row(ws, title, N_COLS)
    ws.row_dimensions[1].height = 26

    # Fixed column widths: Rep col + one col per size
    set_col_width(ws, 1, 10)
    for ci, size in enumerate(sizes, 2):
        # wider for large numbers
        w = 10 if size <= 800 else 12 if size <= 3200 else 14
        set_col_width(ws, ci, w)

    cur_row = 2   # next free row

    # -----------------------------------------------------------------------
    # Table 1 — Raw measurements (pivoted: reps as rows, sizes as cols)
    # -----------------------------------------------------------------------
    # Section label
    ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
    sec = ws.cell(cur_row, 1, value="Tabla 1  —  Mediciones individuales (ms)")
    sec.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec.fill      = PatternFill("solid", fgColor=C["dark"])
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    for row_idx, row in enumerate(rows):
        row_reps = all_reps.get(row, {})

        # --- impl block header ---
        ws.merge_cells(
            f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}"
        )
        bh = ws.cell(cur_row, 1, value=row.label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11,
                            color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C_EXTRA["impl_header"])
        bh.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        bh.border    = _thick_bottom()
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        # --- column sub-header: Rep | N=400 | N=800 | ... ---
        _hdr(ws.cell(cur_row, 1), "Rep",
             bg=C["mid"], size=9)
        for ci, size in enumerate(sizes, 2):
            _hdr(ws.cell(cur_row, ci), f"N = {size:,}",
                 bg=C["mid"], size=9)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # --- rep rows ---
        for rep in range(1, n_reps + 1):
            alt = (rep % 2 == 0)
            bg  = C["alt"] if alt else "FFFFFF"
            _dat(ws.cell(cur_row, 1), rep, fmt="0",
                 bg=C["light"], bold=True, align="center")
            for ci, size in enumerate(sizes, 2):
                val = next(
                    (v for r, v in row_reps.get(size, []) if r == rep),
                    None,
                )
                if val is not None:
                    _dat(ws.cell(cur_row, ci), round(val, 3),
                         fmt="#,##0.000", bg=bg)
                else:
                    _dat(ws.cell(cur_row, ci), "—", bg=bg, align="center")
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # --- summary rows: Avg / StdDev / CV% ---
        summary_labels = ["Promedio", "Desv. Est.", "CV (%)"]
        for s_idx, s_label in enumerate(summary_labels):
            is_last = (s_idx == len(summary_labels) - 1)
            _hdr(ws.cell(cur_row, 1), s_label,
                 bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"],
                 size=9, align="left")
            for ci, size in enumerate(sizes, 2):
                vals = [v for _, v in row_reps.get(size, [])]
                if vals:
                    mean = sum(vals) / len(vals)
                    std  = math.sqrt(
                        sum((v - mean) ** 2 for v in vals) / len(vals)
                    )
                    cv   = std / mean * 100 if mean > 0 else 0.0
                    show = mean if s_idx == 0 else (std if s_idx == 1 else cv)
                    fmt  = "#,##0.000" if s_idx < 2 else "0.00"
                    c = ws.cell(cur_row, ci)
                    _summary_dat(c, round(show, 3 if s_idx < 2 else 2),
                                 fmt=fmt, bold=(s_idx == 0))
                else:
                    _summary_dat(ws.cell(cur_row, ci), "—")
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        # blank separator between impl blocks (except last)
        if row_idx < len(rows) - 1:
            for ci in range(1, N_COLS + 1):
                ws.cell(cur_row, ci).fill = PatternFill(
                    "solid", fgColor=C_EXTRA["sep"]
                )
            ws.row_dimensions[cur_row].height = 6
            cur_row += 1

    cur_row += 2   # space before table 2

    # -----------------------------------------------------------------------
    # Table 2 — Speedup summary
    # -----------------------------------------------------------------------
    ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
    sec2 = ws.cell(cur_row, 1,
                   value="Tabla 2  —  Promedio y Speedup")
    sec2.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec2.fill      = PatternFill("solid", fgColor=C["dark"])
    sec2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    # SP table cols: Impl | [avg, sp] * n_sizes | Avg Sp
    N_SP = 1 + len(sizes) * 2 + 1

    # Re-set widths for speedup table (same col indices, different meaning)
    set_col_width(ws, 1, 28)
    for si, size in enumerate(sizes):
        ac = 2 + si * 2
        sc = ac + 1
        w  = 12 if size <= 800 else 14 if size <= 3200 else 16
        set_col_width(ws, ac, w)
        set_col_width(ws, sc, 10)
    set_col_width(ws, N_SP, 12)

    # SP header
    _hdr(ws.cell(cur_row, 1), "Implementación", bg=C["dark"], size=10)
    for si, size in enumerate(sizes):
        ac, sc = 2 + si * 2, 3 + si * 2
        _hdr(ws.cell(cur_row, ac), f"N={size:,}\nAvg (ms)",
             bg=C["mid"], size=9)
        _hdr(ws.cell(cur_row, sc), f"N={size:,}\nSpeedup",
             bg=C["mid"], size=9)
    _hdr(ws.cell(cur_row, N_SP), "Speedup\nProm.", bg=C["dark"], size=9)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    ref_avgs = avg_data.get(ref, {})

    for ri, row in enumerate(rows, cur_row):
        bg     = C["alt"] if ri % 2 == 0 else "FFFFFF"
        is_ref = (row == ref)
        times  = avg_data.get(row, {})

        _dat(ws.cell(ri, 1), row.label,
             bg=C["light"], bold=True, align="left")

        sp_refs: list[str] = []
        for si, size in enumerate(sizes):
            ac  = 2 + si * 2
            sc  = ac + 1
            avg = times.get(size)

            avg_c               = ws.cell(ri, ac)
            avg_c.value         = round(avg, 3) if avg is not None else "N/A"
            avg_c.number_format = "#,##0.000"
            avg_c.font          = Font(name=FONT_NAME, size=10)
            avg_c.fill          = PatternFill("solid", fgColor=bg)
            avg_c.alignment     = Alignment(horizontal="right",
                                            vertical="center")
            avg_c.border        = _thin_border()

            sp_c    = ws.cell(ri, sc)
            ref_avg = ref_avgs.get(size)
            if is_ref:
                sp_c.value = 1.0
            elif avg and avg > 0 and ref_avg:
                sp_c.value = round(ref_avg / avg, 4)
            else:
                sp_c.value = "N/A"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10,
                                  color=C["green_fg"])
            sp_c.fill      = PatternFill("solid",
                                          fgColor=C["green_bg"])
            sp_c.alignment = Alignment(horizontal="right",
                                       vertical="center")
            sp_c.border    = _thin_border()

            if not is_ref:
                sp_refs.append(f"{get_column_letter(sc)}{ri}")

        avsp           = ws.cell(ri, N_SP)
        avsp.value     = (f"=IFERROR(AVERAGE({','.join(sp_refs)}),\"N/A\")"
                          if sp_refs else 1.0)
        avsp.number_format = "0.00"
        avsp.font      = Font(name=FONT_NAME, bold=True, size=10,
                              color=C["green_fg"])
        avsp.fill      = PatternFill("solid", fgColor=C["green_bg"])
        avsp.alignment = Alignment(horizontal="right", vertical="center")
        avsp.border    = _thin_border()
        ws.row_dimensions[ri].height = 17

    chart_row_anchor = ri + 4  # type: ignore[possibly-undefined]

    # -----------------------------------------------------------------------
    # Charts — side by side
    # -----------------------------------------------------------------------
    for anchor, path in [
        (f"A{chart_row_anchor}", ct),
        (f"L{chart_row_anchor}", cs),
    ]:
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = 620
            img.height = 360
            ws.add_image(img, anchor)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    print("Reading CSVs...")
    df_compiler = load_csv("compiler")
    df_cache    = load_csv("cache")
    df_mixed    = load_csv("mixed")

    frames = [d for d in [df_compiler, df_cache, df_mixed] if d is not None]
    if not frames:
        print("No data found.")
        sys.exit(1)

    sizes = sorted(pd.concat(frames)["matrix_size"].unique().tolist())

    all_reps: AllReps = {}
    for src in frames:
        for row, rdata in build_all_reps(src).items():
            if row not in all_reps:
                all_reps[row] = {}
            all_reps[row].update(rdata)

    avg_data = avgs(all_reps)
    t_counts = sorted({r.threads for r in all_reps if r.impl == "threads"})

    # Sheet row definitions
    ref1  = Row("seq_std", 0, "best")
    rows1 = (
        [ref1]
        + [Row("threads", t, "noopt") for t in t_counts
           if Row("threads", t, "noopt") in all_reps]
    )

    ref2  = Row("seq_std", 0, "best")
    rows2 = (
        [ref2]
        + [Row("threads", t, "best") for t in t_counts
           if Row("threads", t, "best") in all_reps]
    )

    ref3  = Row("seq_std", 0, "best")
    rows3 = [r for r in [ref3,
                          Row("seq_cache", 0, "noopt"),
                          Row("seq_cache", 0, "best")]
             if r in all_reps]

    ref4  = Row("seq_std", 0, "best")
    rows4 = [r for r in [ref4,
                          Row("conc", 0, "noopt"),
                          Row("conc", 0, "best")]
             if r in all_reps]

    ref5  = Row("seq_std", 0, "best")
    rows5 = [r for r in [
        ref5,
        Row("threads", 6, "best"),
        Row("seq_cache", 0, "best"),
        Row("conc", 0, "best"),
    ] if r in all_reps]

    ref6  = Row("seq_std", 0, "best")
    rows6 = [r for r in (
        [ref6]
        + [Row("threads", t, "noopt") for t in t_counts
           if Row("threads", t, "noopt") in all_reps]
        + [Row("conc", 0, "noopt")]
    ) if r in all_reps]

    print("\nGenerating charts...")

    def gen(prefix: str, rows: list[Row], ref: Row,
            t_title: str, s_title: str) -> tuple[str, str]:
        ct = chart_time(rows, avg_data, sizes, t_title, f"{prefix}_time.png")
        cs = chart_speedup(rows, avg_data, ref, sizes, s_title,
                           f"{prefix}_sp.png")
        print(f"  {os.path.basename(ct)}  |  {os.path.basename(cs)}")
        return ct, cs

    ct1, cs1 = gen("s1", rows1, ref1,
        "Tiempo  |  Hilos sin opt  vs  Secuencial naive con O3_full",
        "Speedup  |  T(seq_std/best) / T(threads_Nt/noopt)")
    ct2, cs2 = gen("s2", rows2, ref2,
        "Tiempo  |  Hilos con O3_full  vs  Secuencial naive con O3_full",
        "Speedup  |  T(seq_std/best) / T(threads_Nt/best)")
    ct3, cs3 = gen("s3", rows3, ref3,
        "Tiempo  |  Cache: naive/best vs cache/noopt vs cache/best",
        "Speedup  |  ref = seq_std/best")
    ct4, cs4 = gen("s4", rows4, ref4,
        "Tiempo  |  Procesos: sin opt vs con O3_full  vs  Secuencial naive/best",
        "Speedup  |  ref = seq_std/best")
    ct5, cs5 = gen("s5", rows5, ref5,
        "Tiempo  |  Mejor de cada estrategia",
        "Speedup  |  ref = seq_std/best")
    ct6, cs6 = gen("s6", rows6, ref6,
        "Tiempo  |  seq_std/best vs hilos/noopt vs conc/noopt",
        "Speedup  |  ref = seq_std/best")

    print("\nBuilding workbook...")
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)

    write_sheet(wb, "1. Hilos sin opt",
                "Hilos sin opt  |  Speedup = T(seq_std/best) / T(threads_Nt/noopt)",
                rows1, all_reps, ref1, sizes, ct1, cs1)
    write_sheet(wb, "2. Hilos con opt",
                "Hilos con O3_full  |  Speedup = T(seq_std/best) / T(threads_Nt/best)",
                rows2, all_reps, ref2, sizes, ct2, cs2)
    write_sheet(wb, "3. Cache",
                "Cache line  |  seq_std/best  vs  seq_cache/noopt  vs  seq_cache/best",
                rows3, all_reps, ref3, sizes, ct3, cs3)
    write_sheet(wb, "4. Procesos",
                "Procesos (fork)  |  conc/noopt  vs  conc/best  vs  seq_std/best  |  ref = seq_std/best",
                rows4, all_reps, ref4, sizes, ct4, cs4)
    write_sheet(wb, "5. Comparacion final",
                "Comparacion final  |  Mejor de cada estrategia  |  ref = seq_std/best",
                rows5, all_reps, ref5, sizes, ct5, cs5)
    write_sheet(wb, "6. Efecto compilador",
                "Efecto compilador  |  seq_std/best vs hilos/noopt vs conc/noopt",
                rows6, all_reps, ref6, sizes, ct6, cs6)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")

    print("\nExporting table PNGs...")
    table_specs = [
        (rows1, "1. Hilos sin opt",      "tabla_s1_hilos_sin_opt.png"),
        (rows2, "2. Hilos con opt",      "tabla_s2_hilos_con_opt.png"),
        (rows3, "3. Cache",              "tabla_s3_cache.png"),
        (rows4, "4. Procesos",           "tabla_s4_procesos.png"),
        (rows5, "5. Comparacion final",  "tabla_s5_comparacion.png"),
        (rows6, "6. Efecto compilador",  "tabla_s6_compilador.png"),
    ]
    for t_rows, t_title, t_fname in table_specs:
        path = export_table_png(t_rows, all_reps, sizes, t_title, t_fname)
        print(f"  {os.path.basename(path)}")
    print(f"  Tables saved to: {CHARTS_DIR}/")



# ---------------------------------------------------------------------------
# Table PNG export  (for LaTeX / Word import)
# ---------------------------------------------------------------------------

def export_table_png(rows: list[Row], all_reps: AllReps,
                     sizes: list[int], table_title: str,
                     fname: str) -> str:
    """
    Renders the speedup summary table as a PNG image.
    Suitable for direct inclusion in LaTeX (includegraphics)
    or pasting into Word/Google Docs.
    """
    avg_data  = avgs(all_reps)
    col_hdrs  = ["Implementacion"] + [f"N={s:,}" for s in sizes] + ["Sp. prom."]
    ref_row   = rows[0]
    ref_avgs  = avg_data.get(ref_row, {})

    table_data: list[list[str]] = []
    for row in rows:
        times   = avg_data.get(row, {})
        is_ref  = (row == ref_row)
        row_vals = [row.label]
        sp_list  = []
        for size in sizes:
            avg = times.get(size)
            ref_avg = ref_avgs.get(size)
            if avg is None:
                row_vals.append("N/A")
            else:
                if is_ref:
                    row_vals.append(f"{avg:,.1f} ms\n(ref)")
                elif ref_avg and avg > 0:
                    sp = ref_avg / avg
                    sp_list.append(sp)
                    row_vals.append(f"{avg:,.1f} ms\n{sp:.2f}x")
                else:
                    row_vals.append(f"{avg:,.1f} ms")
        if is_ref:
            row_vals.append("1.00x")
        elif sp_list:
            row_vals.append(f"{sum(sp_list)/len(sp_list):.2f}x")
        else:
            row_vals.append("N/A")
        table_data.append(row_vals)

    n_cols = len(col_hdrs)
    n_rows = len(table_data)

    col_w    = [3.2] + [1.4] * len(sizes) + [1.1]
    fig_w    = sum(col_w) + 0.4
    row_h    = 0.55
    hdr_h    = 0.45
    title_h  = 0.45
    fig_h    = title_h + hdr_h + n_rows * row_h + 0.2

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")
    fig.patch.set_facecolor("white")

    # Title
    fig.text(0.5, 1 - title_h / fig_h * 0.6,
             table_title, ha="center", va="top",
             fontsize=10, fontweight="bold",
             fontfamily="DejaVu Sans")

    # Compute relative positions
    total_w = sum(col_w)
    x_pos   = [sum(col_w[:i]) / total_w for i in range(n_cols)]
    x_pos.append(1.0)

    content_top = 1 - title_h / fig_h
    hdr_top     = content_top
    hdr_bot     = hdr_top - hdr_h / fig_h

    HDR_BG  = "#2E75B6"
    ALT_BG  = "#F2F9FF"
    REF_BG  = "#FFF2CC"

    # Header row
    for ci in range(n_cols):
        x0 = x_pos[ci]
        x1 = x_pos[ci + 1]
        rect = mpatches.Rectangle((x0, hdr_bot), x1 - x0,
                              hdr_h / fig_h,
                              transform=ax.transAxes,
                              facecolor=HDR_BG, edgecolor="white",
                              linewidth=0.5, clip_on=False)
        ax.add_patch(rect)
        ax.text((x0 + x1) / 2, (hdr_top + hdr_bot) / 2,
                col_hdrs[ci],
                ha="center", va="center",
                fontsize=7.5, fontweight="bold", color="white",
                fontfamily="DejaVu Sans",
                transform=ax.transAxes)

    # Data rows
    for ri, row_vals in enumerate(table_data):
        is_ref  = (ri == 0)
        bg      = REF_BG if is_ref else (ALT_BG if ri % 2 == 0 else "white")
        row_top = hdr_bot - ri * row_h / fig_h
        row_bot = row_top - row_h / fig_h

        for ci, val in enumerate(row_vals):
            x0 = x_pos[ci]
            x1 = x_pos[ci + 1]
            rect = mpatches.Rectangle((x0, row_bot), x1 - x0,
                                  row_h / fig_h,
                                  transform=ax.transAxes,
                                  facecolor=bg, edgecolor="#BDD7EE",
                                  linewidth=0.4, clip_on=False)
            ax.add_patch(rect)
            ha   = "left" if ci == 0 else "center"
            pad  = 0.008 if ci == 0 else 0
            bold = ci == 0 or ci == n_cols - 1
            ax.text(x0 + pad + (x1 - x0) / (1 if ci == 0 else 2) * (0 if ci == 0 else 1),
                    (row_top + row_bot) / 2,
                    val,
                    ha=ha, va="center",
                    fontsize=7, fontweight="bold" if bold else "normal",
                    color="#1F4E79" if (ci == n_cols - 1 and not is_ref) else "black",
                    fontfamily="DejaVu Sans",
                    transform=ax.transAxes,
                    multialignment="center")

    fig.tight_layout(pad=0)
    path = os.path.join(CHARTS_DIR, fname)
    fig.savefig(path, dpi=200, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    return path


if __name__ == "__main__":
    main()