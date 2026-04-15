"""
report_profiling.py -- Report for sequential matrix multiplication profiling.

Reads data_profiling.csv (produced by bench_profiling.sh), generates:
  - PNG charts for CPU metrics, memory, and cache misses vs N
  - Excel workbook following project conventions (report_utils.py)
  - LaTeX table snippet

Usage:
    python3 tests/benchmarks/report_profiling.py [path/to/data_profiling.csv]
"""

from __future__ import annotations

import math
import os
import re
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


CSV_PATH   = sys.argv[1] if len(sys.argv) > 1 \
             else "tests/benchmarks/machine1/results_profiling/data_profiling.csv"
OUT_DIR    = os.path.dirname(os.path.abspath(CSV_PATH))
CHARTS_DIR = os.path.join(OUT_DIR, "charts_profiling")
XLSX_PATH  = os.path.join(OUT_DIR, "reporte_profiling.xlsx")
RAW_DIR    = os.path.join(OUT_DIR, "raw")


try:
    sys.path.insert(0, "tests/benchmarks")
    from report_utils import (
        C, CHART_STYLE, FONT_NAME,
        ColConfig, Series,
        make_border, set_col_width, style_data_cell, style_header_cell,
        write_title_row, write_raw_data_sheet, write_charts_sheet,
        save_figure, plot_lines, plot_simple_bars,
        compute_averages,
    )
    _HAS_UTILS = True
except ImportError:
    _HAS_UTILS = False
    FONT_NAME = "Arial"
    C = {
        "dark": "1F4E79", "mid": "2E75B6", "light": "D6E4F0",
        "alt": "F2F9FF", "white": "FFFFFF", "green_bg": "E2EFDA",
        "green_fg": "375623", "red_bg": "FCE4D6", "border": "BDD7EE",
    }
    CHART_STYLE = {
        "figure.facecolor": "white", "axes.facecolor": "white",
        "axes.grid": True, "grid.color": "#E8E8E8",
        "axes.spines.top": False, "axes.spines.right": False,
    }

    def save_figure(fig, charts_dir, name):
        path = os.path.join(charts_dir, name)
        fig.savefig(path, dpi=180, bbox_inches="tight")
        plt.close(fig)
        return path

    def make_border():
        from openpyxl.styles import Border, Side
        s = Side(style="thin", color=C["border"])
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_width(ws, col, width):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width

    def style_header_cell(cell, value, bg="dark", size=10):
        cell.value = value
        cell.font = Font(name=FONT_NAME, bold=True, color=C["white"], size=size)
        cell.fill = PatternFill("solid", fgColor=C.get(bg, bg))
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = make_border()

    def style_data_cell(cell, value, fmt=None, bg="white", fg="000000",
                        bold=False, align="right"):
        cell.value = value
        cell.font = Font(name=FONT_NAME, bold=bold, color=C.get(fg, fg), size=10)
        cell.fill = PatternFill("solid", fgColor=C.get(bg, bg))
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = make_border()
        if fmt:
            cell.number_format = fmt

    def write_title_row(ws, text, n_cols, row=1):
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT_NAME, bold=True, size=13, color=C["white"])
        cell.fill = PatternFill("solid", fgColor=C["dark"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24

def load_data(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = [c.strip().lower() for c in df.columns]
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()
    for col in df.columns:
        if col != "matrix_size":
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["matrix_size"] = df["matrix_size"].astype(int)
    df = df.sort_values("matrix_size").reset_index(drop=True)
    return df

def theoretical_heap_mb(n: int) -> float:
    """3 NxN int matrices + N row-pointer arrays each, all ints = 4 bytes."""
    bytes_ = 3 * (n * n * 4 + n * 8)
    return bytes_ / 1024 / 1024

def parse_gprof_top(n: int, top_k: int = 5) -> list[tuple[str, float]]:
    """Return [(func_name, pct_time)] from the flat profile section."""
    report = os.path.join(RAW_DIR, f"N{n}", "gprof_report.txt")
    if not os.path.exists(report):
        return []
    results: list[tuple[str, float]] = []
    in_flat = False
    for line in open(report):
        if "Flat profile" in line:
            in_flat = True
        if not in_flat:
            continue
        m = re.match(r"^\s*([\d.]+)\s+[\d.]+\s+[\d.]+\s+.*\s+(\S+)\s*$", line)
        if m:
            results.append((m.group(2), float(m.group(1))))
        if len(results) >= top_k:
            break
    return results

COLOR_GFLOPS = "#2E75B6"
COLOR_CACHE  = "#C00000"
COLOR_L1     = "#ED7D31"
COLOR_MEM    = "#70AD47"
COLOR_IPC    = "#7030A0"

def _base_line_chart(ax, sizes, values, color, title, ylabel, marker="o"):
    ax.plot(sizes, values, marker=marker, lw=2.5, color=color)
    ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
    ax.set_xlabel("Dimensión de la matriz (N)", fontsize=10)
    ax.set_ylabel(ylabel, fontsize=10)
    ax.set_xticks(sizes)
    ax.set_xticklabels([f"N={s}" for s in sizes], fontsize=9)
    ax.grid(True, linestyle="--", alpha=0.6)

def chart_gflops(df: pd.DataFrame) -> str:
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5))
        _base_line_chart(ax, df["matrix_size"].tolist(),
                         df["gflops"].tolist(),
                         COLOR_GFLOPS,
                         "Rendimiento Secuencial — GFLOPS vs N",
                         "GFLOPS")
        for x, y in zip(df["matrix_size"], df["gflops"]):
            ax.annotate(f"{y:.3f}", (x, y), textcoords="offset points",
                        xytext=(0, 8), ha="center", fontsize=8)
        fig.tight_layout()
        return save_figure(fig, CHARTS_DIR, "gflops.png")

def chart_cache_miss(df: pd.DataFrame) -> str:
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(df["matrix_size"], df["cache_miss_pct"],
                marker="o", lw=2.5, color=COLOR_CACHE, label="LLC miss rate")
        if "l1_miss_pct" in df.columns:
            ax.plot(df["matrix_size"], df["l1_miss_pct"],
                    marker="s", lw=2.5, color=COLOR_L1, linestyle="--",
                    label="L1-D miss rate")
        ax.set_title("Cache Miss Rate vs Dimensión de Matriz",
                     fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("N", fontsize=10)
        ax.set_ylabel("Miss rate (%)", fontsize=10)
        ax.set_xticks(df["matrix_size"].tolist())
        ax.set_xticklabels([f"N={s}" for s in df["matrix_size"]], fontsize=9)
        ax.legend(fontsize=9)
        ax.grid(True, linestyle="--", alpha=0.6)
        fig.tight_layout()
        return save_figure(fig, CHARTS_DIR, "cache_miss.png")

def chart_memory(df: pd.DataFrame) -> str:
    sizes      = df["matrix_size"].tolist()
    measured   = df["peak_heap_mb"].tolist()
    theoretical = [theoretical_heap_mb(n) for n in sizes]

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(sizes, measured, marker="o", lw=2.5, color=COLOR_MEM,
                label="Medido (valgrind massif)")
        ax.plot(sizes, theoretical, marker="s", lw=2, color="#AAAAAA",
                linestyle="--", label="Teórico (3·N²·4 bytes)")
        ax.set_title("Uso de Memoria Heap vs Dimensión de Matriz",
                     fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("N", fontsize=10)
        ax.set_ylabel("Memoria pico (MB)", fontsize=10)
        ax.set_xticks(sizes)
        ax.set_xticklabels([f"N={s}" for s in sizes], fontsize=9)
        ax.legend(fontsize=9)
        ax.grid(True, linestyle="--", alpha=0.6)
        fig.tight_layout()
        return save_figure(fig, CHARTS_DIR, "memory.png")

def chart_ipc(df: pd.DataFrame) -> str:
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5))
        _base_line_chart(ax, df["matrix_size"].tolist(),
                         df["ipc"].tolist(),
                         COLOR_IPC,
                         "IPC (Instructions per Cycle) vs N",
                         "IPC", marker="D")
        ax.axhline(4, color="#AAAAAA", lw=1, ls="--",
                   label="4-wide decode (Zen4 teórico)")
        ax.legend(fontsize=9)
        fig.tight_layout()
        return save_figure(fig, CHARTS_DIR, "ipc.png")

def chart_gprof_topfuncs(df: pd.DataFrame) -> str:
    largest_n = df["matrix_size"].max()
    funcs = parse_gprof_top(largest_n, top_k=5)
    if not funcs:
        return ""

    labels = [f[0].split("::")[-1][:25] for f in funcs]
    values = [f[1] for f in funcs]
    colors = ["#C00000" if i == 0 else "#2E75B6" for i in range(len(labels))]

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        bars = ax.barh(labels[::-1], values[::-1], color=colors[::-1],
                       edgecolor="white")
        for bar, val in zip(bars, values[::-1]):
            ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%", va="center", fontsize=9)
        ax.set_xlabel("% tiempo de CPU", fontsize=10)
        ax.set_title(f"Top funciones — gprof (N={largest_n})",
                     fontsize=12, fontweight="bold", pad=8)
        ax.set_xlim(0, 110)
        fig.tight_layout()
        return save_figure(fig, CHARTS_DIR, "gprof_top_funcs.png")

def write_raw_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw Data")
    ws.freeze_panes = "A3"
    cols = [
        ("N", 10), ("Time (ms)", 14), ("GFLOPS", 12), ("Cycles", 18),
        ("Instructions", 18), ("IPC", 10), ("Cache Refs", 16),
        ("Cache Misses", 16), ("Cache Miss %", 14), ("L1 Loads", 16),
        ("L1 Misses", 14), ("L1 Miss %", 12), ("Peak Heap (MB)", 16),
    ]
    write_title_row(ws, "Raw Profiling Data | Sequential MatMul", len(cols))
    for ci, (header, width) in enumerate(cols, 1):
        style_header_cell(ws.cell(2, ci), header, bg="mid")
        set_col_width(ws, ci, width)
    ws.row_dimensions[2].height = 20

    for ri, row in enumerate(df.itertuples(index=False), 3):
        bg = "alt" if ri % 2 == 0 else "white"
        for ci, col_name in enumerate(df.columns, 1):
            val = getattr(row, col_name, None)
            if col_name == "matrix_size":
                style_data_cell(ws.cell(ri, ci), int(val), fmt="#,##0",
                                bg=bg, align="center")
            elif isinstance(val, float):
                style_data_cell(ws.cell(ri, ci), round(val, 4), fmt="0.0000",
                                bg=bg)
            else:
                style_data_cell(ws.cell(ri, ci), val, bg=bg)

def write_analysis_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Analysis")
    write_title_row(ws, "Profiling Analysis | Sequential MatMul", 4)

    for col, w in enumerate([28, 18, 18, 55], 1):
        set_col_width(ws, col, w)

    row = [3]

    def section(title: str) -> None:
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f"A{row[0]}:D{row[0]}")
        c = ws.cell(row[0], 1, value=title)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color=C["white"])
        c.fill = PatternFill("solid", fgColor=C["mid"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row[0]].height = 20
        row[0] += 1

    def finding(label: str, value: str, note: str = "") -> None:
        bg = "alt" if row[0] % 2 == 0 else "white"
        style_data_cell(ws.cell(row[0], 1), label, bg=bg, align="left", bold=True)
        style_data_cell(ws.cell(row[0], 2), value, bg=bg, align="center")
        style_data_cell(ws.cell(row[0], 3), "", bg=bg)
        style_data_cell(ws.cell(row[0], 4), note, bg=bg, align="left")
        row[0] += 1

    section("1. Rendimiento (GFLOPS)")
    for _, r in df.iterrows():
        finding(f"N = {int(r['matrix_size'])}", f"{r['gflops']:.4f} GFLOPS",
                "2·N³ operaciones / tiempo")

    row[0] += 1
    section("2. Eficiencia del pipeline (IPC)")
    for _, r in df.iterrows():
        note = ("Buen IPC" if r["ipc"] >= 2 else
                "IPC moderado — stalls por memoria" if r["ipc"] >= 1 else
                "IPC bajo — cuello de botella de memoria")
        finding(f"N = {int(r['matrix_size'])}", f"{r['ipc']:.4f}", note)

    row[0] += 1
    section("3. Cache Miss Rate (LLC)")
    for _, r in df.iterrows():
        note = ("Excelente" if r["cache_miss_pct"] < 1 else
                "Aceptable" if r["cache_miss_pct"] < 10 else
                "Alto — matriz no cabe en LLC → justifica OpenMP con localidad")
        finding(f"N = {int(r['matrix_size'])}", f"{r['cache_miss_pct']:.3f}%", note)

    row[0] += 1
    section("4. Memoria heap pico")
    for _, r in df.iterrows():
        n = int(r["matrix_size"])
        theo = theoretical_heap_mb(n)
        finding(f"N = {n}",
                f"{r['peak_heap_mb']:.3f} MB",
                f"Teórico: {theo:.3f} MB (3·N²·4 bytes + punteros)")

def write_charts_sheet_local(wb: Workbook, paths: list[str]) -> None:
    from openpyxl.drawing.image import Image as XLImage
    ws = wb.create_sheet("Charts")
    write_title_row(ws, "Visual Analysis | CPU & Memory Profiling", 12)
    anchors = ["A2", "M2", "A28", "M28", "A54", "M54"]
    for anchor, path in zip(anchors, paths):
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width = 700
            img.height = 380
            ws.add_image(img, anchor)

def write_latex_table(df: pd.DataFrame) -> None:
    lines = [
        r"\begin{table}[h]",
        r"\centering",
        r"\caption{Caracterización de la implementación secuencial de "
        r"multiplicación de matrices}",
        r"\label{tab:seq_profiling}",
        r"\begin{tabular}{rrrrrrrr}",
        r"\toprule",
        r"$N$ & Time (ms) & GFLOPS & IPC & Cache Miss \% & "
        r"L1 Miss \% & Peak Heap (MB) \\",
        r"\midrule",
    ]
    for _, r in df.iterrows():
        lines.append(
            f"{int(r['matrix_size'])} & {r['time_ms']:.1f} & {r['gflops']:.4f} & "
            f"{r['ipc']:.3f} & {r['cache_miss_pct']:.2f} & "
            f"{r['l1_miss_pct']:.2f} & {r['peak_heap_mb']:.2f} \\\\"
        )
    lines += [r"\bottomrule", r"\end{tabular}", r"\end{table}"]
    latex_path = os.path.join(OUT_DIR, "profiling_table.tex")
    with open(latex_path, "w") as f:
        f.write("\n".join(lines))
    print(f"LaTeX table: {latex_path}")

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    print(f"Reading: {CSV_PATH}")
    df = load_data(CSV_PATH)
    print(df[["matrix_size", "time_ms", "gflops", "ipc",
              "cache_miss_pct", "peak_heap_mb"]].to_string(index=False))

    print("Generating charts...")
    chart_paths = [
        chart_gflops(df),
        chart_cache_miss(df),
        chart_memory(df),
        chart_ipc(df),
        chart_gprof_topfuncs(df),
    ]
    chart_paths = [p for p in chart_paths if p]
    for p in chart_paths:
        print(f"  {os.path.basename(p)}")

    print("Building workbook...")
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    write_raw_sheet(wb, df)
    write_analysis_sheet(wb, df)
    write_charts_sheet_local(wb, chart_paths)

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")

    write_latex_table(df)
    print("Done.")

if __name__ == "__main__":
    main()