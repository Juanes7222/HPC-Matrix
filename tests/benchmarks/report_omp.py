"""
report_omp.py  --  HPC benchmark report: OpenMP matrix multiplication.

Compares the performance of the OpenMP implementation across two or more
machines. Each machine is identified by the 'machine' column in the CSV.

Table layout per sheet (one sheet per thread count):
  One block per machine:
    Header  : machine label (merged, dark)
    Sub-hdr : Rep | N=400 | N=800 | N=1600 | N=3200 | N=6400
    Rows    : 1..10  (individual measurements)
    Summary : Avg / StdDev / CV%
    Blank separator before next block

Speedup table below (one row per machine):
  Machine | N=400 avg | sp | N=800 avg | sp | ... | Avg Speedup
  Speedup = T(reference_machine) / T(row)

Usage:
    python report_omp.py [results_dir/ [machine1 machine2 ...]]
    results_dir   : directory containing data_omp.csv files per machine
                    (default: tests/benchmarks)
    machine1 ...  : machine flag names to include
                    (default: all subdirs containing data_omp.csv)
"""

from __future__ import annotations

import math
import os
import sys
from dataclasses import dataclass

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    Series,
    make_border, set_col_width, style_data_cell, style_header_cell,
    write_title_row, save_figure, plot_lines,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BENCHMARKS_DIR = sys.argv[1].rstrip("/") if len(sys.argv) > 1 else "tests/benchmarks"
EXPLICIT_MACHINES = sys.argv[2:] if len(sys.argv) > 2 else []
OUTPUT_PATH = os.path.join(BENCHMARKS_DIR, "reporte_omp.xlsx")
CHARTS_DIR  = os.path.join(BENCHMARKS_DIR, "charts_omp")

BEST_MARKER = "-march=native"

# Color palette — one color per machine (cycles if more than defined)
MACHINE_COLORS = [
    "#2E75B6",
    "#C00000",
    "#70AD47",
    "#ED7D31",
    "#7030A0",
    "#1F4E79",
    "#843C0C",
]

C_EXTRA: dict[str, str] = {
    "summary_bg":   "EBF3FB",
    "summary_fg":   "1F4E79",
    "machine_hdr":  "1F4E79",
    "sep":          "D9E1F2",
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class MachineRow:
    machine:  str
    threads:  int

    @property
    def label(self) -> str:
        return f"{self.machine}  ({self.threads} threads)"

AllReps = dict[MachineRow, dict[int, list[tuple[int, float]]]]

# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def discover_machines(base: str, explicit: list[str]) -> list[str]:
    if explicit:
        return explicit
    found = []
    for entry in sorted(os.listdir(base)):
        csv_path = os.path.join(base, entry, "results_omp", "data_omp.csv")
        if os.path.isfile(csv_path):
            found.append(entry)
    return found


def load_machine_csv(base: str, machine: str) -> pd.DataFrame | None:
    path = os.path.join(base, machine, "results_omp", "data_omp.csv")
    if not os.path.exists(path):
        print(f"  [--]  {path} not found")
        return None
    df = pd.read_csv(path)
    df.columns        = [c.strip().lower() for c in df.columns]
    df["flags"]       = df["flags"].str.strip('"')
    df["machine"]     = df["machine"].str.strip()
    df["threads"]     = df["threads"].astype(int)
    df["matrix_size"] = df["matrix_size"].astype(int)
    df["repetition"]  = df["repetition"].astype(int)
    df["wall_time_ms"]= df["wall_time_ms"].astype(float)
    print(f"  [OK]  {path}  ({len(df)} rows)")
    return df


def build_all_reps(df: pd.DataFrame) -> AllReps:
    result: AllReps = {}
    for (machine, threads), grp in df.groupby(["machine", "threads"]):
        row = MachineRow(machine=str(machine), threads=int(str(threads)))
        result[row] = {}
        for size, sub in grp.groupby("matrix_size"):
            pairs = sorted(
                [(int(str(r)), float(v))
                 for r, v in zip(sub["repetition"], sub["wall_time_ms"])
                 if float(v) > 0.0],  # discard invalid measurements
                key=lambda x: x[0],
            )
            if pairs:
                result[row][int(str(size))] = pairs
    return result

def slowest_row(rows: list[MachineRow], avg_data: dict[MachineRow, dict[int, float]]) -> MachineRow:
    """Returns the row with the highest mean execution time across all sizes."""
    def mean_time(row: MachineRow) -> float:
        times = avg_data.get(row, {})
        return sum(times.values()) / len(times) if times else 0.0

    return max(rows, key=mean_time)


def avgs(reps: AllReps) -> dict[MachineRow, dict[int, float]]:
    return {
        row: {s: sum(v for _, v in pairs) / len(pairs)
              for s, pairs in sizes.items()}
        for row, sizes in reps.items()
    }


def machine_color(machine: str, machines: list[str]) -> str:
    idx = machines.index(machine) if machine in machines else 0
    return MACHINE_COLORS[idx % len(MACHINE_COLORS)]

# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom() -> Border:
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin",   color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thick)


def _hdr(cell, value: str, bg: str, fg: str = "FFFFFF",
         size: int = 10, bold: bool = True,
         align: str = "center") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    cell.border    = _thin_border()


def _dat(cell, value, fmt: str | None = None,
         bg: str = "FFFFFF", fg: str = "000000",
         bold: bool = False, align: str = "right") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt


def _summary_dat(cell, value, fmt: str | None = None,
                 bold: bool = False) -> None:
    _dat(cell, value, fmt=fmt,
         bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"], bold=bold)

# ---------------------------------------------------------------------------
# Chart helpers
# ---------------------------------------------------------------------------

def chart_time_threads(thread_rows: list[MachineRow],
                       avg_data: dict, sizes: list,
                       machines: list[str],
                       title: str, fname: str) -> str:
    series = [
        Series(
            label=r.label,
            data=avg_data[r],
            color=machine_color(r.machine, machines),
        )
        for r in thread_rows if r in avg_data and avg_data[r]
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        plot_lines(ax, series, log_scale=True)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Dimensión N", fontsize=10)
        ax.set_ylabel("Tiempo promedio (ms)", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)

def chart_bar_comparison(thread_rows: list[MachineRow],
                         avg_data: dict,
                         sizes: list[int],
                         machines: list[str],
                         title: str,
                         fname: str) -> str:
    """Grouped bar chart comparing average time per machine for each matrix size."""
    import numpy as np

    n_sizes   = len(sizes)
    n_machines = len(thread_rows)
    bar_width  = 0.8 / n_machines
    x_base     = np.arange(n_sizes)

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 5))

        for mi, row in enumerate(thread_rows):
            times  = avg_data.get(row, {})
            values = [times.get(s, 0) for s in sizes]
            offset = (mi - (n_machines - 1) / 2) * bar_width
            color  = machine_color(row.machine, machines)

            bars = ax.bar(
                x_base + offset, values,
                width=bar_width,
                label=row.label,
                color=color,
                alpha=0.85,
                edgecolor="white",
                linewidth=0.5,
            )

            for bar, val in zip(bars, values):
                if val > 0:
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height() * 1.02,
                        f"{val/1000:.1f}s" if val >= 1000 else f"{val:.0f}ms",
                        ha="center", va="bottom",
                        fontsize=7, color=color,
                    )

        ax.set_xticks(x_base)
        ax.set_xticklabels([f"N={s:,}" for s in sizes], fontsize=9)
        ax.set_yscale("log")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Dimensión N", fontsize=10)
        ax.set_ylabel("Tiempo promedio (ms, escala log)", fontsize=10)
        ax.legend(fontsize=8)
        fig.tight_layout()

    return save_figure(fig, CHARTS_DIR, fname)

def chart_speedup_machines(thread_rows: list[MachineRow],
                           avg_data: dict, ref: MachineRow,
                           sizes: list, machines: list[str],
                           title: str, fname: str) -> str:
    ref_avgs = avg_data.get(ref, {})
    series = [
        Series(
            label=r.label,
            data={s: ref_avgs[s] / avg_data[r][s]
                  for s in sizes
                  if s in avg_data.get(r, {}) and s in ref_avgs
                  and avg_data[r][s] > 0},
            color=machine_color(r.machine, machines),
        )
        for r in thread_rows if r != ref and r in avg_data
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--",
                   label=f"ref: {ref.label}")
        plot_lines(ax, [s for s in series if s.data], log_scale=False)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Dimensión N", fontsize=10)
        ax.set_ylabel("Speedup vs referencia", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)


def chart_scaling(machine: str, all_reps: AllReps,
                  avg_data: dict, sizes: list,
                  thread_counts: list[int],
                  title: str, fname: str) -> str:
    """Strong scaling chart for a single machine: threads on x-axis."""
    palette = ["#2E75B6", "#C00000", "#70AD47",
               "#ED7D31", "#7030A0", "#1F4E79"]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        for si, size in enumerate(sizes):
            ys = []
            xs = []
            for t in thread_counts:
                row = MachineRow(machine=machine, threads=t)
                avg = avg_data.get(row, {}).get(size)
                if avg is not None:
                    xs.append(t)
                    ys.append(avg)
            if xs and ys:
                ax.plot(xs, ys, marker="o",
                        color=palette[si % len(palette)],
                        label=f"N={size:,}")
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Número de hilos", fontsize=10)
        ax.set_ylabel("Tiempo promedio (ms)", fontsize=10)
        ax.set_yscale("log")
        ax.legend(fontsize=8, loc="upper right")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)

# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def write_thread_sheet(wb: Workbook, sheet_name: str, title: str,
                       rows: list[MachineRow], all_reps: AllReps,
                       ref: MachineRow, sizes: list[int],
                       machines: list[str],
                       ct: str, cs: str, cb: str) -> None:
    """One sheet per thread count: compares all machines at that thread count."""
    ws = wb.create_sheet(sheet_name)
    avg_data = avgs(all_reps)
    n_reps   = max(
        (len(p) for row in rows for p in all_reps.get(row, {}).values()),
        default=10,
    )
    N_COLS = 1 + len(sizes)

    write_title_row(ws, title, N_COLS)
    ws.row_dimensions[1].height = 26

    set_col_width(ws, 1, 14)
    for ci, size in enumerate(sizes, 2):
        set_col_width(ws, ci, 12 if size <= 800 else 14 if size <= 3200 else 16)

    cur_row = 2

    # Table 1 — Raw measurements
    ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
    sec = ws.cell(cur_row, 1, value="Tabla 1  —  Mediciones individuales (ms)")
    sec.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec.fill      = PatternFill("solid", fgColor=C["dark"])
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    for row_idx, row in enumerate(rows):
        row_reps = all_reps.get(row, {})

        ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
        bh = ws.cell(cur_row, 1, value=row.label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C_EXTRA["machine_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        bh.border    = _thick_bottom()
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        _hdr(ws.cell(cur_row, 1), "Rep", bg=C["mid"], size=9)
        for ci, size in enumerate(sizes, 2):
            _hdr(ws.cell(cur_row, ci), f"N = {size:,}", bg=C["mid"], size=9)
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        for rep in range(1, n_reps + 1):
            alt = (rep % 2 == 0)
            bg  = C["alt"] if alt else "FFFFFF"
            _dat(ws.cell(cur_row, 1), rep, fmt="0",
                 bg=C["light"], bold=True, align="center")
            for ci, size in enumerate(sizes, 2):
                val = next(
                    (v for r, v in row_reps.get(size, []) if r == rep), None)
                if val is not None:
                    _dat(ws.cell(cur_row, ci), round(val, 3),
                         fmt="#,##0.000", bg=bg)
                else:
                    _dat(ws.cell(cur_row, ci), "—", bg=bg, align="center")
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        for s_idx, s_label in enumerate(["Promedio", "Desv. Est.", "CV (%)"]):
            _hdr(ws.cell(cur_row, 1), s_label,
                 bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"],
                 size=9, align="left")
            for ci, size in enumerate(sizes, 2):
                vals = [v for _, v in row_reps.get(size, [])]
                if vals:
                    mean = sum(vals) / len(vals)
                    std  = math.sqrt(
                        sum((v - mean) ** 2 for v in vals) / len(vals))
                    cv   = std / mean * 100 if mean > 0 else 0.0
                    show = mean if s_idx == 0 else (std if s_idx == 1 else cv)
                    fmt  = "#,##0.000" if s_idx < 2 else "0.00"
                    _summary_dat(ws.cell(cur_row, ci),
                                 round(show, 3 if s_idx < 2 else 2),
                                 fmt=fmt, bold=(s_idx == 0))
                else:
                    _summary_dat(ws.cell(cur_row, ci), "—")
            ws.row_dimensions[cur_row].height = 16
            cur_row += 1

        if row_idx < len(rows) - 1:
            for ci in range(1, N_COLS + 1):
                ws.cell(cur_row, ci).fill = PatternFill(
                    "solid", fgColor=C_EXTRA["sep"])
            ws.row_dimensions[cur_row].height = 6
            cur_row += 1

    cur_row += 2

    # Table 2 — Speedup summary
    ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
    sec2 = ws.cell(cur_row, 1,
                   value="Tabla 2  —  Promedio y Speedup entre máquinas")
    sec2.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec2.fill      = PatternFill("solid", fgColor=C["dark"])
    sec2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    N_SP = 1 + len(sizes) * 2 + 1
    set_col_width(ws, 1, 30)
    for si, size in enumerate(sizes):
        set_col_width(ws, 2 + si * 2, 14 if size <= 800 else 16)
        set_col_width(ws, 3 + si * 2, 10)
    set_col_width(ws, N_SP, 12)

    _hdr(ws.cell(cur_row, 1), "Máquina", bg=C["dark"], size=10)
    for si, size in enumerate(sizes):
        _hdr(ws.cell(cur_row, 2 + si * 2), f"N={size:,}\nAvg (ms)",
             bg=C["mid"], size=9)
        _hdr(ws.cell(cur_row, 3 + si * 2), f"N={size:,}\nSpeedup",
             bg=C["mid"], size=9)
    _hdr(ws.cell(cur_row, N_SP), "Speedup\nProm.", bg=C["dark"], size=9)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    ref_avgs = avg_data.get(ref, {})

    for ri, row in enumerate(rows, cur_row):
        bg      = C["alt"] if ri % 2 == 0 else "FFFFFF"
        is_ref  = (row == ref)
        times   = avg_data.get(row, {})

        _dat(ws.cell(ri, 1), row.label,
             bg=C["light"], bold=True, align="left")

        sp_refs: list[str] = []
        for si, size in enumerate(sizes):
            ac  = 2 + si * 2
            sc  = ac + 1
            avg = times.get(size)

            avg_c               = ws.cell(ri, ac)
            avg_c.value         = round(avg, 3) if avg is not None else "N/A"
            avg_c.number_format = "#,##0.000"
            avg_c.font          = Font(name=FONT_NAME, size=10)
            avg_c.fill          = PatternFill("solid", fgColor=bg)
            avg_c.alignment     = Alignment(horizontal="right", vertical="center")
            avg_c.border        = _thin_border()

            sp_c    = ws.cell(ri, sc)
            ref_avg = ref_avgs.get(size)
            if is_ref:
                sp_c.value = 1.0
            elif avg and avg > 0 and ref_avg:
                sp_c.value = round(ref_avg / avg, 4)
            else:
                sp_c.value = "N/A"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            sp_c.fill      = PatternFill("solid", fgColor=C["green_bg"])
            sp_c.alignment = Alignment(horizontal="right", vertical="center")
            sp_c.border    = _thin_border()

            if not is_ref:
                sp_refs.append(f"{get_column_letter(sc)}{ri}")

        avsp           = ws.cell(ri, N_SP)
        avsp.value     = (f"=IFERROR(AVERAGE({','.join(sp_refs)}),\"N/A\")"
                          if sp_refs else 1.0)
        avsp.number_format = "0.00"
        avsp.font      = Font(name=FONT_NAME, bold=True, size=10,
                              color=C["green_fg"])
        avsp.fill      = PatternFill("solid", fgColor=C["green_bg"])
        avsp.alignment = Alignment(horizontal="right", vertical="center")
        avsp.border    = _thin_border()
        ws.row_dimensions[ri].height = 17

    chart_anchor = ri + 4  # type: ignore[possibly-undefined]

    for anchor, path in [
        (f"A{chart_anchor}", ct),
        (f"L{chart_anchor}", cs),
        (f"A{chart_anchor + 22}", cb),
    ]:
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = 620
            img.height = 360
            ws.add_image(img, anchor)


def write_scaling_sheet(wb: Workbook, machine: str,
                        all_reps: AllReps, avg_data: dict,
                        sizes: list[int], thread_counts: list[int],
                        ct: str) -> None:
    """One sheet per machine showing strong scaling across thread counts."""
    ws = wb.create_sheet(f"Scaling  {machine}")
    N_COLS = 1 + len(sizes)

    title = f"Strong Scaling  —  {machine}"
    write_title_row(ws, title, N_COLS)
    ws.row_dimensions[1].height = 26

    set_col_width(ws, 1, 12)
    for ci, size in enumerate(sizes, 2):
        set_col_width(ws, ci, 14 if size <= 800 else 16)

    cur_row = 2

    ws.merge_cells(f"A{cur_row}:{get_column_letter(N_COLS)}{cur_row}")
    sec = ws.cell(cur_row, 1, value="Tabla  —  Tiempo promedio (ms) por hilos y tamaño")
    sec.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec.fill      = PatternFill("solid", fgColor=C["dark"])
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    _hdr(ws.cell(cur_row, 1), "Hilos", bg=C["dark"], size=10)
    for ci, size in enumerate(sizes, 2):
        _hdr(ws.cell(cur_row, ci), f"N = {size:,}", bg=C["mid"], size=9)
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1
    
    baseline_threads = thread_counts[0]
    ref_row_baseline = MachineRow(machine=machine, threads=baseline_threads)
    ref_avgs_baseline = avg_data.get(ref_row_baseline, {})


    for ti, t in enumerate(thread_counts):
        row = MachineRow(machine=machine, threads=t)
        bg  = C["alt"] if ti % 2 == 0 else "FFFFFF"
        times = avg_data.get(row, {})

        _dat(ws.cell(cur_row, 1), f"{t}t",
             bg=C["light"], bold=True, align="center")
        for ci, size in enumerate(sizes, 2):
            avg     = times.get(size)
            ref_avg = ref_avgs_baseline.get(size)
            if avg is not None:
                if ref_avg and avg > 0:
                    sp = ref_avg / avg
                    val = f"{avg:,.1f} ms\n{sp:.2f}x vs {baseline_threads}t"
                else:
                    val = f"{avg:,.1f} ms"
                cell = ws.cell(cur_row, ci)
                cell.value         = val
                cell.font          = Font(name=FONT_NAME, size=10)
                cell.fill          = PatternFill("solid", fgColor=bg)
                cell.alignment     = Alignment(horizontal="center",
                                               vertical="center",
                                               wrap_text=True)
                cell.border        = _thin_border()
            else:
                _dat(ws.cell(cur_row, ci), "—", bg=bg, align="center")
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

    cur_row += 2
    if ct and os.path.exists(ct):
        img        = XLImage(ct)
        img.width  = 700
        img.height = 400
        ws.add_image(img, f"A{cur_row}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    machines = discover_machines(BENCHMARKS_DIR, EXPLICIT_MACHINES)
    if not machines:
        print("No machine directories with data_omp.csv found.")
        sys.exit(1)
    print(f"Machines: {machines}")

    print("\nReading CSVs...")
    frames: list[pd.DataFrame] = []
    for m in machines:
        df = load_machine_csv(BENCHMARKS_DIR, m)
        if df is not None:
            frames.append(df)

    if not frames:
        print("No data found.")
        sys.exit(1)

    combined = pd.concat(frames, ignore_index=True)
    sizes    = sorted(combined["matrix_size"].unique().tolist())
    t_counts = sorted(combined["threads"].unique().tolist())

    all_reps: AllReps = {}
    for src in frames:
        for row, rdata in build_all_reps(src).items():
            if row not in all_reps:
                all_reps[row] = {}
            all_reps[row].update(rdata)

    avg_data = avgs(all_reps)

    print("\nGenerating charts...")
    os.makedirs(CHARTS_DIR, exist_ok=True)

    # --- Charts per thread count (cross-machine comparison) ---
    thread_charts: dict[int, tuple[str, str, str]] = {}
    for t in t_counts:
        t_rows = [MachineRow(machine=m, threads=t)
                  for m in machines
                  if MachineRow(machine=m, threads=t) in all_reps]
        if not t_rows:
            continue
        ref = slowest_row(t_rows, avg_data)
        ct = chart_time_threads(
            t_rows, avg_data, sizes, machines,
            title=f"Tiempo  |  {t} hilos  —  comparación entre máquinas",
            fname=f"time_{t}t.png",
        )
        cs = chart_speedup_machines(
            t_rows, avg_data, ref, sizes, machines,
            title=f"Speedup  |  {t} hilos  —  ref: {ref.machine}",
            fname=f"speedup_{t}t.png",
        )
        cb = chart_bar_comparison(
            t_rows, avg_data, sizes, machines,
            title=f"Comparación de tiempos  |  {t} hilos",
            fname=f"bar_{t}t.png",
        )
        thread_charts[t] = (ct, cs, cb)
        print(f"  {os.path.basename(ct)}  |  {os.path.basename(cs)}")

    # --- Scaling charts per machine ---
    scaling_charts: dict[str, str] = {}
    for m in machines:
        ct = chart_scaling(
            machine=m, all_reps=all_reps, avg_data=avg_data,
            sizes=sizes, thread_counts=t_counts,
            title=f"Strong Scaling  —  {m}",
            fname=f"scaling_{m}.png",
        )
        scaling_charts[m] = ct
        print(f"  {os.path.basename(ct)}")

    print("\nBuilding workbook...")
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)

    # --- One sheet per thread count ---
    for t in t_counts:
        t_rows = [MachineRow(machine=m, threads=t)
                  for m in machines
                  if MachineRow(machine=m, threads=t) in all_reps]
        if not t_rows:
            continue
        ref = t_rows[0]
        ct, cs, cb = thread_charts.get(t, ("", "", ""))
        cb = chart_bar_comparison(
            t_rows, avg_data, sizes, machines,
            title=f"Comparación de tiempos  |  {t} hilos",
            fname=f"bar_{t}t.png",
        )
        thread_charts[t] = (ct, cs, cb)
        write_thread_sheet(
            wb,
            sheet_name=f"{t} hilos",
            title=f"Comparación entre máquinas  |  {t} hilos  |  ref: {ref.machine}",
            rows=t_rows,
            all_reps=all_reps,
            ref=ref,
            sizes=sizes,
            machines=machines,
            ct=ct,
            cs=cs,
            cb=cb,
        )

    # --- One sheet per machine (strong scaling) ---
    for m in machines:
        m_rows = [MachineRow(machine=m, threads=t)
                  for t in t_counts
                  if MachineRow(machine=m, threads=t) in all_reps]
        if not m_rows:
            continue
        write_scaling_sheet(
            wb,
            machine=m,
            all_reps=all_reps,
            avg_data=avg_data,
            sizes=sizes,
            thread_counts=t_counts,
            ct=scaling_charts.get(m, ""),
        )

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()