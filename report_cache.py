import os
import sys

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    ColConfig, Series,
    make_border, set_col_width, style_data_cell, style_header_cell,
    write_title_row, write_raw_data_sheet, write_charts_sheet,
    save_figure, plot_lines, plot_grouped_bars, plot_simple_bars,
    cv_per_group, compute_averages,
)

CSV_PATH    = sys.argv[1] if len(sys.argv) > 1 else "results_cache/data_cache.csv"
OUT_DIR     = os.path.dirname(os.path.abspath(CSV_PATH))
OUTPUT_PATH = os.path.join(OUT_DIR, "reporte_cache.xlsx")
CHARTS_DIR  = os.path.join(OUT_DIR, "charts_cache")

IMPLS_ORDER = ["std", "cache"]

IMPL_LABELS = {
    "std":   "Standard  (i-k-j)",
    "cache": "Transposed  (row x row)",
}

IMPL_COLORS = {
    "std":   "#C00000",
    "cache": "#2E75B6",
}

IMPL_DESC = {
    "std": (
        "Loop order i-k-j. matrix2[k][j] accesses row k sequentially — "
        "already cache-friendly. matrix1[i][k] is a scalar reused across "
        "all j iterations (register). One of the better naive orderings."
    ),
    "cache": (
        "Transposes matrix2 before multiplying. Inner loop accesses "
        "matrix2T[j][k] row-by-row, matching matrix1[i][k] access pattern. "
        "Both operands advance sequentially in memory on every inner step."
    ),
}

def load_data(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns         = [c.strip().lower() for c in df.columns]
    df["impl"]         = df["impl"].str.strip('"')
    df["matrix_size"]  = df["matrix_size"].astype(int)
    df["repetition"]   = df["repetition"].astype(int)
    df["wall_time_ms"] = df["wall_time_ms"].astype(float)
    return df

def chart_time_grouped(avgs: dict, sizes: list, impls: list) -> str:
    series = [
        Series(label=IMPL_LABELS[i], data=avgs[i], color=IMPL_COLORS[i])
        for i in impls
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 6))
        plot_grouped_bars(ax, sizes, series, label_values=True)
        ax.set_xticklabels([f"N={s}" for s in sizes], fontsize=11)
        ax.set_title("Average Wall Time by Implementation and Matrix Size",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_ylabel("Avg wall time (ms)", fontsize=11)
        ax.legend(fontsize=10)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "time_grouped.png")


def chart_time_line(avgs: dict, sizes: list, impls: list) -> str:
    series = [
        Series(label=IMPL_LABELS[i], data=avgs[i], color=IMPL_COLORS[i])
        for i in impls
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(10, 6))
        plot_lines(ax, series, log_scale=True)
        ax.set_title("Execution Time vs Matrix Dimension  (log scale)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Matrix dimension N", fontsize=11)
        ax.set_ylabel("Avg wall time (ms)", fontsize=11)
        ax.legend(fontsize=10)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "time_line.png")


def chart_speedup(avgs: dict, sizes: list, std_avgs: dict) -> str:
    speedups = [std_avgs.get(s, 1) / avgs["cache"].get(s, 1) for s in sizes]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(8, 5))
        plot_simple_bars(ax, [f"N={s}" for s in sizes], speedups,
                         color=IMPL_COLORS["cache"])
        ax.axhline(1, color=IMPL_COLORS["std"], lw=1.5, ls="--",
                   label="std baseline")
        ax.set_title("Speedup: Transposed vs Standard  (T_std / T_cache)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_ylabel("Speedup factor", fontsize=11)
        ax.set_ylim(0, max(speedups) * 1.2)
        ax.legend(fontsize=10)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "speedup.png")


def chart_cv(df: pd.DataFrame, sizes: list, impls: list) -> str:
    series = [
        Series(
            label=IMPL_LABELS[i],
            data={s: cv for s, cv in zip(
                sizes, cv_per_group(df, "impl", "matrix_size", "wall_time_ms", i, sizes)
            )},
            color=IMPL_COLORS[i],
        )
        for i in impls
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        plot_grouped_bars(ax, sizes, series)
        ax.set_xticklabels([f"N={s}" for s in sizes], fontsize=11)
        ax.set_title("Coefficient of Variation  (lower = more stable)",
                     fontsize=12, fontweight="bold", pad=10)
        ax.set_ylabel("CV (%)", fontsize=11)
        ax.legend(fontsize=10)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "cv.png")

def write_averages(wb: Workbook, sizes: list, impls: list) -> None:
    ws = wb.create_sheet("Averages")
    ws.freeze_panes = "B3"

    n_cols = 1 + len(sizes) * 2 + 1
    write_title_row(ws, "Average Wall Time & Speedup  |  Speedup = T(std) / T(impl)", n_cols)

    style_header_cell(ws.cell(2, 1), "Implementation", bg="dark")
    set_col_width(ws, 1, 22)
    ws.row_dimensions[2].height = 32

    for si, size in enumerate(sizes):
        ac, sc = 2 + si * 2, 3 + si * 2
        style_header_cell(ws.cell(2, ac), f"N={size}\nAvg (ms)", bg="mid")
        style_header_cell(ws.cell(2, sc), f"N={size}\nSpeedup",  bg="mid")
        set_col_width(ws, ac, 14)
        set_col_width(ws, sc, 12)

    avg_sp_col = 2 + len(sizes) * 2
    style_header_cell(ws.cell(2, avg_sp_col), "Avg\nSpeedup", bg="dark")
    set_col_width(ws, avg_sp_col, 12)

    for ri, impl in enumerate(impls, 3):
        bg = "alt" if ri % 2 == 0 else "white"
        style_data_cell(ws.cell(ri, 1), IMPL_LABELS[impl],
                        bg="light", bold=True, align="left")

        sp_refs = []
        for si, size in enumerate(sizes):
            ac = 2 + si * 2
            sc = ac + 1
            al = get_column_letter(ac)

            avg_c = ws.cell(ri, ac)
            avg_c.value = (
                f"=IFERROR(AVERAGEIFS('Raw Data'!D:D,"
                f"'Raw Data'!A:A,\"{impl}\","
                f"'Raw Data'!B:B,{size}),\"N/A\")"
            )
            avg_c.number_format = "0.000"
            avg_c.font      = Font(name=FONT_NAME, size=10)
            avg_c.fill      = PatternFill("solid", fgColor=C[bg])
            avg_c.alignment = Alignment(horizontal="right")
            avg_c.border    = make_border()

            sp_c       = ws.cell(ri, sc)
            sp_c.value = 1.0 if impl == "std" else f"=IFERROR(${al}$3/{al}{ri},\"N/A\")"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            sp_c.fill      = PatternFill("solid", fgColor=C["green_bg"])
            sp_c.alignment = Alignment(horizontal="right")
            sp_c.border    = make_border()
            sp_refs.append(f"{get_column_letter(sc)}{ri}")

        avg_sp = ws.cell(ri, avg_sp_col)
        avg_sp.value         = f"=IFERROR(AVERAGE({','.join(sp_refs)}),\"N/A\")"
        avg_sp.number_format = "0.00"
        avg_sp.font      = Font(name=FONT_NAME, bold=True, size=10, color=C["green_fg"])
        avg_sp.fill      = PatternFill("solid", fgColor=C["green_bg"])
        avg_sp.alignment = Alignment(horizontal="right")
        avg_sp.border    = make_border()

def write_analysis(wb: Workbook, avgs: dict, sizes: list,
                   df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Analysis")
    write_title_row(ws, "Key Findings  |  Cache Line Optimization Benchmark", 4)

    for col, w in enumerate([28, 16, 16, 58], 1):
        set_col_width(ws, col, w)

    row = [3]

    def section(title: str) -> None:
        ws.merge_cells(f"A{row[0]}:D{row[0]}")
        c = ws.cell(row[0], 1, value=title)
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color=C["white"])
        c.fill      = PatternFill("solid", fgColor=C["mid"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row[0]].height = 20
        row[0] += 1

    def finding(label: str, value: str, note: str = "") -> None:
        bg = "alt" if row[0] % 2 == 0 else "white"
        style_data_cell(ws.cell(row[0], 1), label, bg=bg, align="left", bold=True)
        style_data_cell(ws.cell(row[0], 2), value, bg=bg, align="center")
        style_data_cell(ws.cell(row[0], 3), "",    bg=bg)
        style_data_cell(ws.cell(row[0], 4), note,  bg=bg, align="left")
        row[0] += 1

    std_avgs = avgs["std"]

    section("1. Speedup por tamano  (T_std / T_cache)")
    for s in sizes:
        if s not in avgs["cache"] or s not in std_avgs:
            continue
        sp        = std_avgs[s] / avgs["cache"][s]
        direction = "cache mas rapido" if sp > 1.0 else "std igual o mas rapido"
        finding(f"N = {s}", f"{sp:.3f}x", direction)

    row[0] += 1
    section("2. Comportamiento de cache por implementacion")
    for impl in IMPLS_ORDER:
        finding(IMPL_LABELS[impl], "", IMPL_DESC[impl])

    row[0] += 1
    section("3. Estabilidad de mediciones  (Coeficiente de Variacion)")
    for impl in IMPLS_ORDER:
        for s in sizes:
            vals  = df[(df["impl"] == impl) & (df["matrix_size"] == s)]["wall_time_ms"]
            cv    = vals.std() / vals.mean() * 100 if len(vals) > 1 else 0
            label = "Excelente (<1%)" if cv < 1 else ("Buena (<3%)" if cv < 3 else "Revisar")
            finding(f"{impl}  N={s}", f"CV = {cv:.2f}%", label)

    row[0] += 1
    section("4. Tendencia con el tamano de la matriz")
    prev_sp = None
    for s in sizes:
        if s not in avgs["cache"] or s not in std_avgs:
            continue
        sp   = std_avgs[s] / avgs["cache"][s]
        note = ""
        if prev_sp is not None:
            if sp > prev_sp + 0.05:
                note = "Speedup crece: presion de cache aumenta con N"
            elif sp < prev_sp - 0.05:
                note = "Speedup decrece: ambas implementaciones sufren el LLC miss"
            else:
                note = "Speedup estable: patron de acceso consistente en este rango"
        finding(f"N = {s}", f"{sp:.3f}x", note)
        prev_sp = sp

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    print(f"Reading: {CSV_PATH}")
    df    = load_data(CSV_PATH)
    impls = [i for i in IMPLS_ORDER if i in df["impl"].unique()]
    sizes = sorted(df["matrix_size"].unique().tolist())
    avgs  = compute_averages(df, "impl", "matrix_size", "wall_time_ms", impls, sizes)

    print("Generating charts...")
    chart_paths = [
        chart_time_grouped(avgs, sizes, impls),
        chart_speedup(avgs, sizes, avgs.get("std", {})),
        chart_time_line(avgs, sizes, impls),
        chart_cv(df, sizes, impls),
    ]
    for p in chart_paths:
        print(f"  {os.path.basename(p)}")

    print("Building workbook...")
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    raw_cols = [
        ColConfig("Implementation", 22, "impl",        align="left"),
        ColConfig("Matrix Size",    14, "matrix_size", fmt="#,##0"),
        ColConfig("Repetition",     12, "repetition",  fmt="0"),
        ColConfig("Wall Time (ms)", 16, "wall_time_ms",fmt="0.000"),
    ]
    write_raw_data_sheet(wb, df,
                         "Raw Measurements  |  Cache Line Optimization Benchmark",
                         raw_cols)
    write_averages(wb, sizes, impls)
    write_analysis(wb, avgs, sizes, df)
    write_charts_sheet(wb, "Visual Analysis  |  Cache Line Optimization Benchmark",
                       chart_paths)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()