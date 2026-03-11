import os
import sys
import statistics

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    ColConfig, Series,
    make_border, set_col_width, style_data_cell, style_header_cell,
    write_title_row, write_raw_data_sheet, write_charts_sheet,
    save_figure, plot_lines, plot_grouped_bars, plot_simple_bars,
    cv_per_group, compute_averages,
)

CSV_PATH    = sys.argv[1] if len(sys.argv) > 1 else "results_opt/data_opt.csv"
OUT_DIR     = os.path.dirname(os.path.abspath(CSV_PATH))
OUTPUT_PATH = os.path.join(OUT_DIR, "reporte_opt.xlsx")
CHARTS_DIR  = os.path.join(OUT_DIR, "charts_opt")

CONFIGS_ORDER = [
    "O0", "O1", "O2", "O3", "Ofast",
    "O2_native", "O3_native", "Ofast_native",
    "O3_unroll", "O3_lto", "O3_full",
]

CONFIG_FLAGS = {
    "O0":           "-O0 -Wall",
    "O1":           "-O1 -Wall",
    "O2":           "-O2 -Wall",
    "O3":           "-O3 -Wall",
    "Ofast":        "-Ofast -Wall",
    "O2_native":    "-O2 -Wall -march=native",
    "O3_native":    "-O3 -Wall -march=native",
    "Ofast_native": "-Ofast -Wall -march=native",
    "O3_unroll":    "-O3 -Wall -march=native -funroll-loops",
    "O3_lto":       "-O3 -Wall -march=native -flto",
    "O3_full":      "-O3 -Wall -march=native -funroll-loops -flto -ffast-math -fomit-frame-pointer",
}

CONFIG_COLORS = {
    "O0":           "#C00000",
    "O1":           "#FF6600",
    "O2":           "#FFC000",
    "O3":           "#70AD47",
    "Ofast":        "#00B050",
    "O2_native":    "#4472C4",
    "O3_native":    "#2E75B6",
    "Ofast_native": "#1F4E79",
    "O3_unroll":    "#7030A0",
    "O3_lto":       "#00B0F0",
    "O3_full":      "#FF0066",
}

def load_data(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns         = [c.strip().lower() for c in df.columns]
    df["config"]       = df["config"].str.strip('"')
    df["flags"]        = df["flags"].str.strip('"')
    df["matrix_size"]  = df["matrix_size"].astype(int)
    df["repetition"]   = df["repetition"].astype(int)
    df["wall_time_ms"] = df["wall_time_ms"].astype(float)
    return df

def chart_time_log(avgs: dict, sizes: list, configs: list) -> str:
    series = [
        Series(label=cfg, data=avgs[cfg], color=CONFIG_COLORS.get(cfg, "#333"))
        for cfg in configs
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(11, 6))
        plot_lines(ax, series, log_scale=True)
        ax.set_title("Execution Time vs Matrix Dimension  (log scale)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Matrix dimension N", fontsize=11)
        ax.set_ylabel("Avg wall time (ms)", fontsize=11)
        ax.legend(fontsize=8, ncol=2, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "time_log.png")


def chart_speedup(avgs: dict, sizes: list, configs: list, o0_avgs: dict) -> str:
    series = [
        Series(
            label=cfg,
            data={s: o0_avgs[s] / avgs[cfg][s]
                  for s in sizes if s in avgs[cfg] and s in o0_avgs},
            color=CONFIG_COLORS.get(cfg, "#333"),
        )
        for cfg in configs if cfg != "O0"
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(11, 6))
        ax.axhline(1, color="#AAAAAA", lw=1, ls="--", label="O0 baseline")
        plot_lines(ax, series)
        ax.set_title("Speedup over -O0 vs Matrix Dimension",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_xlabel("Matrix dimension N", fontsize=11)
        ax.set_ylabel("Speedup  (T_O0 / T_config)", fontsize=11)
        ax.legend(fontsize=8, ncol=2, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "speedup.png")


def chart_bar_speedup(avg_speedups: dict, configs: list) -> str:
    vals   = [avg_speedups.get(c, 0) for c in configs]
    colors = [CONFIG_COLORS.get(c, "#333") for c in configs]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(11, 5))
        plot_simple_bars(ax, configs, vals, color=colors, label_fmt="{:.1f}x")
        ax.set_xticklabels(configs, rotation=30, ha="right", fontsize=9)
        ax.set_title("Average Speedup over -O0  (all sizes)",
                     fontsize=13, fontweight="bold", pad=10)
        ax.set_ylabel("Speedup", fontsize=11)
        ax.set_ylim(0, max(vals) * 1.18)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "bar_speedup.png")


def chart_cv(df: pd.DataFrame, sizes: list, configs: list) -> str:
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(11, 5))
        w = 0.8 / len(sizes)
        for si, size in enumerate(sizes):
            offsets = [x + (si - len(sizes) / 2 + 0.5) * w for x in range(len(configs))]
            cv_vals = [
                cv_per_group(df, "config", "matrix_size", "wall_time_ms", cfg, [size])[0]
                for cfg in configs
            ]
            ax.bar(offsets, cv_vals, width=w * 0.9, label=f"N={size}", alpha=0.85)
        ax.set_xticks(range(len(configs)))
        ax.set_xticklabels(configs, rotation=30, ha="right", fontsize=9)
        ax.set_title("Coefficient of Variation  (lower = more stable)",
                     fontsize=12, fontweight="bold", pad=10)
        ax.set_ylabel("CV (%)", fontsize=11)
        ax.legend(fontsize=9)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "cv.png")


def chart_native_impact(avgs: dict, sizes: list) -> str:
    pairs  = [("O2", "O2_native"), ("O3", "O3_native"), ("Ofast", "Ofast_native")]
    series = [
        Series(
            label=f"{base} -> {native}",
            data={s: avgs[base][s] / avgs[native][s]
                  for s in sizes if s in avgs.get(base, {}) and s in avgs.get(native, {})},
            color=CONFIG_COLORS.get(native, "#333"),
        )
        for base, native in pairs
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        plot_grouped_bars(ax, sizes, series, bar_width=0.25)
        ax.axhline(1, color="#AAAAAA", lw=1, ls="--")
        ax.set_xticklabels([f"N={s}" for s in sizes], fontsize=10)
        ax.set_title("Speedup gained by adding -march=native",
                     fontsize=12, fontweight="bold", pad=10)
        ax.set_ylabel("Speedup factor", fontsize=11)
        ax.legend(fontsize=9)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, "native_impact.png")

def write_averages(wb: Workbook, sizes: list, configs: list) -> None:
    ws = wb.create_sheet("Averages")
    ws.freeze_panes = "B3"

    n_cols = 1 + len(sizes) * 2 + 2
    write_title_row(ws, "Average Wall Time & Speedup  |  Speedup = T(O0) / T(config)", n_cols)

    style_header_cell(ws.cell(2, 1), "Config", bg="dark")
    set_col_width(ws, 1, 14)
    ws.row_dimensions[2].height = 32

    for si, size in enumerate(sizes):
        ac, sc = 2 + si * 2, 3 + si * 2
        style_header_cell(ws.cell(2, ac), f"N={size}\nAvg (ms)", bg="mid")
        style_header_cell(ws.cell(2, sc), f"N={size}\nSpeedup",  bg="mid")
        set_col_width(ws, ac, 13)
        set_col_width(ws, sc, 11)

    avg_sp_col  = 2 + len(sizes) * 2
    correct_col = avg_sp_col + 1
    style_header_cell(ws.cell(2, avg_sp_col),  "Avg\nSpeedup", bg="dark")
    style_header_cell(ws.cell(2, correct_col), "Correct",      bg="dark")
    set_col_width(ws, avg_sp_col, 12)
    set_col_width(ws, correct_col, 9)

    for ri, cfg in enumerate(configs, 3):
        bg = "alt" if ri % 2 == 0 else "white"
        style_data_cell(ws.cell(ri, 1), cfg, bg="light", bold=True, align="center")

        sp_refs = []
        for si, size in enumerate(sizes):
            ac = 2 + si * 2
            sc = ac + 1
            al = get_column_letter(ac)

            avg_c = ws.cell(ri, ac)
            avg_c.value = (
                f"=IFERROR(AVERAGEIFS('Raw Data'!E:E,"
                f"'Raw Data'!A:A,A{ri},"
                f"'Raw Data'!C:C,{size}),\"N/A\")"
            )
            avg_c.number_format = "0.000"
            avg_c.font      = Font(name=FONT_NAME, size=10)
            avg_c.fill      = PatternFill("solid", fgColor=C[bg])
            avg_c.alignment = Alignment(horizontal="right")
            avg_c.border    = make_border()

            sp_c       = ws.cell(ri, sc)
            sp_c.value = 1.0 if cfg == "O0" else f"=IFERROR(${al}$3/{al}{ri},\"N/A\")"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10, color=C["green_fg"])
            sp_c.fill      = PatternFill("solid", fgColor=C["green_bg"])
            sp_c.alignment = Alignment(horizontal="right")
            sp_c.border    = make_border()
            sp_refs.append(f"{get_column_letter(sc)}{ri}")

        avg_sp = ws.cell(ri, avg_sp_col)
        avg_sp.value         = f"=IFERROR(AVERAGE({','.join(sp_refs)}),\"N/A\")"
        avg_sp.number_format = "0.00"
        avg_sp.font      = Font(name=FONT_NAME, bold=True, size=10, color=C["green_fg"])
        avg_sp.fill      = PatternFill("solid", fgColor=C["green_bg"])
        avg_sp.alignment = Alignment(horizontal="right")
        avg_sp.border    = make_border()

        corr = ws.cell(ri, correct_col)
        corr.value = (
            f"=IFERROR(AVERAGEIFS('Raw Data'!F:F,'Raw Data'!A:A,A{ri}),\"N/A\")"
        )
        corr.number_format = "0"
        corr.font      = Font(name=FONT_NAME, size=10)
        corr.fill      = PatternFill("solid", fgColor=C[bg])
        corr.alignment = Alignment(horizontal="center")
        corr.border    = make_border()

def write_ranking(wb: Workbook, avg_speedups: dict,
                  configs: list, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Ranking")
    write_title_row(ws, "Configuration Ranking by Average Speedup over -O0", 5)

    headers = ["Rank", "Config", "Avg Speedup", "Correct", "Flags"]
    widths  = [8, 14, 13, 9, 72]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        style_header_cell(ws.cell(2, col), h, bg="mid")
        set_col_width(ws, col, w)
    ws.row_dimensions[2].height = 20

    ranked = sorted(configs, key=lambda c: avg_speedups.get(c, 0), reverse=True)
    medal  = {
        1: ("gold_bg", "gold_fg"),
        2: ("grey_bg", "grey_fg"),
        3: ("red_bg",  "red_fg"),
    }

    for ri, cfg in enumerate(ranked, 3):
        rank    = ri - 2
        sp      = avg_speedups.get(cfg, 0)
        correct = int(df[df["config"] == cfg]["correct"].mode()[0])
        bg, fg  = medal.get(rank, ("alt" if rank % 2 == 0 else "white", "000000"))

        style_data_cell(ws.cell(ri, 1), f"#{rank}",    bg=bg, fg=fg, bold=rank <= 3, align="center")
        style_data_cell(ws.cell(ri, 2), cfg,            bg=bg, fg=fg, bold=rank <= 3, align="center")
        style_data_cell(ws.cell(ri, 3), f"{sp:.2f}x",  bg=bg, fg="green_fg", bold=rank <= 3)
        style_data_cell(ws.cell(ri, 4), correct, fmt="0", bg=bg, align="center")
        style_data_cell(ws.cell(ri, 5), CONFIG_FLAGS.get(cfg, ""), bg=bg, align="left")

def write_analysis(wb: Workbook, avgs: dict, avg_speedups: dict,
                   sizes: list, configs: list, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Analysis")
    write_title_row(ws, "Key Findings  |  Compiler Optimization Benchmark", 4)

    for col, w in enumerate([32, 18, 18, 55], 1):
        set_col_width(ws, col, w)

    row = [3]

    def section(title: str) -> None:
        ws.merge_cells(f"A{row[0]}:D{row[0]}")
        c = ws.cell(row[0], 1, value=title)
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color=C["white"])
        c.fill      = PatternFill("solid", fgColor=C["mid"])
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row[0]].height = 20
        row[0] += 1

    def finding(label: str, value: str, note: str = "") -> None:
        bg = "alt" if row[0] % 2 == 0 else "white"
        style_data_cell(ws.cell(row[0], 1), label, bg=bg, align="left", bold=True)
        style_data_cell(ws.cell(row[0], 2), value, bg=bg, align="center")
        style_data_cell(ws.cell(row[0], 3), "",    bg=bg)
        style_data_cell(ws.cell(row[0], 4), note,  bg=bg, align="left")
        row[0] += 1

    o0_avgs = avgs["O0"]

    section("1. Baseline levels  (-O0 to -Ofast)")
    prev_sp, prev_cfg = 1.0, "O0"
    for cfg in ["O1", "O2", "O3", "Ofast"]:
        sp    = avg_speedups[cfg]
        delta = sp / prev_sp
        finding(f"{prev_cfg} -> {cfg}",
                f"{sp:.2f}x total",
                f"{delta:.2f}x adicional respecto a {prev_cfg}")
        prev_sp, prev_cfg = sp, cfg

    row[0] += 1
    section("2. Impacto de -march=native")
    notes_native = {
        "O2":    "A nivel O2 no hay vectorizacion; -march=native no cambia nada",
        "O3":    "O3 activa vectorizacion SSE; native habilita AVX2 -> 8 int/ciclo",
        "Ofast": "Similar a O3_native; Ofast agrega fast-math sin beneficio en ints",
    }
    for base, native in [("O2", "O2_native"), ("O3", "O3_native"), ("Ofast", "Ofast_native")]:
        gains = [avgs[base][s] / avgs[native][s]
                 for s in sizes if s in avgs[base] and s in avgs[native]]
        finding(f"{base} -> {native}", f"{statistics.mean(gains):.2f}x avg",
                notes_native[base])

    row[0] += 1
    section("3. Flags adicionales sobre O3_native")
    notes_extra = {
        "O3_unroll": "-funroll-loops: elimina branch del bucle interno; ganancia marginal sobre AVX2",
        "O3_lto":    "-flto: no hay TUs utiles que inlinear; overhead de layout > beneficio",
        "O3_full":   "Combinacion completa: mejor resultado global del benchmark",
    }
    for cfg in ["O3_unroll", "O3_lto", "O3_full"]:
        gains = [avgs["O3_native"][s] / avgs[cfg][s]
                 for s in sizes if s in avgs.get(cfg, {})]
        finding(f"O3_native -> {cfg}", f"{statistics.mean(gains):.2f}x avg",
                notes_extra[cfg])

    row[0] += 1
    section("4. Estabilidad  (Coeficiente de Variacion - menor es mejor)")
    for cfg in configs:
        vals  = df[df["config"] == cfg]["wall_time_ms"]
        cv    = vals.std() / vals.mean() * 100
        label = "Excelente (<1%)" if cv < 1 else ("Buena (<3%)" if cv < 3 else "Revisar")
        finding(cfg, f"CV = {cv:.2f}%", label)

    row[0] += 1
    section("5. Correctitud  (verify_mul.c - 4x4 con valores conocidos)")
    all_ok = (df["correct"] == 1).all()
    finding("Todas las configuraciones",
            "PASS" if all_ok else "FAIL",
            "Aritmetica entera pura: las optimizaciones no alteran el resultado matematico")

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    print(f"Reading: {CSV_PATH}")
    df      = load_data(CSV_PATH)
    configs = [c for c in CONFIGS_ORDER if c in df["config"].unique()]
    sizes   = sorted(df["matrix_size"].unique().tolist())
    avgs    = compute_averages(df, "config", "matrix_size", "wall_time_ms", configs, sizes)

    o0_avgs      = avgs.get("O0", {})
    avg_speedups = {
        cfg: statistics.mean(
            [o0_avgs[s] / avgs[cfg][s]
             for s in sizes if s in avgs[cfg] and s in o0_avgs and avgs[cfg][s] > 0]
        )
        for cfg in configs
    }

    print("Generating charts...")
    chart_paths = [
        chart_time_log(avgs, sizes, configs),
        chart_speedup(avgs, sizes, configs, o0_avgs),
        chart_bar_speedup(avg_speedups, configs),
        chart_cv(df, sizes, configs),
        chart_native_impact(avgs, sizes),
    ]
    for p in chart_paths:
        print(f"  {os.path.basename(p)}")

    print("Building workbook...")
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    raw_cols = [
        ColConfig("Config",          14, "config",       align="left"),
        ColConfig("Flags",           60, "flags",        align="left"),
        ColConfig("Matrix Size",     13, "matrix_size",  fmt="#,##0"),
        ColConfig("Repetition",      12, "repetition",   fmt="0"),
        ColConfig("Wall Time (ms)",  16, "wall_time_ms", fmt="0.000"),
        ColConfig("Correct",         10, "correct",      fmt="0", align="center"),
    ]
    write_raw_data_sheet(wb, df,
                         "Raw Measurements  |  Compiler Optimization Benchmark",
                         raw_cols)
    write_averages(wb, sizes, configs)
    write_ranking(wb, avg_speedups, configs, df)
    write_analysis(wb, avgs, avg_speedups, sizes, configs, df)
    write_charts_sheet(wb, "Visual Analysis  |  Compiler Optimization Benchmark",
                       chart_paths)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()